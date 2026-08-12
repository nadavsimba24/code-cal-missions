"""ייצוא הלוח לאקסל — POST /api/boards/{id}/export/xlsx.

The client ships the grid it is already showing (filters, sort, column
structure, computed values) and the server turns it into a real .xlsx: one
sheet, RTL, frozen header, sub-items nested and collapsible under their item.
"""
import io

import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import load_workbook  # noqa: E402


HEADERS = ["פריט", "סטטוס", "תאריך יעד", "תקציב"]


def _grid(rows=None):
    return {"columns": HEADERS, "rows": rows if rows is not None else [
        {"group": True, "level": 0, "cells": ["בביצוע", "", "", ""]},
        {"level": 1, "cells": ["גשר הרצל", {"v": "בתהליך", "c": "#fdab3d"},
                               {"v": "2026-09-01", "t": "date"}, 125000]},
        {"level": 2, "cells": ["הכנת מכרז", {"v": "הושלם", "c": "#00c875"}, "", 4000]},
    ]}


def _export(client, bid, uid, grid=None):
    return client.post(f"/api/boards/{bid}/export/xlsx",
                       json={"user_id": uid, **(grid or _grid())})


def _sheet(resp):
    return load_workbook(io.BytesIO(resp.content)).active


def test_export_returns_a_real_xlsx(client, admin_id):
    r = _export(client, 1, admin_id)
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "attachment" in r.headers["content-disposition"]
    assert r.content[:2] == b"PK"                      # a real zip/xlsx container


def test_headers_values_and_hierarchy(client, admin_id):
    ws = _sheet(_export(client, 1, admin_id))
    assert [c.value for c in ws[1]] == HEADERS
    assert ws.cell(row=2, column=1).value == "בביצוע"          # group header
    assert ws.cell(row=3, column=1).value == "גשר הרצל"        # item
    assert ws.cell(row=4, column=1).value == "הכנת מכרז"       # sub-item
    # sub-items collapse under their item, items under their group
    assert ws.row_dimensions[3].outlineLevel == 1
    assert ws.row_dimensions[4].outlineLevel == 2
    assert ws.sheet_properties.outlinePr.summaryBelow is False


def test_types_are_preserved_for_excel(client, admin_id):
    ws = _sheet(_export(client, 1, admin_id))
    budget = ws.cell(row=3, column=4)
    assert budget.value == 125000 and budget.data_type == "n"   # a number, not text
    due = ws.cell(row=3, column=3)
    assert due.value.strftime("%Y-%m-%d") == "2026-09-01"       # a real date
    assert "DD/MM/YYYY" in due.number_format


def test_status_color_and_rtl_and_freeze(client, admin_id):
    ws = _sheet(_export(client, 1, admin_id))
    assert ws.cell(row=3, column=2).fill.fgColor.rgb == "FFFDAB3D"   # opaque, not alpha-00
    assert ws.sheet_view.rightToLeft is True
    assert ws.freeze_panes == "A2"


def test_text_starting_with_equals_is_never_a_formula(client, admin_id):
    """Spreadsheet-injection guard: an item titled "=1+1" stays text."""
    ws = _sheet(_export(client, 1, admin_id, _grid([
        {"level": 1, "cells": ["=1+1", "@SUM(A1)", "-2+3", "+9"]}])))
    row = [ws.cell(row=2, column=i) for i in range(1, 5)]
    assert [c.value for c in row] == ["=1+1", "@SUM(A1)", "-2+3", "+9"]
    assert {c.data_type for c in row} == {"s"}


def test_long_text_is_capped_and_extra_columns_dropped(client, admin_id):
    ws = _sheet(_export(client, 1, admin_id, _grid([
        {"level": 1, "cells": ["x" * 9000, "", "", "", "עודף", "עודף"]}])))
    assert len(ws.cell(row=2, column=1).value) == 4000
    assert ws.max_column == len(HEADERS)


def test_empty_columns_are_rejected(client, admin_id):
    r = client.post("/api/boards/1/export/xlsx",
                    json={"user_id": admin_id, "columns": [], "rows": []})
    assert r.status_code == 400


def test_missing_board_is_404(client, admin_id):
    assert _export(client, 999999, admin_id).status_code == 404


def test_non_member_cannot_export(client, admin_id, guinea_id):
    bid = client.post("/api/boards", json={"name": "לוח פרטי", "user_id": admin_id}).json()["id"]
    assert _export(client, bid, guinea_id).status_code == 403
    assert _export(client, bid, admin_id).status_code == 200
