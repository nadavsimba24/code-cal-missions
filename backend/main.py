"""
CityOS — FastAPI Backend Server
"""
import os, sys, json, uuid, csv, io, re, secrets, urllib.parse
from datetime import datetime, timezone
from typing import Optional
from fastapi import Depends, FastAPI, HTTPException, Query, UploadFile, File, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, Response
import os
from pydantic import BaseModel
from sqlalchemy import create_engine, func, or_
from sqlalchemy.orm import Session

sys.path.insert(0, os.path.dirname(__file__))
from auth import (auth_mode, current_user, current_user_id, effective_auth_mode,
                  init_auth, resolve_user)
from models import (
    Organization, Department, Environment, EnvironmentMember, Folder, User, Board, Group, Task, Comment, BoardMember, WorkspaceMember, RolePermission,
    Permit, CitizenRequest, PublicTransportStop, InfrastructureAsset,
    TaskStatus, Priority, BoardType, init_db,
    AnnualWorkPlan, Project, ProjectStep, BudgetLineItem,
    Approval, ChangeRequest, KPI, Dependency, Document, AuditLog,
    task_assignees, task_watchers,
    ProjectStatus, ApprovalStatus, ChangeRequestStatus,
    DependencyType, BudgetItemType, DocumentType, LoginEvent, UploadedFile, Notification
)

try:
    from dotenv import load_dotenv
    # Load .env from cityos/ root (parent of backend/)
    load_dotenv(os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env"))
except Exception:
    pass

app = FastAPI(title="CODE-CAL MISSIONS", version="0.1.0")
# The SPA is served from the same origin as the API, so it needs no CORS grant at
# all. Extra origins (a staging frontend, a local vite server) go in
# CITYOS_CORS_ORIGINS as a comma-separated list. This was allow_origins=["*"].
_cors_origins = [o.strip() for o in (os.environ.get("CITYOS_CORS_ORIGINS") or "").split(",") if o.strip()]
if _cors_origins:
    app.add_middleware(CORSMiddleware, allow_origins=_cors_origins,
                       allow_credentials=True,
                       allow_methods=["GET", "POST", "PATCH", "PUT", "DELETE", "OPTIONS"],
                       allow_headers=["Content-Type", "X-CityOS-User"])

# Endpoints that may be reached without an identity. Everything else under /api/
# requires one — so a route added tomorrow is protected by default instead of
# being protected only if its author remembered the dependency.
PUBLIC_API_PATHS = {"/api/status"}


@app.middleware("http")
async def require_authentication(request, call_next):
    path = request.url.path
    if (path.startswith("/api/")
            and path not in PUBLIC_API_PATHS
            and request.method != "OPTIONS"):
        try:
            with Session(engine) as db:
                resolve_user(db, request)
        except HTTPException as exc:
            return JSONResponse({"detail": exc.detail}, status_code=exc.status_code)
    return await call_next(request)


@app.middleware("http")
async def security_headers(request, call_next):
    resp = await call_next(request)
    if "text/html" in resp.headers.get("content-type", ""):
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("X-Frame-Options", "DENY")
    resp.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    return resp

# Persistent Postgres (production) via DATABASE_URL; SQLite otherwise (local).
DATABASE_URL = os.environ.get("DATABASE_URL")
if DATABASE_URL:
    if DATABASE_URL.startswith("postgres://"):  # SQLAlchemy needs postgresql://
        DATABASE_URL = "postgresql://" + DATABASE_URL[len("postgres://"):]
    engine = init_db(DATABASE_URL)
    IS_SQLITE = False
else:
    DB_PATH = os.environ.get("CITYOS_DB_PATH", os.path.join(os.path.dirname(__file__), "cityos.db"))
    engine = init_db(f"sqlite:///{DB_PATH}")
    IS_SQLITE = True

# Identity comes from the transport (Entra ID via Azure Easy Auth), never from a
# request parameter. See backend/auth.py for the two modes and why the default
# fails closed.
init_auth(engine)
if auth_mode() == "dev":
    print("[auth] DEV MODE — identity is taken from the X-CityOS-User header. "
          "Never set CITYOS_AUTH_MODE=dev in production.")
elif auth_mode() == "entra":
    # The SSO routes live at /auth/* — outside /api/, so the middleware above
    # lets them through unauthenticated, which is the whole point of a login
    # page. They are registered here, well before the StaticFiles mount at "/",
    # because the first matching route wins.
    import entra
    app.include_router(entra.router)
    if entra.is_configured():
        print(f"[auth] ENTRA SSO — tenant {entra.tenant_id()}, client {entra.client_id()}")
    else:
        print("[auth] ENTRA SSO selected but not configured — set ENTRA_TENANT_ID, "
              "ENTRA_CLIENT_ID and ENTRA_CLIENT_SECRET. Every request will be refused.")

def _migrate():
    """Lightweight additive migrations for SQLite (create_all won't ALTER)."""
    from sqlalchemy import text
    with engine.begin() as conn:
        cols = {r[1] for r in conn.execute(text("PRAGMA table_info(comments)"))}
        for name, ddl in [("parent_id", "INTEGER"), ("mentions", "JSON"), ("likes", "JSON"), ("seen_by", "JSON")]:
            if name not in cols:
                conn.execute(text(f"ALTER TABLE comments ADD COLUMN {name} {ddl}"))
        tcols = {r[1] for r in conn.execute(text("PRAGMA table_info(tasks)"))}
        if "permissions" not in tcols:
            conn.execute(text("ALTER TABLE tasks ADD COLUMN permissions JSON"))
        # "מזהה פריט" — automatic 11-digit item identifier (+ the board it belongs to)
        if "item_uid" not in tcols:
            conn.execute(text("ALTER TABLE tasks ADD COLUMN item_uid VARCHAR(11)"))
        if "item_uid_board" not in tcols:
            conn.execute(text("ALTER TABLE tasks ADD COLUMN item_uid_board INTEGER"))
        # boards: environment membership (Monday-style environments/workspaces)
        bcols = {r[1] for r in conn.execute(text("PRAGMA table_info(boards)"))}
        if "environment_id" not in bcols:
            conn.execute(text("ALTER TABLE boards ADD COLUMN environment_id INTEGER"))
        if "folder_id" not in bcols:
            conn.execute(text("ALTER TABLE boards ADD COLUMN folder_id INTEGER"))
        if "position" not in bcols:
            conn.execute(text("ALTER TABLE boards ADD COLUMN position INTEGER DEFAULT 0"))
        # environments: primary-workspace flag (may predate the column)
        ecols = {r[1] for r in conn.execute(text("PRAGMA table_info(environments)"))}
        if ecols and "is_primary" not in ecols:
            conn.execute(text("ALTER TABLE environments ADD COLUMN is_primary BOOLEAN DEFAULT 0"))
        # environment members: per-environment role (manager | member)
        emcols = {r[1] for r in conn.execute(text("PRAGMA table_info(environment_members)"))}
        if emcols and "role" not in emcols:
            conn.execute(text("ALTER TABLE environment_members ADD COLUMN role VARCHAR(20) DEFAULT 'member'"))
        # notifications table may predate the task_id column (created before mentions)
        ncols = {r[1] for r in conn.execute(text("PRAGMA table_info(notifications)"))}
        if ncols and "task_id" not in ncols:
            conn.execute(text("ALTER TABLE notifications ADD COLUMN task_id INTEGER"))
        # users: job title + contact phone (for the profile card)
        ucols = {r[1] for r in conn.execute(text("PRAGMA table_info(users)"))}
        if "title" not in ucols:
            conn.execute(text("ALTER TABLE users ADD COLUMN title VARCHAR(120)"))
        if "email_notifications" not in ucols:
            conn.execute(text("ALTER TABLE users ADD COLUMN email_notifications BOOLEAN DEFAULT 1"))
        if "notif_prefs" not in ucols:
            conn.execute(text("ALTER TABLE users ADD COLUMN notif_prefs JSON"))
        if "is_active" not in ucols:
            conn.execute(text("ALTER TABLE users ADD COLUMN is_active BOOLEAN DEFAULT 1"))
        _demo = {
            1: ("מנהל אגף הנדסה ותשתיות", "050-2345678"),
            2: ("מנהלת מחלקת תחבורה",     "052-3456789"),
            3: ("מנהל מחלקת תכנון",       "054-4567890"),
            4: ("רכזת פרויקטים",          "053-5678901"),
            5: ("מהנדס ביצוע",            "050-6789012"),
            6: ("רכזת שירות לתושב",       "052-7890123"),
            7: ("עובד תחזוקה",            "058-8901234"),
        }
        for uid, (title, phone) in _demo.items():
            conn.execute(text("UPDATE users SET title=COALESCE(title,:t), phone=COALESCE(phone,:p) WHERE id=:id"),
                         {"t": title, "p": phone, "id": uid})
# The additive migrations use SQLite PRAGMA/ALTER; on Postgres create_all()
# already builds every table with all current columns, so they're not needed.
if IS_SQLITE:
    _migrate()
else:
    # Postgres: create_all builds new tables fully, but columns added later to
    # already-existing tables must be back-filled. Each ALTER runs in its OWN
    # transaction so one failure can't roll back the others (notably item_uid,
    # which the current model always SELECTs — if it's missing, task loads break).
    from sqlalchemy import text as _text
    _PG_MIGRATIONS = [
        "ALTER TABLE comments ADD COLUMN IF NOT EXISTS seen_by JSON",
        "ALTER TABLE notifications ADD COLUMN IF NOT EXISTS task_id INTEGER",
        "ALTER TABLE boards ADD COLUMN IF NOT EXISTS environment_id INTEGER",
        "ALTER TABLE boards ADD COLUMN IF NOT EXISTS folder_id INTEGER",
        "ALTER TABLE boards ADD COLUMN IF NOT EXISTS position INTEGER DEFAULT 0",
        "ALTER TABLE environments ADD COLUMN IF NOT EXISTS is_primary BOOLEAN DEFAULT FALSE",
        "ALTER TABLE environment_members ADD COLUMN IF NOT EXISTS role VARCHAR(20) DEFAULT 'member'",
        "ALTER TABLE tasks ADD COLUMN IF NOT EXISTS item_uid VARCHAR(11)",
        "ALTER TABLE tasks ADD COLUMN IF NOT EXISTS item_uid_board INTEGER",
    ]
    for _sql in _PG_MIGRATIONS:
        try:
            with engine.begin() as _conn:
                _conn.execute(_text(_sql))
        except Exception as _e:
            print(f"⚠️  pg migration skipped: {_sql!r} → {_e}", flush=True)

# Seed on first run only — on a persistent DB (Postgres) skip if data exists,
# so real data isn't duplicated or overwritten on every cold start.
from seed import seed_database, seed_work_plan
with Session(engine) as _seed_db:
    _db_empty = _seed_db.query(Board).count() == 0
if _db_empty:
    seed_database(engine)
    seed_work_plan(engine)

REMOVED_DEMO_USERS = {
    "רחל לוי": "rachel@hodhasharon.gov.il",
    "שרה ברק": "sarah@hodhasharon.gov.il",
    "יוסי אברהם": "yossi@hodhasharon.gov.il",
    "נעה שטרן": "noa@hodhasharon.gov.il",
    "עומר גולן": "omer@hodhasharon.gov.il",
}

def _purge_removed_demo_users():
    """Keep removed demo users out of existing local/serverless databases."""
    names = set(REMOVED_DEMO_USERS)
    emails = set(REMOVED_DEMO_USERS.values())
    with Session(engine) as db:
        users = db.query(User).filter(or_(User.name.in_(names), User.email.in_(emails))).all()
        ids = [u.id for u in users]
        if not ids:
            return

        db.execute(task_assignees.delete().where(task_assignees.c.user_id.in_(ids)))
        db.execute(task_watchers.delete().where(task_watchers.c.user_id.in_(ids)))
        db.query(BoardMember).filter(BoardMember.user_id.in_(ids)).delete(synchronize_session=False)
        db.query(WorkspaceMember).filter(WorkspaceMember.user_id.in_(ids)).delete(synchronize_session=False)
        db.query(LoginEvent).filter(LoginEvent.user_id.in_(ids)).delete(synchronize_session=False)
        db.query(Notification).filter(Notification.user_id.in_(ids)).delete(synchronize_session=False)

        db.query(Task).filter(Task.created_by.in_(ids)).update({Task.created_by: None}, synchronize_session=False)
        db.query(Comment).filter(Comment.user_id.in_(ids)).update({Comment.user_id: None}, synchronize_session=False)
        db.query(Permit).filter(Permit.assigned_to.in_(ids)).update({Permit.assigned_to: None}, synchronize_session=False)
        db.query(CitizenRequest).filter(CitizenRequest.assigned_to.in_(ids)).update({CitizenRequest.assigned_to: None}, synchronize_session=False)
        db.query(Project).filter(Project.manager_id.in_(ids)).update({Project.manager_id: None}, synchronize_session=False)
        db.query(ProjectStep).filter(ProjectStep.owner_id.in_(ids)).update({ProjectStep.owner_id: None}, synchronize_session=False)
        db.query(Approval).filter(Approval.approver_user_id.in_(ids)).update({Approval.approver_user_id: None}, synchronize_session=False)
        db.query(ChangeRequest).filter(ChangeRequest.requested_by.in_(ids)).update({ChangeRequest.requested_by: None}, synchronize_session=False)
        db.query(ChangeRequest).filter(ChangeRequest.approved_by.in_(ids)).update({ChangeRequest.approved_by: None}, synchronize_session=False)
        db.query(Document).filter(Document.uploaded_by.in_(ids)).update({Document.uploaded_by: None}, synchronize_session=False)
        db.query(AuditLog).filter(AuditLog.changed_by.in_(ids)).update({AuditLog.changed_by: None}, synchronize_session=False)

        db.query(User).filter(User.id.in_(ids)).delete(synchronize_session=False)
        db.commit()

def _safe_startup(fn):
    """Run a one-time startup/seed step; never let it crash the whole process.
    On a persistent prod DB these are mostly idempotent no-ops, so a failure
    against existing data must not take the server down — log and carry on."""
    try:
        fn()
    except Exception as _e:
        print(f"⚠️  startup step {getattr(fn,'__name__',fn)} failed: {_e}", flush=True)

_safe_startup(_purge_removed_demo_users)

def _seed_memberships():
    """One-time: give existing boards their members so nothing disappears.
    New boards start private to their creator (see create_board)."""
    with Session(engine) as db:
        users = db.query(User).all()
        # workspace/environment members
        if db.query(WorkspaceMember).count() == 0:
            for u in users:
                # only a true system admin (User.role == "admin") is a workspace/system
                # admin; a "manager" is NOT — they get access only to granted environments
                erole = "admin" if u.role == "admin" else ("viewer" if u.role == "viewer" else "member")
                db.add(WorkspaceMember(user_id=u.id, role=erole))
            db.commit()
        # board members
        if db.query(BoardMember).count() == 0:
            for b in db.query(Board).all():
                for u in users:
                    role = "viewer" if u.role == "viewer" else ("admin" if u.role in ("admin", "manager") else "editor")
                    db.add(BoardMember(board_id=b.id, user_id=u.id, role=role))
            db.commit()
_safe_startup(_seed_memberships)

# The municipality's environments (Monday-style workspaces, like 'עיריית הוד השרון').
# Seeded once; sysadmins manage them afterwards via /api/environments.
ENVIRONMENTS_SEED = [
    ("בינוי", "🏗️"), ("היסעים", "🚌"), ("חדשנות טכנולוגיה ומערכות מידע", "💡"),
    ("כספים", "💰"), ("מגזר ערבי דרוזי וצ'רקסי", "🕌"), ("מטה", "🏛️"),
    ("מינהל ואיכות", "📋"), ("משאבי אנוש", "👥"), ("משפטית", "⚖️"),
    ("סוכנות הביטוח", "🛡️"), ("פיקוח ובקרה", "🔎"), ("פיתוח תשתיות ואחזקה", "🛠️"),
    ('פמ"א', "🏭"), ("קשרי לקוחות", "🤝"), ("רגולציה וקשרי ממשל", "📜"),
    ("שיווק", "📣"), ('תבו"ר', "💵"),
]
_ENV_COLORS = ["#6366f1", "#0ea5e9", "#10b981", "#f59e0b", "#ef4444", "#8b5cf6", "#ec4899", "#14b8a6"]

# The primary workspace is matched BY NAME on every startup — renaming it in the
# UI without changing this would make the seeder create a second primary
# environment under the old name.
PRIMARY_ENV_NAME = "בדיקה"

def _seed_environments():
    """Seed environments on first run (idempotent by name):
    - a primary workspace (PRIMARY_ENV_NAME) that holds the existing/legacy boards
    - the municipality's 17 environments
    Also gives every current user access to the primary workspace so nothing
    disappears, and attaches legacy (unassigned) boards to the primary workspace."""
    with Session(engine) as db:
        org = db.query(Organization).first()
        existing = {e.name: e for e in db.query(Environment).all()}
        # 1) primary workspace first. An existing primary wins over the name:
        # renaming it (it is an ordinary editable workspace) must not make this
        # seeder mint a second primary under the old name on the next boot.
        #
        # Repair first: more than one workspace flagged primary is never valid.
        # It is the residue of the older name-matching logic — renaming the
        # default workspace made the next boot mint another one under the old
        # name, each flagged primary. A primary workspace cannot be renamed away
        # from, moved out of or deleted, so every extra one is a workspace the
        # admin can no longer get rid of. Keep the one that actually holds the
        # boards (ties: the oldest) and demote the rest to ordinary workspaces.
        prims = db.query(Environment).filter(Environment.is_primary == True).order_by(Environment.id).all()
        if len(prims) > 1:
            counts = {e.id: db.query(Board).filter(Board.environment_id == e.id).count() for e in prims}
            keep = max(prims, key=lambda e: (counts[e.id], -e.id))
            for e in prims:
                if e.id != keep.id:
                    e.is_primary = False
            db.flush()
            print(f"↻ demoted {len(prims) - 1} extra default workspace(s); kept '{keep.name}'")
            prims = [keep]
        primary = (prims[0] if prims else None) or existing.get(PRIMARY_ENV_NAME)
        if not primary:
            primary = Environment(name=PRIMARY_ENV_NAME, icon="🏛️", color=_ENV_COLORS[0],
                                  position=0, is_primary=True, organization_id=org.id if org else None)
            db.add(primary); db.flush()
        elif not primary.is_primary:
            primary.is_primary = True
        # 2) the 17 municipality environments after it
        pos = max([e.position for e in existing.values()], default=0) + 1
        for name, icon in ENVIRONMENTS_SEED:
            if name in existing:
                continue
            db.add(Environment(name=name, icon=icon, color=_ENV_COLORS[pos % len(_ENV_COLORS)],
                               position=pos, organization_id=org.id if org else None))
            pos += 1
        # 3) legacy boards (no environment) belong to the primary workspace
        db.query(Board).filter(Board.environment_id == None).update({Board.environment_id: primary.id})
        # 4) every current user gets access to the primary workspace (once)
        if db.query(EnvironmentMember).filter(EnvironmentMember.environment_id == primary.id).count() == 0:
            for u in db.query(User).all():
                db.add(EnvironmentMember(environment_id=primary.id, user_id=u.id))
        db.commit()
_safe_startup(_seed_environments)


AUTO_COLS = (("created_at", "מועד יצירה", "sys_created_at"),
             ("created_by", "יוצר הרשומה", "sys_created_by"))

def _backfill_auto_columns():
    """Give every board the two automatic columns — who created an item and when.

    New boards ship with them; boards that predate them did not have them at
    all, and the creator column on the oldest ones is a plain people column
    that only auto-filled for items created while it existed. Both are derived
    from the item itself, so adding them is enough — no values to write, and
    they are right for items that predate the column.

    Runs once per board (`auto_cols_v1`): an admin who then removes a column
    must not have it come back on the next restart.
    """
    with Session(engine) as db:
        added = 0
        for b in db.query(Board).all():
            s = dict(b.settings or {})
            cols = [dict(c) for c in (s.get("columns") or [])]
            changed = False
            # the legacy creator column is a people column that has to be written
            # on creation — the derived type reads the item's own creator instead
            for c in cols:
                if c.get("id") == "sys_created_by" and c.get("type") == "people":
                    c["type"] = "created_by"
                    changed = True
            if not s.get("auto_cols_v1"):
                have = {c.get("type") for c in cols}
                used = {c.get("id") for c in cols}
                for typ, title, sid in AUTO_COLS:
                    if typ in have:
                        continue
                    cid = sid if sid not in used else ("col_" + uuid.uuid4().hex[:8])
                    cols.append({"id": cid, "type": typ, "title": title, "options": None,
                                 "formula": None, "connect": None, "copy_mode": None, "perms": {}})
                    used.add(cid)
                    added += 1
                    changed = True
                s["auto_cols_v1"] = True
                changed = True
            if changed:
                s["columns"] = cols
                b.settings = s
        db.commit()
        if added:
            print(f"↻ added {added} automatic column(s) across boards", flush=True)
_safe_startup(_backfill_auto_columns)

# ── Configurable capability matrix (role → feature) ──────────────────
# 'admin' always has every capability (implicit, never stored). These roles are
# editable by a system admin from the SysAdmin → role-permissions panel.
CAPABILITIES = ["manage_system", "create_environment", "create_board"]
CAP_HE = {
    "manage_system": "ניהול מערכת",
    "create_environment": "יצירת סביבה חדשה",
    "create_board": "יצירת לוח חדש",
}
EDITABLE_ROLES = ["manager", "member", "viewer", "guest"]
ROLE_HE_BE = {"admin": "מנהל מערכת", "manager": "מנהל", "member": "חבר צוות",
              "viewer": "צפייה בלבד", "guest": "אורח"}

def _seed_role_permissions():
    """Ensure a row exists for every (editable role × capability). Defaults to
    False — only 'admin' (implicit) starts with access, matching prior behavior."""
    with Session(engine) as db:
        existing = {(r.role, r.capability) for r in db.query(RolePermission).all()}
        added = False
        for role in EDITABLE_ROLES:
            for cap in CAPABILITIES:
                if (role, cap) not in existing:
                    db.add(RolePermission(role=role, capability=cap, allowed=False))
                    added = True
        if added:
            db.commit()
_safe_startup(_seed_role_permissions)

def _role_perms(db):
    """Return {role: {capability: bool}} for the editable roles."""
    m = {role: {cap: False for cap in CAPABILITIES} for role in EDITABLE_ROLES}
    for r in db.query(RolePermission).all():
        if r.role in m and r.capability in CAPABILITIES:
            m[r.role][r.capability] = bool(r.allowed)
    return m

def _cap(db, uid, cap):
    """Whether the user (by User.role, with 'admin' always-on) has a capability."""
    if uid is None:
        return False
    if _ws_role(db, uid) == "admin":
        return True
    u = db.query(User).filter(User.id == uid).first()
    role = (u.role if u else None) or "guest"
    if role == "admin":
        return True
    row = (db.query(RolePermission)
           .filter(RolePermission.role == role, RolePermission.capability == cap).first())
    return bool(row and row.allowed)

def get_db():
    with Session(engine) as session:
        yield session

# ── API Models ───────────────────────────────────────────────────────

class TaskOut(BaseModel):
    id: int; board_id: int; group_id: Optional[int]
    title: str; description: Optional[str]
    status: str; priority: str; position: int
    due_date: Optional[datetime]; start_date: Optional[datetime]
    estimated_hours: Optional[float]; actual_hours: Optional[float]
    location_lat: Optional[float]; location_lng: Optional[float]
    address: Optional[str]; gis_layer_id: Optional[str]
    tags: list; custom_fields: dict; is_archived: bool
    created_by: Optional[int]
    created_at: datetime; updated_at: datetime
    assignees: list = []
    subtask_count: int = 0
    comment_count: int = 0

class BoardOut(BaseModel):
    id: int; name: str; description: Optional[str]
    board_type: str; icon: str; color: str
    is_archived: bool
    department_name: Optional[str] = ""
    groups: list = []
    tasks: list = []
    task_count: int = 0

class DashboardOut(BaseModel):
    total_tasks: int
    tasks_by_status: dict
    tasks_by_priority: dict
    overdue_tasks: int
    citizen_requests_open: int
    permits_pending: int
    recent_activity: list = []

# ── API Routes ───────────────────────────────────────────────────────

@app.get("/api/status")
def status(request: Request):
    # The SPA shows the local user picker only in dev mode, so this has to be
    # the mode THIS request is actually authenticated under — dev is honoured
    # only when the app is served locally. It exposes no user data.
    return {"status": "ok", "app": "CODE-CAL MISSIONS", "version": "0.1.0",
            "auth_mode": effective_auth_mode(request)}

@app.get("/api/dashboard")
def dashboard(user_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        visible = _visible_board_ids(db, user_id)
        tasks = db.query(Task).filter(
            Task.is_archived == False, Task.board_id.in_(visible)
        ).all() if visible else []
        total = len(tasks)
        by_status = {}
        by_priority = {}
        overdue = 0
        now = datetime.now(timezone.utc).replace(tzinfo=None)
        for t in tasks:
            s = t.status.value if hasattr(t.status, 'value') else t.status
            p = t.priority.value if hasattr(t.priority, 'value') else t.priority
            by_status[s] = by_status.get(s, 0) + 1
            by_priority[p] = by_priority.get(p, 0) + 1
            due = t.due_date
            if due and hasattr(due, 'tzinfo') and due.tzinfo:
                due = due.replace(tzinfo=None)
            if due and due < now and s not in ("done", "cancelled"):
                overdue += 1
        citizen_open = db.query(CitizenRequest).filter(
            CitizenRequest.status.in_(["new", "assigned", "in_progress"])
        ).count()
        permits_pending = db.query(Permit).filter(
            Permit.status.in_(["draft", "submitted", "in_review"])
        ).count()
        return {
            "total_tasks": total,
            "tasks_by_status": by_status,
            "tasks_by_priority": by_priority,
            "overdue_tasks": overdue,
            "citizen_requests_open": citizen_open,
            "permits_pending": permits_pending,
            "recent_activity": [],
        }

# Views a board can expose. "table" is always present and cannot be removed.
ALL_VIEWS = ["table", "kanban", "gantt", "dashboard", "ceo", "calendar", "map", "gis"]

def _board_views(b):
    """Enabled views for a board. New boards start with only the main table;
    older/seeded boards (no explicit setting) keep all views for compatibility."""
    s = b.settings or {}
    v = s.get("views")
    if not v:
        return list(ALL_VIEWS)
    # always keep table first and present
    out = ["table"] + [x for x in v if x in ALL_VIEWS and x != "table"]
    return out

# Per-board status vocabulary. Each entry is {key,label,color} where `key` is one
# of the underlying TaskStatus values. Keeping the enum key under the hood means
# group auto-move, parent roll-up, kanban and charts keep working (they all key
# off task.status) while a board admin may rename/recolor/reorder the statuses
# and expose only the ones the board actually uses.
STATUS_DEFAULTS = [
    {"key": "backlog", "label": "בתכנון", "color": "#9699a6"},
    {"key": "todo", "label": "לביצוע", "color": "#c4c4c4"},
    {"key": "in_progress", "label": "בתהליך", "color": "#fdab3d"},
    {"key": "review", "label": "בבדיקה", "color": "#579bfc"},
    {"key": "on_hold", "label": "בהמתנה", "color": "#808080"},
    {"key": "done", "label": "הושלם", "color": "#00c875"},
    {"key": "cancelled", "label": "בוטל", "color": "#e2445c"},
]
_STATUS_KEYS = {s["key"] for s in STATUS_DEFAULTS}

def _valid_hex(c):
    c = str(c or "")
    if len(c) == 7 and c[0] == "#":
        try:
            int(c[1:], 16); return True
        except ValueError:
            return False
    return False

# A custom "status" column carries its own vocabulary — the board admin names and
# colours the labels of every such column separately (the built-in status column
# has its own, board-wide list above). Stored on the column as
# options=[{label,color}]; an empty/absent list means "use the client defaults".
STATUS_COL_MAX = 30

def _status_col_options(raw):
    """Normalise a status column's own vocabulary, or None when unset.

    Each option carries a stable id. A cell stores that id, so renaming or
    recolouring the option is a change in one place that every cell already
    referencing it picks up — no rewriting of item data, and nothing to go
    stale if a rewrite were to fail halfway.
    """
    if not isinstance(raw, list):
        return None
    out, seen, ids = [], set(), set()
    for it in raw:
        if isinstance(it, str):
            it = {"label": it}
        if not isinstance(it, dict):
            continue
        label = (str(it.get("label") or "").strip())[:40]
        if not label or label in seen:
            continue
        seen.add(label)
        oid = str(it.get("id") or "").strip()
        if not oid or oid in ids:
            oid = "o_" + uuid.uuid4().hex[:8]
        ids.add(oid)
        out.append({"id": oid, "label": label,
                    "color": it.get("color") if _valid_hex(it.get("color")) else "#c4c4c4"})
        if len(out) >= STATUS_COL_MAX:
            break
    return out or None

# A board may define up to this many statuses. Only seven of them can be a
# TaskStatus value — the ones the engine keys off for group auto-move, kanban
# columns and charts. Any status beyond those seven is a board-defined label
# that *behaves as* one of the seven (its `base`): the item stores the base in
# task.status so all of that keeps working, and carries the chosen label and
# colour in custom_fields, which is what the grid renders.
BOARD_STATUS_MAX = 30

def _norm_board_status(it, seen):
    """One validated {key,label,color,base} entry, or None to skip it."""
    if not isinstance(it, dict):
        return None
    key = str(it.get("key") or "").strip()
    label = (str(it.get("label") or "").strip())[:40]
    if key in _STATUS_KEYS:                       # one of the seven built-ins
        base = key
    elif not key or key.startswith("x_"):          # a board-defined status
        # no key means "newly added" — the server mints it. An arbitrary
        # unknown key is still a client mistake and is dropped, as before.
        base = it.get("base") if it.get("base") in _STATUS_KEYS else "in_progress"
        key = key or ("x_" + uuid.uuid4().hex[:8])
    else:
        return None
    if not label:
        label = key
    if key in seen:
        return None
    seen.add(key)
    color = it.get("color") if _valid_hex(it.get("color")) else "#c4c4c4"
    return {"key": key, "label": label, "color": color, "base": base}

def _board_statuses(b):
    """The board's ordered status list (defaults when the admin hasn't customised)."""
    raw = (b.settings or {}).get("statuses")
    if not raw:
        return [dict(x, base=x["key"]) for x in STATUS_DEFAULTS]
    out, seen = [], set()
    for it in raw:
        e = _norm_board_status(it, seen)
        if e:
            out.append(e)
        if len(out) >= BOARD_STATUS_MAX:
            break
    return out or [dict(x, base=x["key"]) for x in STATUS_DEFAULTS]

# ── Priority is a status column too ──────────────────────────────────────
# "עדיפות" is the same kind of thing as "סטטוס": one value out of a named,
# coloured vocabulary. So it gets the same treatment — a board admin renames,
# recolours, reorders, adds and removes its values, up to BOARD_PRIORITY_MAX,
# exactly as with the statuses above. Five of them can be a Priority enum value
# (what sorting and the charts key off); anything beyond behaves as one of those
# five (its `base`), with the chosen label and colour carried on the item.
PRIORITY_DEFAULTS = [
    {"key": "low", "label": "נמוכה", "color": "#579bfc"},
    {"key": "medium", "label": "בינונית", "color": "#5559df"},
    {"key": "high", "label": "גבוהה", "color": "#fdab3d"},
    {"key": "critical", "label": "קריטית", "color": "#e2445c"},
    {"key": "emergency", "label": "חירום", "color": "#bb3354"},
]
_PRIORITY_KEYS = {p["key"] for p in PRIORITY_DEFAULTS}
BOARD_PRIORITY_MAX = 30

def _norm_board_priority(it, seen):
    """One validated {key,label,color,base} priority, or None to skip it."""
    if not isinstance(it, dict):
        return None
    key = str(it.get("key") or "").strip()
    label = (str(it.get("label") or "").strip())[:40]
    if key in _PRIORITY_KEYS:                      # one of the five built-ins
        base = key
    elif not key or key.startswith("x_"):           # a board-defined priority
        base = it.get("base") if it.get("base") in _PRIORITY_KEYS else "medium"
        key = key or ("x_" + uuid.uuid4().hex[:8])
    else:
        return None
    if not label:
        label = key
    if key in seen:
        return None
    seen.add(key)
    color = it.get("color") if _valid_hex(it.get("color")) else "#c4c4c4"
    return {"key": key, "label": label, "color": color, "base": base}

def _board_priorities(b):
    """The board's ordered priority list (defaults when it hasn't customised)."""
    raw = (b.settings or {}).get("priorities")
    if not raw:
        return [dict(x, base=x["key"]) for x in PRIORITY_DEFAULTS]
    out, seen = [], set()
    for it in raw:
        e = _norm_board_priority(it, seen)
        if e:
            out.append(e)
        if len(out) >= BOARD_PRIORITY_MAX:
            break
    return out or [dict(x, base=x["key"]) for x in PRIORITY_DEFAULTS]

def _resolve_priority(board, value):
    """Map a chosen priority key to (engine value, the fields the item carries).

    A board-defined priority is not a Priority enum value, so it is stored as its
    `base` with the chosen label and colour on the item — the same arrangement as
    a board-defined status. Returns (None, None) for a key this board does not
    know, so the caller can leave the priority alone rather than crash on it.
    """
    key = str(value or "").strip()
    if not key:
        return None, None
    cleared = {"priority_key": None, "priority_label": None,
               "priority_color": None, "priority_unset": None}
    if key in _PRIORITY_KEYS:
        return key, cleared
    e = next((p for p in _board_priorities(board) if p["key"] == key), None) if board else None
    if not e:
        return None, None
    return e["base"], {"priority_key": e["key"], "priority_label": e["label"],
                       "priority_color": e["color"], "priority_unset": None}

def _board_role(db, board_id, user_id):
    """Board-scoped role (admin/editor/viewer) for a user, or None if not a member."""
    if user_id is None:
        return None
    m = db.query(BoardMember).filter(BoardMember.board_id == board_id,
                                     BoardMember.user_id == user_id).first()
    if m:
        return m.role
    # system (workspace) admins have admin access to every board, even ones
    # they were never explicitly added to
    if _ws_role(db, user_id) == "admin":
        return "admin"
    return None

def _is_board_admin(db, board_id, user_id):
    return _board_role(db, board_id, user_id) == "admin"

_BROLE_ORDER = {"viewer": 0, "editor": 1, "admin": 2}
def _require_board_edit(db, board_id, actor, need="editor"):
    """Enforce board-level write permission (closes IDOR on board mutations).
    The actor must be a board member whose role meets `need` (editor|admin).

    This used to return silently when `actor` was None, which meant dropping the
    parameter skipped the check entirely. Identity now comes from the transport,
    so None means unauthenticated and must be refused."""
    if actor is None:
        raise HTTPException(401, "נדרשת התחברות")
    role = _board_role(db, board_id, actor)
    if role is None or _BROLE_ORDER.get(role, -1) < _BROLE_ORDER.get(need, 1):
        raise HTTPException(403, "אין לך הרשאה לשנות לוח זה")

def _visible_board_ids(db, user_id):
    """Set of board IDs a user is allowed to see (membership-based).
    Mirrors /api/boards visibility so aggregate views (dashboard, CEO,
    insights) never leak boards the user was not invited to.
    An unauthenticated caller (user_id=None) sees nothing — this used to return
    every board, so omitting the parameter exposed the whole workspace."""
    all_ids = {b for (b,) in db.query(Board.id).filter(Board.is_archived == False).all()}
    if user_id is None:
        return set()
    if _ws_role(db, user_id) == "admin":   # system admins see every board
        return all_ids
    member_ids = {m.board_id for m in db.query(BoardMember).filter(BoardMember.user_id == user_id).all()}
    return {i for i in all_ids if i in member_ids}

def _ws_role(db, user_id):
    if user_id is None:
        return None
    m = db.query(WorkspaceMember).filter(WorkspaceMember.user_id == user_id).first()
    return m.role if m else None

# ── Multi-level permission model (סביבה → לוח → פריט → עמודה) ────────
def _board_caps(role):
    """Capabilities implied by a board role."""
    return {
        "admin":  {"view", "edit", "delete", "manage"},
        "editor": {"view", "edit"},
        "viewer": {"view"},
    }.get(role, set())

def _item_perm(task, user_id, board_role):
    """Effective permission on a specific item: none|view|edit|delete.
    A per-user item override wins over the board role."""
    ov = (task.permissions or {}).get(str(user_id))
    if ov in ("none", "view", "edit", "delete"):
        return ov
    caps = _board_caps(board_role)
    return "delete" if "delete" in caps else ("edit" if "edit" in caps else ("view" if "view" in caps else "none"))

def _col_perm(col, user_id, board_role):
    """Effective permission on a column for a user: none|view|edit."""
    ov = (col.get("perms") or {}).get(str(user_id))
    if ov in ("none", "view", "edit"):
        return ov
    caps = _board_caps(board_role)
    return "edit" if "edit" in caps else ("view" if "view" in caps else "none")

@app.get("/api/boards")
def list_boards(user_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        boards = db.query(Board).filter(Board.is_archived == False).order_by(
            Board.position, Board.id).all()
        # system (workspace) admins see every board; everyone else only the boards
        # they were invited to (membership-based visibility)
        if user_id is not None and _ws_role(db, user_id) != "admin":
            member_ids = {m.board_id for m in db.query(BoardMember).filter(BoardMember.user_id == user_id).all()}
            boards = [b for b in boards if b.id in member_ids]
        result = []
        for b in boards:
            task_count = db.query(Task).filter(Task.board_id == b.id, Task.parent_id == None).count()
            dept_name = db.query(Department.name).filter(Department.id == b.department_id).scalar() or ""
            groups = db.query(Group).filter(Group.board_id == b.id).order_by(Group.position).all()
            result.append({
                "id": b.id, "name": b.name, "description": b.description,
                "board_type": b.board_type.value if hasattr(b.board_type, 'value') else b.board_type,
                "icon": b.icon, "color": b.color,
                "is_archived": b.is_archived,
                "department_name": dept_name,
                "environment_id": b.environment_id,
                "folder_id": b.folder_id,
                "position": b.position or 0,
                "task_count": task_count,
                "my_role": _board_role(db, b.id, user_id) if user_id is not None else None,
                "can_manage_env": _can_manage_env(db, b.environment_id, user_id) if user_id is not None else False,
                "groups": [{"id": g.id, "name": g.name, "position": g.position, "color": g.color, "task_status": g.task_status.value if hasattr(g.task_status, 'value') else g.task_status} for g in groups],
            })
        return result

# ── "מזהה פריט" (item id column) ─────────────────────────────────────
# A purely technical identifier: 11 digits, generated automatically, unique
# inside its board, with no prefix and no business meaning. It cannot be edited
# by hand. Sub-items are issued one by the same logic. Because the identifier is
# tied to the board it was minted for, an item that lands in another board (and
# later comes back) is issued a new one rather than keeping the old number.
ITEM_UID_MIN = 10_000_000_000     # smallest 11-digit number
ITEM_UID_MAX = 99_999_999_999

def _new_item_uid(db, board_id):
    for _ in range(60):
        cand = str(secrets.randbelow(ITEM_UID_MAX - ITEM_UID_MIN + 1) + ITEM_UID_MIN)
        taken = db.query(Task.id).filter(Task.board_id == board_id, Task.item_uid == cand).first()
        if not taken:
            return cand
    return str(secrets.randbelow(ITEM_UID_MAX - ITEM_UID_MIN + 1) + ITEM_UID_MIN)

def _ensure_item_uid(db, task):
    """Give the task an identifier if it has none, or a fresh one if the one it
    carries was minted for a different board. Returns the identifier."""
    if task.item_uid and task.item_uid_board == task.board_id:
        return task.item_uid
    task.item_uid = _new_item_uid(db, task.board_id)
    task.item_uid_board = task.board_id
    return task.item_uid

def _board_has_item_id_col(board):
    return any(c.get("type") == "item_id" for c in ((board.settings or {}).get("columns") or []))

def _backfill_item_uids(db, board_id):
    """Every item and sub-item on the board gets an identifier — used when the
    column is added to a board that already holds items."""
    rows = db.query(Task).filter(Task.board_id == board_id).all()
    changed = False
    for t in rows:
        before = t.item_uid
        _ensure_item_uid(db, t)
        changed = changed or (t.item_uid != before)
    if changed:
        db.commit()

def _serialize_task(t, db, with_subs=True, user_id=None, board_role=None, columns=None):
    assignees = [{"id": u.id, "name": u.name, "avatar_url": u.avatar_url} for u in t.assignees] if t.assignees else []
    subtask_count = db.query(Task).filter(Task.parent_id == t.id).count()
    comment_count = db.query(Comment).filter(Comment.task_id == t.id).count()
    my_perm = _item_perm(t, user_id, board_role) if user_id is not None else "delete"
    cf = dict(t.custom_fields or {})
    col_perms = {}
    if user_id is not None:
        for c in (columns or []):
            cp = _col_perm(c, user_id, board_role)
            col_perms[c["id"]] = cp
            if cp == "none":         # view-restricted column → mask its value
                cf.pop(c["id"], None)
    d = {
        "id": t.id, "board_id": t.board_id, "group_id": t.group_id,
        "parent_id": t.parent_id,
        "item_uid": t.item_uid,          # "מזהה פריט" — automatic, read-only
        "title": t.title, "description": t.description,
        "status": t.status.value if hasattr(t.status, 'value') else t.status,
        "priority": t.priority.value if hasattr(t.priority, 'value') else t.priority,
        "position": t.position, "due_date": t.due_date,
        "start_date": t.start_date,
        "estimated_hours": t.estimated_hours, "actual_hours": t.actual_hours,
        "location_lat": t.location_lat, "location_lng": t.location_lng,
        "address": t.address, "tags": t.tags or [],
        "custom_fields": cf,
        "permissions": t.permissions or {},
        "my_perm": my_perm,
        "col_perms": col_perms,
        "is_archived": t.is_archived,
        "created_by": t.created_by,
        "created_at": t.created_at, "updated_at": t.updated_at,
        "assignees": assignees,
        "subtask_count": subtask_count,
        "comment_count": comment_count,
    }
    if with_subs:
        subs = db.query(Task).filter(Task.parent_id == t.id, Task.is_archived == False).order_by(Task.position).all()
        subs = [s for s in subs if user_id is None or _item_perm(s, user_id, board_role) != "none"]
        d["subtasks"] = [_serialize_task(s, db, with_subs=False, user_id=user_id, board_role=board_role, columns=columns) for s in subs]
    return d

@app.get("/api/boards/{board_id}")
def get_board(board_id: int, user_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        b = db.query(Board).filter(Board.id == board_id).first()
        if not b:
            raise HTTPException(404, "Board not found")
        my_role = _board_role(db, board_id, user_id)
        if user_id is not None and my_role is None:
            raise HTTPException(403, "אין לך גישה ללוח זה")
        groups = db.query(Group).filter(Group.board_id == b.id).order_by(Group.position).all()
        # only top-level items as rows; sub-items are nested under their parent
        tasks = db.query(Task).filter(Task.board_id == b.id, Task.is_archived == False,
                                      Task.parent_id == None).order_by(Task.position).all()
        dept_name = db.query(Department.name).filter(Department.id == b.department_id).scalar() or ""
        columns = (b.settings or {}).get("columns", [])
        # a board carrying a "מזהה פריט" column shows an identifier for every
        # item — including ones that predate the column, or arrived from elsewhere
        if _board_has_item_id_col(b):
            _backfill_item_uids(db, b.id)
        # hide items the user has no view permission on
        if user_id is not None:
            tasks = [t for t in tasks if _item_perm(t, user_id, my_role) != "none"]
        tasks_out = [_serialize_task(t, db, user_id=user_id, board_role=my_role, columns=columns) for t in tasks]

        # board owners = members with the board-scoped admin role (creator + any
        # promoted admin). Exposed to every member so anyone can see who manages it.
        owners = []
        for m in db.query(BoardMember).filter(BoardMember.board_id == b.id, BoardMember.role == "admin").all():
            ou = db.query(User).filter(User.id == m.user_id).first()
            if ou:
                owners.append({"id": ou.id, "name": ou.name, "avatar_url": ou.avatar_url})

        return {
            "id": b.id, "name": b.name, "description": b.description,
            "board_type": b.board_type.value if hasattr(b.board_type, 'value') else b.board_type,
            "icon": b.icon, "color": b.color,
            "is_archived": b.is_archived,
            "department_name": dept_name,
            "views": _board_views(b),
            "view_only": (b.settings or {}).get("view_only", False),
            "my_role": my_role,
            "owners": owners,
            "columns": (b.settings or {}).get("columns", []),
            # which columns sub-items show (ids, name column first). None = never
            # configured → sub-items mirror the item columns.
            "sub_cols": (b.settings or {}).get("sub_cols"),
            "col_widths": (b.settings or {}).get("col_widths", {}),
            "col_labels": (b.settings or {}).get("col_labels", {}),
            "col_hidden": (b.settings or {}).get("col_hidden", []),
            "col_order": (b.settings or {}).get("col_order", []),
            "notifications_enabled": bool((b.settings or {}).get("notifications_enabled", True)),
            "statuses": _board_statuses(b),
            "priorities": _board_priorities(b),
            "form": (b.settings or {}).get("form"),
            "groups": [{"id": g.id, "name": g.name, "position": g.position, "color": g.color, "task_status": g.task_status.value if hasattr(g.task_status, 'value') else g.task_status} for g in groups],
            "tasks": tasks_out,
        }

@app.post("/api/boards")
def create_board(data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        # only a system (workspace) admin may create boards; the creator becomes
        # the board's first admin, so a board is never created without a manager.
        creator = actor_id
        # a new board belongs to the environment it was created in (default: primary workspace)
        env_id = data.get("environment_id")
        if not env_id:
            primary = db.query(Environment).filter(Environment.is_primary == True).first()
            env_id = primary.id if primary else None
        # allowed if the user has the global create-board capability, or manages the
        # target environment (an environment manager runs their own environment)
        if not (_cap(db, creator, "create_board") or _can_manage_env(db, env_id, creator)):
            raise HTTPException(403, "אין לך הרשאה ליצור לוח חדש")
        dept_id = data.get("department_id")
        if not dept_id:
            dept_id = db.query(Department.id).order_by(Department.id).limit(1).scalar()
        folder_id = data.get("folder_id")
        # append the new board at the bottom of its container (env root or folder)
        maxpos = (db.query(func.max(Board.position))
                  .filter(Board.environment_id == env_id, Board.folder_id == folder_id).scalar())
        b = Board(
            name=data.get("name", "לוח חדש"),
            description=data.get("description", ""),
            department_id=dept_id,
            environment_id=env_id,
            folder_id=folder_id,
            position=(maxpos + 1) if maxpos is not None else 0,
            board_type=BoardType.KANBAN,
            icon=data.get("icon", "📋"),
            color=data.get("color", "#0073ea"),
            settings={
                "views": ["table"],  # new boards start with only the main table
                # every new board ships with these auto columns (read-only, from the item's metadata)
                "columns": [
                    {"id": "sys_item_id", "type": "item_id", "title": "מספר מזהה"},
                    {"id": "sys_created_at", "type": "created_at", "title": "מועד יצירה"},
                    # derived from the item's own creator, like "מועד יצירה" above:
                    # nothing to fill in, so it is right on every item on every
                    # board — including boards and items that predate the column
                    {"id": "sys_created_by", "type": "created_by", "title": "יוצר הרשומה"},
                ],
            },
        )
        db.add(b)
        db.flush()
        # default groups: 3 clean stages (no "review"). More can be added in-board.
        g1 = Group(board_id=b.id, name="בתכנון", position=0, color="#579bfc", task_status=TaskStatus.BACKLOG)
        db.add_all([
            g1,
            Group(board_id=b.id, name="בביצוע", position=1, color="#fdab3d", task_status=TaskStatus.IN_PROGRESS),
            Group(board_id=b.id, name="הושלם", position=2, color="#00c875", task_status=TaskStatus.DONE),
        ])
        db.flush()
        # 3 starter items in the first group so the board isn't blank — the creator
        # is recorded as their author and auto-filled into the "יוצר הרשומה" column.
        # Nobody has chosen a status or a priority for them, so both read
        # "טרם הוגדר" until someone does, exactly like an item added by hand.
        for i in range(1, 4):
            cf = {"status_unset": True, "priority_unset": True}
            if creator:
                cf["sys_created_by"] = [creator]
            db.add(Task(board_id=b.id, group_id=g1.id, title=f"פריט {i}",
                        status=TaskStatus.BACKLOG, priority=Priority.MEDIUM, position=i,
                        created_by=creator, custom_fields=cf))
        # creator becomes the board admin; the board is private until they invite others
        db.add(BoardMember(board_id=b.id, user_id=creator, role="admin"))
        db.commit()
        db.refresh(b)
        return {"id": b.id, "name": b.name, "icon": b.icon, "views": _board_views(b), "my_role": "admin"}

@app.patch("/api/boards/{board_id}")
def update_board(board_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        b = db.query(Board).filter(Board.id == board_id).first()
        if not b:
            raise HTTPException(404, "board not found")
        if data.get("name") is not None:
            b.name = data["name"]
        if data.get("icon") is not None:
            b.icon = data["icon"]
        if data.get("color") is not None:
            b.color = data["color"]
        if data.get("views") is not None:
            views = ["table"] + [v for v in data["views"] if v in ALL_VIEWS and v != "table"]
            s = dict(b.settings or {})
            s["views"] = views
            b.settings = s
        if "view_only" in data:
            s = dict(b.settings or {})
            s["view_only"] = bool(data["view_only"])
            b.settings = s
        if "form" in data:
            # only a board admin can add/update the form
            if _board_role(db, board_id, actor_id) != "admin":
                raise HTTPException(403, "רק מנהל הלוח יכול לערוך את הטופס")
            s = dict(b.settings or {})
            s["form"] = data["form"]
            b.settings = s
        if data.get("col_widths") is not None:
            # only a board admin may set column widths (drag-to-resize)
            if _board_role(db, board_id, actor_id) != "admin":
                raise HTTPException(403, "רק מנהל הלוח יכול לשנות רוחב עמודות")
            widths = {}
            for k, v in (data["col_widths"] or {}).items():
                try:
                    widths[str(k)] = max(60, min(600, int(v)))
                except (TypeError, ValueError):
                    continue
            s = dict(b.settings or {})
            s["col_widths"] = widths
            b.settings = s
        if data.get("col_labels") is not None:
            # rename a built-in column header (e.g. "פריט") — board admin
            if _board_role(db, board_id, actor_id) != "admin":
                raise HTTPException(403, "רק מנהל הלוח יכול לשנות שם עמודה")
            builtin = {"item", "assignees", "status", "priority", "due", "tags"}
            labels = {}
            for k, v in (data["col_labels"] or {}).items():
                if k in builtin:
                    lv = str(v or "").strip()[:40]
                    if lv:
                        labels[str(k)] = lv
            s = dict(b.settings or {})
            s["col_labels"] = labels
            b.settings = s
        if data.get("col_hidden") is not None:
            # hide built-in columns (the "פריט" name column can never be hidden) — board admin
            if _board_role(db, board_id, actor_id) != "admin":
                raise HTTPException(403, "רק מנהל הלוח יכול להסתיר עמודות")
            hideable = {"assignees", "status", "priority", "due", "tags"}
            hidden = [str(k) for k in (data["col_hidden"] or []) if str(k) in hideable]
            s = dict(b.settings or {})
            s["col_hidden"] = list(dict.fromkeys(hidden))  # de-dup, preserve order
            b.settings = s
        if data.get("col_order") is not None:
            # persisted display order of columns (built-in keys + custom ids) — board admin
            if _board_role(db, board_id, actor_id) != "admin":
                raise HTTPException(403, "רק מנהל הלוח יכול לשנות סדר עמודות")
            order = [str(k) for k in (data["col_order"] or []) if isinstance(k, (str, int))]
            s = dict(b.settings or {})
            s["col_order"] = list(dict.fromkeys(order))  # de-dup, preserve first occurrence
            b.settings = s
        if "notifications_enabled" in data:
            # board admin toggles all notifications for this board on/off
            if _board_role(db, board_id, actor_id) != "admin":
                raise HTTPException(403, "רק מנהל הלוח יכול לשנות התראות ללוח")
            s = dict(b.settings or {})
            s["notifications_enabled"] = bool(data["notifications_enabled"])
            b.settings = s
        if data.get("statuses") is not None:
            # only a board admin may rename/recolor/reorder/add statuses
            if _board_role(db, board_id, actor_id) != "admin":
                raise HTTPException(403, "רק מנהל הלוח יכול לערוך סטטוסים")
            out, seen = [], set()
            for it in data["statuses"]:
                e = _norm_board_status(it, seen)
                if e:
                    out.append(e)
                if len(out) >= BOARD_STATUS_MAX:
                    break
            if not out:
                raise HTTPException(400, "חובה סטטוס אחד לפחות")
            s = dict(b.settings or {})
            s["statuses"] = out
            b.settings = s
        if data.get("priorities") is not None:
            # the priority column is a status column, so it is edited under the
            # same rule: only a board admin may rename/recolor/reorder/add
            if _board_role(db, board_id, actor_id) != "admin":
                raise HTTPException(403, "רק מנהל הלוח יכול לערוך עדיפויות")
            out, seen = [], set()
            for it in data["priorities"]:
                e = _norm_board_priority(it, seen)
                if e:
                    out.append(e)
                if len(out) >= BOARD_PRIORITY_MAX:
                    break
            if not out:
                raise HTTPException(400, "חובה עדיפות אחת לפחות")
            s = dict(b.settings or {})
            s["priorities"] = out
            b.settings = s
        if data.get("columns") is not None:
            # only a board admin may add/edit/remove columns and their options
            if _board_role(db, board_id, actor_id) != "admin":
                raise HTTPException(403, "רק מנהל הלוח יכול לערוך עמודות")
            # full replacement of the custom-column definitions
            allowed = {"timeline", "text", "number", "date", "rating", "status",
                       "people", "dropdown", "files", "accounts", "checkbox", "formula",
                       "connect", "created_at", "created_by", "item_id", "item_kind"}
            old_cols = {c.get("id"): c for c in (b.settings or {}).get("columns", [])}
            is_ws_admin = _ws_role(db, actor_id) == "admin"
            cols = []
            for c in data["columns"]:
                if not isinstance(c, dict) or c.get("type") not in allowed:
                    continue
                cid = c.get("id") or ("col_" + uuid.uuid4().hex[:8])
                if c["type"] == "connect":
                    # only a workspace (system) admin may create/change a connect column
                    prev = old_cols.get(cid)
                    changed = (not prev) or prev.get("type") != "connect" or prev.get("connect") != c.get("connect")
                    if changed and not is_ws_admin:
                        raise HTTPException(403, "רק מנהל מערכת יכול להוסיף או לשנות עמודת קישור בין לוחות")
                # a status column and a "סוג" column carry their own {label,color}
                # vocabulary; every other type keeps its options untouched
                opts = (_status_col_options(c.get("options"))
                        if c["type"] in ("status", "item_kind") else c.get("options"))
                cols.append({
                    "id": cid,
                    "type": c["type"],
                    "title": c.get("title") or c["type"],
                    "options": opts,
                    "formula": c.get("formula"),
                    "connect": c.get("connect") if c["type"] == "connect" else None,
                    # "מזהה פריט": what a click on the cell copies — the number
                    # itself ("id", the default) or a link to the item ("url")
                    "copy_mode": (c.get("copy_mode") if c.get("copy_mode") in ("id", "url") else "id")
                                 if c["type"] == "item_id" else None,
                    "perms": c.get("perms") or {},
                })
            s = dict(b.settings or {})
            s["columns"] = cols
            b.settings = s
            # adding the column issues an identifier to every item already on the
            # board (and to its sub-items)
            if any(c.get("type") == "item_id" for c in cols):
                for t in db.query(Task).filter(Task.board_id == board_id).all():
                    _ensure_item_uid(db, t)
        if data.get("sub_cols") is not None:
            # Sub-items have their own column structure: identical for every
            # sub-item on the board, but not necessarily the item's columns.
            # Stored as ids picked from the board's column pool (built-in keys +
            # custom column ids); the name column is always present and first.
            if _board_role(db, board_id, actor_id) != "admin":
                raise HTTPException(403, "רק מנהל הלוח יכול להגדיר את עמודות תת-הפריט")
            s = dict(b.settings or {})
            pool = {"assignees", "status", "priority", "due", "tags"}
            pool |= {str(c.get("id")) for c in s.get("columns", []) if c.get("id")}
            picked = [str(k) for k in (data["sub_cols"] or []) if str(k) in pool]
            s["sub_cols"] = ["item"] + list(dict.fromkeys(picked))
            b.settings = s
        db.commit()
        db.refresh(b)
        # statuses and priorities come back because the server mints the key for
        # a newly added one — the client cannot know it otherwise
        return {"id": b.id, "name": b.name, "icon": b.icon, "color": b.color,
                "views": _board_views(b), "columns": (b.settings or {}).get("columns", []),
                "statuses": _board_statuses(b), "priorities": _board_priorities(b)}

@app.delete("/api/boards/{board_id}")
def delete_board(board_id: int, user_id: int = Depends(current_user_id)):
    """Delete a board and everything under it (groups, items, comments,
    memberships). Only a board admin may delete it."""
    with Session(engine) as db:
        b = db.query(Board).filter(Board.id == board_id).first()
        if not b:
            raise HTTPException(404, "board not found")
        # a board admin, an environment manager (of the board's environment), or a
        # system admin may delete a board
        if user_id is not None and _board_role(db, board_id, user_id) != "admin" \
                and not _can_manage_env(db, b.environment_id, user_id):
            raise HTTPException(403, "רק מנהל הלוח או מנהל הסביבה יכול למחוק אותו")
        tasks = db.query(Task).filter(Task.board_id == board_id).all()
        task_ids = [t.id for t in tasks]
        if task_ids:
            db.query(Comment).filter(Comment.task_id.in_(task_ids)).delete(synchronize_session=False)
        for t in tasks:          # clear assignee links via the ORM relationship
            t.assignees = []
        db.flush()
        db.query(Task).filter(Task.board_id == board_id).delete(synchronize_session=False)
        db.query(Group).filter(Group.board_id == board_id).delete(synchronize_session=False)
        db.query(BoardMember).filter(BoardMember.board_id == board_id).delete(synchronize_session=False)
        db.delete(b)
        db.commit()
        return {"status": "deleted", "id": board_id}

# ── Board membership & per-board permissions ────────────────────────
BOARD_ROLES = ("admin", "editor", "viewer")
BROLE_HE = {"admin": "מנהל לוח", "editor": "עורך", "viewer": "צופה"}


def _notify(db, user_id, type, title, body="", board_id=None, task_id=None):
    """Create an in-app notification for a recipient (best-effort, no-op on bad input)."""
    if not user_id:
        return
    db.add(Notification(user_id=user_id, type=type, title=title, body=body,
                        board_id=board_id, task_id=task_id))

def _board_notify_on(board):
    """Whether the board emits notifications (board admin can switch it off wholesale)."""
    return bool((board.settings or {}).get("notifications_enabled", True)) if board else True


def _send_email(to_email, subject, html):
    """Send a transactional email via Resend. No-ops (logs) when RESEND_API_KEY
    isn't configured, so an invite never fails just because email isn't set up."""
    import urllib.request, json as _json
    api_key = os.environ.get("RESEND_API_KEY")
    if not api_key or not to_email:
        print(f"[email] skipped (no RESEND_API_KEY) → {to_email}: {subject}")
        return False
    from_addr = os.environ.get("INVITE_FROM_EMAIL", "CityOS <onboarding@resend.dev>")
    # development mode: until a domain is verified in Resend, real recipients are
    # blocked. Redirect every message to a single verified inbox so invites can be
    # tested end-to-end, keeping the original recipient visible in the email.
    redirect = os.environ.get("EMAIL_DEV_REDIRECT")
    actual_to = to_email
    if redirect:
        actual_to = redirect
        subject = f"[פיתוח · נועד ל-{to_email}] {subject}"
        html = (f"<div dir='rtl' style='background:#fff3cd;border:1px solid #ffe08a;"
                f"border-radius:8px;padding:10px 12px;margin-bottom:14px;font-size:12.5px;"
                f"font-family:Arial'>⚙️ מצב פיתוח — הנמען המקורי של ההזמנה הוא "
                f"<b>{to_email}</b>. מייל זה הופנה אליך לצורכי בדיקה.</div>") + html
    payload = _json.dumps({"from": from_addr, "to": [actual_to],
                           "subject": subject, "html": html}).encode("utf-8")
    req = urllib.request.Request("https://api.resend.com/emails", data=payload,
                                 headers={"Authorization": f"Bearer {api_key}",
                                          "Content-Type": "application/json",
                                          "User-Agent": "CityOS/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            r.read()
        print(f"[email] sent → {actual_to}: {subject}")
        return True
    except Exception as e:
        print(f"[email] send failed → {actual_to}: {e}")
        return False


def _board_invite_html(invitee_name, board_name, inviter_name, role):
    role_he = BROLE_HE.get(role, role)
    base = os.environ.get("APP_BASE_URL", "https://code-cal-missions.vercel.app")
    return f"""<div dir="rtl" style="font-family:Arial,Helvetica,sans-serif;max-width:520px;margin:0 auto;color:#1a1a2e">
      <div style="background:#0073ea;color:#fff;padding:18px 22px;border-radius:12px 12px 0 0">
        <h2 style="margin:0;font-size:20px">הוזמנת ללוח ב-CityOS</h2>
      </div>
      <div style="border:1px solid #e6e9ef;border-top:none;border-radius:0 0 12px 12px;padding:22px">
        <p style="font-size:15px">שלום {invitee_name},</p>
        <p style="font-size:15px"><b>{inviter_name}</b> הזמין/ה אותך ללוח <b>{board_name}</b> במערכת CityOS של עיריית הוד השרון.</p>
        <table style="width:100%;font-size:14px;border-collapse:collapse;margin:16px 0">
          <tr><td style="padding:8px 0;color:#676879">שם הלוח</td><td style="padding:8px 0;font-weight:600">{board_name}</td></tr>
          <tr><td style="padding:8px 0;color:#676879">ההרשאה שלך</td><td style="padding:8px 0;font-weight:600">{role_he}</td></tr>
          <tr><td style="padding:8px 0;color:#676879">הוזמנת על ידי</td><td style="padding:8px 0;font-weight:600">{inviter_name}</td></tr>
        </table>
        <a href="{base}" style="display:inline-block;background:#0073ea;color:#fff;text-decoration:none;padding:12px 26px;border-radius:8px;font-weight:700;font-size:15px">כניסה למערכת</a>
        <p style="color:#9699a6;font-size:12px;margin-top:22px">מייל זה נשלח אוטומטית ממערכת CityOS. אם לא ציפית לו, אפשר להתעלם ממנו.</p>
      </div>
    </div>"""

@app.get("/api/boards/{board_id}/members")
def list_board_members(board_id: int):
    with Session(engine) as db:
        ms = db.query(BoardMember).filter(BoardMember.board_id == board_id).all()
        member_ids = {m.user_id for m in ms}
        out = []
        for m in ms:
            u = db.query(User).filter(User.id == m.user_id).first()
            out.append({"user_id": m.user_id, "name": u.name if u else "—",
                        "email": u.email if u else "", "role": m.role})
        # users who could still be invited
        available = [{"id": u.id, "name": u.name} for u in db.query(User).all() if u.id not in member_ids]
        return {"members": out, "available": available}

@app.post("/api/boards/{board_id}/members")
def add_board_member(board_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        if not _is_board_admin(db, board_id, actor_id):
            raise HTTPException(403, "רק מנהל הלוח יכול לנהל הרשאות")
        uid = data.get("user_id")
        role = data.get("role", "editor")
        if role not in BOARD_ROLES:
            role = "editor"
        m = db.query(BoardMember).filter(BoardMember.board_id == board_id, BoardMember.user_id == uid).first()
        is_new = m is None
        if m:
            m.role = role
        else:
            db.add(BoardMember(board_id=board_id, user_id=uid, role=role))
        board = db.query(Board).filter(Board.id == board_id).first()
        inviter = db.query(User).filter(User.id == actor_id).first()
        # in-app notification to the added user — every role, only on a real add,
        # and only if the board hasn't switched notifications off
        if is_new and _board_notify_on(board):
            _notify(db, uid, "board_add", f"נוספת ללוח '{board.name if board else ''}'",
                    f"{(inviter.name + ' ') if inviter else ''}הוסיף/ה אותך ללוח בתפקיד {BROLE_HE.get(role, role)}.",
                    board_id=board_id)
        db.commit()
        # email the invitee — only on a genuine new invitation, any role
        if is_new:
            invitee = db.query(User).filter(User.id == uid).first()
            if invitee and invitee.email and board:
                _send_email(invitee.email,
                            f"הוזמנת ללוח '{board.name}' ב-CityOS",
                            _board_invite_html(invitee.name, board.name,
                                               inviter.name if inviter else "מנהל הלוח", role))
        return {"status": "ok", "invited": is_new}

@app.patch("/api/boards/{board_id}/members/{uid}")
def update_board_member(board_id: int, uid: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        if not _is_board_admin(db, board_id, actor_id):
            raise HTTPException(403, "רק מנהל הלוח יכול לנהל הרשאות")
        m = db.query(BoardMember).filter(BoardMember.board_id == board_id, BoardMember.user_id == uid).first()
        if not m:
            raise HTTPException(404, "member not found")
        role = data.get("role")
        if role in BOARD_ROLES:
            # don't allow demoting the last admin
            if m.role == "admin" and role != "admin":
                admins = db.query(BoardMember).filter(BoardMember.board_id == board_id, BoardMember.role == "admin").count()
                if admins <= 1:
                    raise HTTPException(400, "חייב להישאר לפחות מנהל אחד ללוח")
            m.role = role
        db.commit()
        return {"status": "ok"}

@app.delete("/api/boards/{board_id}/members/{uid}")
def remove_board_member(board_id: int, uid: int, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        if not _is_board_admin(db, board_id, actor_id):
            raise HTTPException(403, "רק מנהל הלוח יכול לנהל הרשאות")
        m = db.query(BoardMember).filter(BoardMember.board_id == board_id, BoardMember.user_id == uid).first()
        if not m:
            return {"status": "removed"}
        if m.role == "admin":
            admins = db.query(BoardMember).filter(BoardMember.board_id == board_id, BoardMember.role == "admin").count()
            if admins <= 1:
                raise HTTPException(400, "לא ניתן להסיר את מנהל הלוח האחרון")
        db.delete(m)
        board = db.query(Board).filter(Board.id == board_id).first()
        actor = db.query(User).filter(User.id == actor_id).first()
        # notify the removed user (any role) — no board deep-link (no longer a member)
        if _board_notify_on(board):
            _notify(db, uid, "board_remove", f"הוסרת מהלוח '{board.name if board else ''}'",
                    f"{(actor.name + ' ') if actor else ''}הסיר/ה אותך מהלוח.")
        db.commit()
        return {"status": "removed"}

# ── In-app notifications (התראות) ───────────────────────────────────
def _serialize_notif(n):
    return {"id": n.id, "type": n.type, "title": n.title, "body": n.body,
            "board_id": n.board_id, "task_id": n.task_id,
            "is_read": bool(n.is_read), "created_at": n.created_at}

@app.get("/api/notifications")
def list_notifications(user_id: int = Depends(current_user_id), limit: int = 30):
    if user_id is None:
        raise HTTPException(400, "user_id required")
    with Session(engine) as db:
        q = db.query(Notification).filter(Notification.user_id == user_id)
        unread = q.filter(Notification.is_read == False).count()
        items = q.order_by(Notification.created_at.desc()).limit(max(1, min(100, limit))).all()
        return {"notifications": [_serialize_notif(n) for n in items], "unread_count": unread}

@app.post("/api/notifications/{nid}/read")
def read_notification(nid: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        n = db.query(Notification).filter(Notification.id == nid,
                                          Notification.user_id == actor_id).first()
        if n and not n.is_read:
            n.is_read = True
            db.commit()
        return {"status": "ok"}

@app.post("/api/notifications/read-all")
def read_all_notifications(data: dict, actor_id: int = Depends(current_user_id)):
    uid = actor_id  # you may only clear your own notifications
    with Session(engine) as db:
        db.query(Notification).filter(Notification.user_id == uid,
                                      Notification.is_read == False).update({Notification.is_read: True})
        db.commit()
        return {"status": "ok"}

# ── ייצוא הלוח לאקסל ────────────────────────────────────────────────
# The client renders the board; the export therefore ships the grid it is
# already showing (current filters, sort, column structure and the values it
# computes — formulas, per-column status labels, linked-item titles) and this
# endpoint only turns that grid into a real .xlsx. The caller can never see
# more than the board API already handed it, so no data is widened here; the
# limits below just keep a malformed payload from eating memory.
XLSX_MAX_ROWS, XLSX_MAX_COLS, XLSX_MAX_TEXT = 20000, 80, 4000
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

def _xlsx_argb(hex_color):
    """#rrggbb → opaque ARGB. Without the explicit FF alpha openpyxl writes 00,
    which some viewers read as a fully transparent fill."""
    return "FF" + hex_color[1:].upper()

def _xlsx_luma(hex_color):
    """Perceived brightness of #rrggbb (0–255) — picks black or white text."""
    try:
        r, g, b = (int(hex_color[i:i + 2], 16) for i in (1, 3, 5))
    except (ValueError, IndexError):
        return 255
    return 0.299 * r + 0.587 * g + 0.114 * b

def _xlsx_cell(ws, row, col, spec):
    """Write one exported cell. `spec` is a scalar, or {v,c,t} with an optional
    background colour and type hint. Text is always written as text — a value
    starting with '=' must never become a live formula in the exported file."""
    from openpyxl.styles import Alignment, Font, PatternFill

    val, color, kind = spec, None, None
    if isinstance(spec, dict):
        val, color, kind = spec.get("v"), spec.get("c"), spec.get("t")
    cell = ws.cell(row=row, column=col)
    if val is None or val == "":
        cell.value = None
    elif kind == "date":
        try:
            cell.value = datetime.fromisoformat(str(val).replace("Z", "+00:00")).replace(tzinfo=None)
            cell.number_format = "DD/MM/YYYY"
        except ValueError:
            cell.value = str(val)[:XLSX_MAX_TEXT]
            cell.data_type = "s"
    elif isinstance(val, bool):
        cell.value = "כן" if val else ""
        cell.data_type = "s"
    elif isinstance(val, (int, float)):
        cell.value = val
    else:
        cell.value = str(val)[:XLSX_MAX_TEXT]
        cell.data_type = "s"          # never a formula, whatever the text is
    if color and _valid_hex(color):
        cell.fill = PatternFill("solid", fgColor=_xlsx_argb(color))
        # one regular weight everywhere, as in the app — colour carries the meaning
        cell.font = Font(color="FFFFFFFF" if _xlsx_luma(color) < 150 else "FF1F2937")
    cell.alignment = Alignment(vertical="center", wrap_text=False)
    return cell

@app.post("/api/boards/{board_id}/export/xlsx")
def export_board_xlsx(board_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    """Board → .xlsx: one sheet, sub-items nested (and collapsible) under their
    item, groups as headers. See the note above on where the grid comes from."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    with Session(engine) as db:
        b = db.query(Board).filter(Board.id == board_id).first()
        if not b:
            raise HTTPException(404, "board not found")
        # export is a full read of the board — require membership unconditionally
        # (this used to skip the check when the caller omitted user_id)
        if _board_role(db, board_id, actor_id) is None:
            raise HTTPException(403, "אין לך גישה ללוח זה")
        board_color = b.color if _valid_hex(b.color) else "#0073ea"

    headers = [str(h)[:XLSX_MAX_TEXT] for h in (data.get("columns") or [])][:XLSX_MAX_COLS]
    rows = (data.get("rows") or [])[:XLSX_MAX_ROWS]
    if not headers:
        raise HTTPException(400, "אין עמודות לייצוא")

    wb = Workbook()
    ws = wb.active
    ws.title = (re.sub(r"[\\/*?:\[\]]", " ", b.name or "לוח").strip() or "לוח")[:31]
    ws.sheet_view.rightToLeft = True
    # the parent row sits ABOVE its detail rows, so Excel's collapse arrows line
    # up with the board: a group folds its items, an item folds its sub-items
    ws.sheet_properties.outlinePr.summaryBelow = False

    head_fill = PatternFill("solid", fgColor=_xlsx_argb(board_color))
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.data_type = "s"
        c.font = Font(color="FFFFFFFF")     # regular weight; the fill marks the header
        c.fill = head_fill
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    widths = [len(h) + 2 for h in headers]
    r = 1
    for row in rows:
        if not isinstance(row, dict):
            continue
        r += 1
        level = row.get("level")
        level = level if isinstance(level, int) and 0 <= level <= 2 else 0
        cells = (row.get("cells") or [])[:len(headers)]
        for i, spec in enumerate(cells, start=1):
            cell = _xlsx_cell(ws, r, i, spec)
            if row.get("group"):
                cell.font = Font(color=_xlsx_argb(board_color))   # colour, not weight
            if i == 1 and level:
                cell.alignment = Alignment(vertical="center", indent=level * 2)
            text = "" if cell.value is None else str(cell.value)
            widths[i - 1] = max(widths[i - 1], min(len(text) + 2 + level * 2, 60))
        if level:
            ws.row_dimensions[r].outlineLevel = level
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    # "<שם הלוח>_<מספר הלוח>.xlsx" — the number keeps exports of same-named
    # boards apart, and says which board a file on disk came from
    fname = re.sub(r"[^\w֐-׿ .()-]", "", (b.name or "board")).strip() or "board"
    quoted = urllib.parse.quote(f"{fname}_{board_id}.xlsx")
    return Response(content=buf.getvalue(), media_type=XLSX_MIME,
                    headers={"Content-Disposition": f"attachment; filename=board_{board_id}.xlsx; "
                                                    f"filename*=UTF-8''{quoted}"})

# ── Column-level permissions (הרשאות עמודה) ─────────────────────────
@app.post("/api/boards/{board_id}/columns/{col_id}/permissions")
def set_col_permissions(board_id: int, col_id: str, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        if _board_role(db, board_id, actor_id) != "admin":
            raise HTTPException(403, "רק מנהל לוח יכול לקבוע הרשאות עמודה")
        b = db.query(Board).filter(Board.id == board_id).first()
        if not b:
            raise HTTPException(404, "board not found")
        s = dict(b.settings or {})
        cols = [dict(c) for c in s.get("columns", [])]
        found = False
        for c in cols:
            if c["id"] == col_id:
                perms = dict(c.get("perms") or {})
                uid = str(data.get("user_id")); perm = data.get("perm")
                if perm in ("view", "edit", "none"):
                    perms[uid] = perm
                else:
                    perms.pop(uid, None)
                c["perms"] = perms; found = True
        if not found:
            raise HTTPException(404, "column not found")
        s["columns"] = cols; b.settings = s; db.commit()
        return {"columns": cols}

# ── Environment / workspace members (הזמנה לסביבה) ──────────────────
WS_ROLES = ("admin", "member", "viewer")

@app.get("/api/workspace/members")
def workspace_members():
    with Session(engine) as db:
        ms = db.query(WorkspaceMember).all()
        member_ids = {m.user_id for m in ms}
        out = []
        for m in ms:
            u = db.query(User).filter(User.id == m.user_id).first()
            out.append({"user_id": m.user_id, "name": u.name if u else "—",
                        "email": u.email if u else "", "role": m.role})
        available = [{"id": u.id, "name": u.name} for u in db.query(User).all() if u.id not in member_ids]
        return {"members": out, "available": available}

@app.post("/api/workspace/members")
def workspace_add(data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        if _ws_role(db, actor_id) != "admin":
            raise HTTPException(403, "רק מנהל סביבה יכול להזמין")
        uid = data.get("user_id"); role = data.get("role", "member")
        if role not in WS_ROLES:
            role = "member"
        m = db.query(WorkspaceMember).filter(WorkspaceMember.user_id == uid).first()
        if m:
            m.role = role
        else:
            db.add(WorkspaceMember(user_id=uid, role=role))
        db.commit()
        return {"status": "ok"}

@app.patch("/api/workspace/members/{uid}")
def workspace_update(uid: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        if _ws_role(db, actor_id) != "admin":
            raise HTTPException(403, "רק מנהל סביבה")
        m = db.query(WorkspaceMember).filter(WorkspaceMember.user_id == uid).first()
        if not m:
            raise HTTPException(404, "member not found")
        role = data.get("role")
        if role in WS_ROLES:
            if m.role == "admin" and role != "admin":
                if db.query(WorkspaceMember).filter(WorkspaceMember.role == "admin").count() <= 1:
                    raise HTTPException(400, "חייב להישאר מנהל סביבה אחד")
            m.role = role
        db.commit()
        return {"status": "ok"}

@app.delete("/api/workspace/members/{uid}")
def workspace_remove(uid: int, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        if _ws_role(db, actor_id) != "admin":
            raise HTTPException(403, "רק מנהל סביבה")
        m = db.query(WorkspaceMember).filter(WorkspaceMember.user_id == uid).first()
        if m and m.role == "admin":
            if db.query(WorkspaceMember).filter(WorkspaceMember.role == "admin").count() <= 1:
                raise HTTPException(400, "לא ניתן להסיר את מנהל הסביבה האחרון")
        if m:
            db.delete(m); db.commit()
        return {"status": "removed"}

# ── Environments (Monday-style workspaces) ───────────────────────────
# Reads are open to all; create/edit/delete are restricted to system (workspace) admins.
def _env_role(db, env_id, user_id):
    """A user's role within one environment: 'admin' for a system (workspace) admin
    (manages every environment), else the per-environment role (manager|member), or None."""
    if user_id is None:
        return None
    if _ws_role(db, user_id) == "admin":
        return "admin"
    m = (db.query(EnvironmentMember)
         .filter(EnvironmentMember.environment_id == env_id, EnvironmentMember.user_id == user_id).first())
    return (m.role or "member") if m else None

def _can_manage_env(db, env_id, user_id):
    """System admins and environment managers may manage an environment."""
    return _env_role(db, env_id, user_id) in ("admin", "manager")

def _env_out(e, board_count=0, my_role=None):
    return {"id": e.id, "name": e.name, "icon": e.icon, "color": e.color,
            "position": e.position, "is_primary": bool(e.is_primary), "board_count": board_count,
            "my_role": my_role, "can_manage": my_role in ("admin", "manager")}

@app.get("/api/environments")
def list_environments(user_id: int = Depends(current_user_id)):
    """System admins see every environment; managers/members see only the
    environments they were granted access to."""
    with Session(engine) as db:
        counts = dict(db.query(Board.environment_id, func.count(Board.id))
                        .group_by(Board.environment_id).all())
        # Environments always read in alphabetical (א-ב) order by name — including
        # newly created ones. Sorted in Python so SQLite (local) and Postgres (prod)
        # order Hebrew identically, regardless of the DB's collation.
        allenvs = db.query(Environment).all()
        allenvs.sort(key=lambda e: (e.name or "").strip().casefold())
        is_sysadmin = _ws_role(db, user_id) == "admin"
        roles = {}
        if user_id is not None:
            roles = {m.environment_id: (m.role or "member") for m in
                     db.query(EnvironmentMember).filter(EnvironmentMember.user_id == user_id).all()}
        if is_sysadmin:
            envs = allenvs
        elif user_id is not None:
            envs = [e for e in allenvs if e.id in roles]
        else:
            envs = []
        def _role_for(e):
            return "admin" if is_sysadmin else roles.get(e.id)
        return {"environments": [_env_out(e, counts.get(e.id, 0), _role_for(e)) for e in envs]}

@app.post("/api/environments")
def create_environment(data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        if not _cap(db, actor_id, "create_environment"):
            raise HTTPException(403, "אין לך הרשאה ליצור סביבה")
        name = (data.get("name") or "").strip()
        if not name:
            raise HTTPException(400, "נדרש שם לסביבה")
        org = db.query(Organization).first()
        pos = db.query(Environment).count()
        e = Environment(name=name, icon=data.get("icon") or "🏢",
                        color=data.get("color") or _ENV_COLORS[pos % len(_ENV_COLORS)],
                        position=pos, organization_id=org.id if org else None)
        db.add(e); db.commit(); db.refresh(e)
        return _env_out(e, 0)

@app.patch("/api/environments/{env_id}")
def update_environment(env_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        actor = actor_id
        if not _can_manage_env(db, env_id, actor):
            raise HTTPException(403, "רק מנהל מערכת או מנהל הסביבה יכול לערוך אותה")
        e = db.query(Environment).filter(Environment.id == env_id).first()
        if not e:
            raise HTTPException(404, "סביבה לא נמצאה")
        if "name" in data and (data.get("name") or "").strip():
            e.name = data["name"].strip()
        if data.get("icon"):
            e.icon = data["icon"]
        if data.get("color"):
            e.color = data["color"]
        db.commit()
        return _env_out(e, my_role=_env_role(db, env_id, actor))

@app.delete("/api/environments/{env_id}")
def delete_environment(env_id: int, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        if _ws_role(db, actor_id) != "admin":
            raise HTTPException(403, "רק מנהל מערכת יכול למחוק סביבה")
        e = db.query(Environment).filter(Environment.id == env_id).first()
        if not e:
            raise HTTPException(404, "סביבה לא נמצאה")
        if e.is_primary:
            raise HTTPException(400, "לא ניתן למחוק את סביבת ברירת המחדל")
        # move boards to the primary workspace (don't delete them) so nothing disappears
        primary = db.query(Environment).filter(Environment.is_primary == True).first()
        db.query(Board).filter(Board.environment_id == env_id).update(
            {Board.environment_id: primary.id if primary else None})
        db.query(EnvironmentMember).filter(EnvironmentMember.environment_id == env_id).delete()
        db.delete(e); db.commit()
        return {"status": "deleted"}

ENV_ROLES = ("manager", "member")
ENV_ROLE_HE = {"admin": "מנהל מערכת", "manager": "מנהל סביבה", "member": "חבר"}

# ── Environment members — system admins and environment managers ──
@app.get("/api/environments/{env_id}/members")
def environment_members(env_id: int, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        if not _can_manage_env(db, env_id, actor_id):
            raise HTTPException(403, "רק מנהל מערכת או מנהל הסביבה")
        if not db.query(Environment).filter(Environment.id == env_id).first():
            raise HTTPException(404, "סביבה לא נמצאה")
        rows = db.query(EnvironmentMember).filter(EnvironmentMember.environment_id == env_id).all()
        users = {u.id: u for u in db.query(User).all()}
        member_ids = {m.user_id for m in rows}
        members = [{"user_id": m.user_id, "name": users[m.user_id].name,
                    "email": users[m.user_id].email,
                    "avatar_url": users[m.user_id].avatar_url, "role": m.role or "member",
                    "role_he": ENV_ROLE_HE.get(m.role or "member", m.role)}
                   for m in rows if m.user_id in users]
        members.sort(key=lambda x: (x["role"] != "manager", x["name"]))
        available = [{"id": u.id, "name": u.name} for u in users.values() if u.id not in member_ids]
        # only a system admin may hand out the manager role; a plain env manager can add members
        can_set_manager = _ws_role(db, actor_id) == "admin"
        return {"members": members, "available": available, "can_set_manager": can_set_manager}

@app.post("/api/environments/{env_id}/members")
def environment_add_member(env_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        actor = actor_id
        if not _can_manage_env(db, env_id, actor):
            raise HTTPException(403, "רק מנהל מערכת או מנהל הסביבה")
        if not db.query(Environment).filter(Environment.id == env_id).first():
            raise HTTPException(404, "סביבה לא נמצאה")
        uid = data.get("user_id")
        role = data.get("role") if data.get("role") in ENV_ROLES else "member"
        # only a system admin may grant the manager role
        if role == "manager" and _ws_role(db, actor) != "admin":
            role = "member"
        if not uid:
            return {"status": "ok"}
        m = db.query(EnvironmentMember).filter(
            EnvironmentMember.environment_id == env_id, EnvironmentMember.user_id == uid).first()
        if m:
            m.role = role
        else:
            db.add(EnvironmentMember(environment_id=env_id, user_id=uid, role=role))
        db.commit()
        return {"status": "ok"}

@app.patch("/api/environments/{env_id}/members/{uid}")
def environment_set_member_role(env_id: int, uid: int, data: dict, actor_id: int = Depends(current_user_id)):
    """Promote/demote an environment member (manager|member). System admin only —
    handing out the manager role is a system-level grant."""
    with Session(engine) as db:
        if _ws_role(db, actor_id) != "admin":
            raise HTTPException(403, "רק מנהל מערכת יכול לשנות תפקיד בסביבה")
        role = data.get("role") if data.get("role") in ENV_ROLES else "member"
        m = db.query(EnvironmentMember).filter(
            EnvironmentMember.environment_id == env_id, EnvironmentMember.user_id == uid).first()
        if not m:
            raise HTTPException(404, "החבר לא נמצא בסביבה")
        m.role = role
        db.commit()
        return {"status": "ok", "role": role}

@app.delete("/api/environments/{env_id}/members/{uid}")
def environment_remove_member(env_id: int, uid: int, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        if not _can_manage_env(db, env_id, actor_id):
            raise HTTPException(403, "רק מנהל מערכת או מנהל הסביבה")
        # a user may be removed from an environment only if they don't belong to any
        # board within it — otherwise they'd lose access to boards they're part of
        env_boards = {b.id: b.name for b in db.query(Board).filter(
            Board.environment_id == env_id, Board.is_archived == False).all()}
        if env_boards:
            linked = (db.query(BoardMember)
                      .filter(BoardMember.user_id == uid,
                              BoardMember.board_id.in_(list(env_boards.keys()))).all())
            names = [env_boards[bm.board_id] for bm in linked if bm.board_id in env_boards]
            if names:
                raise HTTPException(400, "לא ניתן להסיר משתתף זה מאחר והוא משוייך ללוחות הבאים בסביבה: "
                                    + ", ".join(names))
        m = db.query(EnvironmentMember).filter(
            EnvironmentMember.environment_id == env_id, EnvironmentMember.user_id == uid).first()
        if m:
            db.delete(m); db.commit()
        return {"status": "removed"}

# ── Folders (group boards inside an environment) — env managers + sysadmins ──
def _folder_out(db, f, user_id):
    return {"id": f.id, "environment_id": f.environment_id, "name": f.name,
            "position": f.position or 0,
            "can_manage": _can_manage_env(db, f.environment_id, user_id)}

@app.get("/api/folders")
def list_folders(user_id: int = Depends(current_user_id)):
    """Folders across every environment the user can access."""
    with Session(engine) as db:
        folders = db.query(Folder).order_by(Folder.position, Folder.id).all()
        if user_id is not None and _ws_role(db, user_id) != "admin":
            env_ids = {m.environment_id for m in
                       db.query(EnvironmentMember).filter(EnvironmentMember.user_id == user_id).all()}
            folders = [f for f in folders if f.environment_id in env_ids]
        return {"folders": [_folder_out(db, f, user_id) for f in folders]}

@app.post("/api/folders")
def create_folder(data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        actor = actor_id
        env_id = data.get("environment_id")
        if not _can_manage_env(db, env_id, actor):
            raise HTTPException(403, "רק מנהל מערכת או מנהל הסביבה יכול ליצור תיקייה")
        if not db.query(Environment).filter(Environment.id == env_id).first():
            raise HTTPException(404, "סביבה לא נמצאה")
        name = (data.get("name") or "").strip() or "תיקייה חדשה"
        maxpos = db.query(func.max(Folder.position)).filter(Folder.environment_id == env_id).scalar()
        f = Folder(environment_id=env_id, name=name, position=(maxpos + 1) if maxpos is not None else 0)
        db.add(f); db.commit(); db.refresh(f)
        return _folder_out(db, f, actor)

@app.patch("/api/folders/{folder_id}")
def update_folder(folder_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        f = db.query(Folder).filter(Folder.id == folder_id).first()
        if not f:
            raise HTTPException(404, "תיקייה לא נמצאה")
        if not _can_manage_env(db, f.environment_id, actor_id):
            raise HTTPException(403, "רק מנהל מערכת או מנהל הסביבה")
        if (data.get("name") or "").strip():
            f.name = data["name"].strip()
        db.commit()
        return _folder_out(db, f, actor_id)

@app.delete("/api/folders/{folder_id}")
def delete_folder(folder_id: int, actor_id: int = Depends(current_user_id)):
    """Delete a folder. Its boards are detached to the environment root — never deleted."""
    with Session(engine) as db:
        f = db.query(Folder).filter(Folder.id == folder_id).first()
        if not f:
            raise HTTPException(404, "תיקייה לא נמצאה")
        if not _can_manage_env(db, f.environment_id, actor_id):
            raise HTTPException(403, "רק מנהל מערכת או מנהל הסביבה")
        db.query(Board).filter(Board.folder_id == folder_id).update({Board.folder_id: None})
        db.delete(f); db.commit()
        return {"status": "deleted", "id": folder_id}

@app.post("/api/environments/{env_id}/reorder")
def reorder_environment(env_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    """Persist the ordering of folders and boards in an environment. Body:
    {folders:[{id,position}], boards:[{id,folder_id,position}]}. Only items that
    belong to this environment are touched. Env managers / sysadmins only."""
    with Session(engine) as db:
        if not _can_manage_env(db, env_id, actor_id):
            raise HTTPException(403, "רק מנהל מערכת או מנהל הסביבה יכול לסדר")
        env_folder_ids = {f.id for f in db.query(Folder).filter(Folder.environment_id == env_id).all()}
        for row in (data.get("folders") or []):
            f = db.query(Folder).filter(Folder.id == row.get("id"),
                                        Folder.environment_id == env_id).first()
            if f:
                f.position = int(row.get("position", 0))
        for row in (data.get("boards") or []):
            b = db.query(Board).filter(Board.id == row.get("id"),
                                       Board.environment_id == env_id).first()
            if not b:
                continue
            b.position = int(row.get("position", 0))
            fid = row.get("folder_id")
            # only allow assigning to a folder that belongs to this environment
            b.folder_id = fid if (fid in env_folder_ids) else None
        db.commit()
        return {"status": "ok"}

@app.get("/api/tasks")
def list_tasks(board_id: Optional[int] = None, status: Optional[str] = None):
    with Session(engine) as db:
        q = db.query(Task).filter(Task.is_archived == False)
        if board_id:
            q = q.filter(Task.board_id == board_id)
        if status:
            q = q.filter(Task.status == status)
        tasks = q.order_by(Task.position).limit(100).all()
        result = []
        for t in tasks:
            assignees = [{"id": u.id, "name": u.name} for u in (t.assignees or [])]
            result.append({
                "id": t.id, "board_id": t.board_id, "group_id": t.group_id,
                "title": t.title, "description": t.description,
                "status": t.status.value if hasattr(t.status, 'value') else t.status,
                "priority": t.priority.value if hasattr(t.priority, 'value') else t.priority,
                "due_date": t.due_date,
                "location_lat": t.location_lat, "location_lng": t.location_lng,
                "tags": t.tags or [],
                "assignees": assignees,
            })
        return result

@app.post("/api/boards/{board_id}/environment")
def move_board_to_environment(board_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    """Move a board to another environment (workspace).

    Requires managing BOTH ends: you may not push a board into a workspace you
    do not run, nor pull one out of a workspace you do not run. A board admin
    who manages neither cannot move it. The board keeps its members, items and
    settings; it only leaves its folder, which belongs to the old environment.
    """
    with Session(engine) as db:
        actor = actor_id
        b = db.query(Board).filter(Board.id == board_id).first()
        if not b:
            raise HTTPException(404, "לוח לא נמצא")
        env_id = data.get("environment_id")
        target = db.query(Environment).filter(Environment.id == env_id).first()
        if not target:
            raise HTTPException(404, "סביבה לא נמצאה")
        if b.environment_id == target.id:
            raise HTTPException(400, "הלוח כבר נמצא בסביבה הזו")
        if not _can_manage_env(db, target.id, actor):
            raise HTTPException(403, "רק מנהל מערכת או מנהל הסביבה יכול להעביר לוח לסביבה זו")
        if b.environment_id and not _can_manage_env(db, b.environment_id, actor):
            raise HTTPException(403, "רק מנהל מערכת או מנהל הסביבה הנוכחית יכול להוציא ממנה לוח")
        maxpos = db.query(func.max(Board.position)).filter(Board.environment_id == target.id).scalar()
        b.environment_id = target.id
        b.folder_id = None            # folders belong to the environment it just left
        b.position = (maxpos + 1) if maxpos is not None else 0
        db.commit()
        return {"status": "moved", "board_id": b.id,
                "environment_id": target.id, "environment_name": target.name}

@app.post("/api/tasks")
def create_task(data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        parent_id = data.get("parent_id")
        board_id = data.get("board_id")
        group_id = data.get("group_id")
        if parent_id:
            parent = db.query(Task).filter(Task.id == parent_id).first()
            if not parent:
                raise HTTPException(404, "parent task not found")
            # sub-items inherit board (and group if not given) from their parent
            board_id = parent.board_id
            if group_id is None:
                group_id = parent.group_id
        # a task must belong to a board (directly or via its parent) — reject
        # orphan tasks instead of silently creating board-less junk data
        if not board_id:
            raise HTTPException(422, "board_id (or a valid parent_id) is required")
        board = db.query(Board).filter(Board.id == board_id).first()
        if not board:
            raise HTTPException(404, "board not found")
        # a board-defined priority is not an enum value — resolve it to the one
        # it behaves as, and carry its label/colour on the item
        prio, prio_cf = _resolve_priority(board, data.get("priority"))
        task = Task(
            board_id=board_id,
            group_id=group_id,
            parent_id=parent_id,
            title=data.get("title", "Untitled"),
            description=data.get("description", ""),
            priority=prio or "medium",
            tags=data.get("tags", []),
            location_lat=data.get("location_lat"),
            location_lng=data.get("location_lng"),
            address=data.get("address"),
            created_by=actor_id,  # record the creator
        )
        due = data.get("due_date")
        if due:
            try:
                task.due_date = datetime.fromisoformat(str(due).replace("Z", "+00:00"))
            except ValueError:
                pass
        status = data.get("status")
        if status:
            try:
                task.status = TaskStatus(status)
            except ValueError:
                pass
        ids = data.get("assignee_ids") or []
        if ids:
            task.assignees = db.query(User).filter(User.id.in_(ids)).all()
        # Boards created before "יוצר הרשומה" became a derived column carry it as
        # a plain people column, whose value has to be written on creation. The
        # derived `created_by` column needs nothing — it reads task.created_by.
        creator = actor_id
        cf = dict(data.get("custom_fields") or {})
        # A new item has no status yet — it shows "טרם הוגדר" until someone picks
        # one. task.status still holds the group's value so grouping, kanban and
        # the charts behave normally; this only says a human has not chosen. It
        # is cleared the moment a status is set. A caller that names a status on
        # creation (the form, a move) is choosing one, so the flag is not set.
        if not data.get("status"):
            cf.setdefault("status_unset", True)
        # and the same for the priority column, for the same reason: task.priority
        # keeps its "medium" default so sorting and the charts behave, while the
        # cell reads "טרם הוגדר" until someone actually picks a priority.
        if not data.get("priority"):
            cf.setdefault("priority_unset", True)
        elif prio_cf:
            for k, v in prio_cf.items():
                if v is None:
                    cf.pop(k, None)
                else:
                    cf.setdefault(k, v)
        if creator:
            cols = (board.settings or {}).get("columns", []) if board else []
            for c in cols:
                if c.get("id") == "sys_created_by" and c.get("type") == "people":
                    cf.setdefault("sys_created_by", [creator])
        if cf:
            task.custom_fields = cf
        # place the new item last — bottom of its group (or its parent's sub-list)
        from sqlalchemy import func as _func
        if parent_id is not None:
            maxpos = db.query(_func.max(Task.position)).filter(Task.parent_id == parent_id).scalar()
        else:
            _q = db.query(_func.max(Task.position)).filter(Task.board_id == board_id, Task.parent_id == None)
            _q = _q.filter(Task.group_id == group_id) if group_id is not None else _q.filter(Task.group_id == None)
            maxpos = _q.scalar()
        task.position = (maxpos + 1) if maxpos is not None else 0
        db.add(task)
        db.flush()   # get task.id for assignment notifications
        # a brand-new item (or sub-item) gets its own automatic identifier — an
        # item created after another was deleted never reuses that one's number
        _ensure_item_uid(db, task)
        # notify anyone assigned at creation time (skip the creator), honoring the switch
        if ids:
            board = db.query(Board).filter(Board.id == board_id).first()
            if _board_notify_on(board):
                actor = actor_id
                for aid in set(ids):
                    if aid != actor:
                        _notify(db, aid, "assign", "שויכת למשימה",
                                f"שויכת למשימה '{task.title}'.", board_id=board_id, task_id=task.id)
        db.commit()
        db.refresh(task)
        # return the fully-serialized task so the client can insert the new row
        # in place (no full-board refetch/re-render on every add)
        my_role = _board_role(db, board_id, creator)
        cols = (db.query(Board.settings).filter(Board.id == board_id).scalar() or {}).get("columns", [])
        return _serialize_task(task, db, user_id=creator, board_role=my_role, columns=cols)

# ── Cross-board item linking (connect column) ───────────────────────
@app.get("/api/items/search")
def items_search(board_ids: str = "", q: str = "", exclude_task: Optional[int] = None, limit: int = 60):
    """Top-level items across the given boards, for the connect-column picker."""
    ids = [int(x) for x in board_ids.split(",") if x.strip().isdigit()]
    with Session(engine) as db:
        query = db.query(Task).filter(Task.is_archived == False, Task.parent_id == None)
        if ids:
            query = query.filter(Task.board_id.in_(ids))
        if q.strip():
            query = query.filter(Task.title.ilike(f"%{q.strip()}%"))
        if exclude_task:
            query = query.filter(Task.id != exclude_task)
        rows = query.order_by(Task.board_id, Task.position).limit(max(1, min(200, limit))).all()
        bnames = {b.id: b.name for b in db.query(Board).all()}
        return {"items": [{"id": t.id, "title": t.title, "board_id": t.board_id,
                           "board_name": bnames.get(t.board_id, ""),
                           "status": t.status.value if hasattr(t.status, "value") else t.status}
                          for t in rows]}

@app.get("/api/tasks/lookup")
def tasks_lookup(ids: str = ""):
    """Resolve item ids → title/board, to display linked items in a connect cell."""
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    if not id_list:
        return {"items": []}
    with Session(engine) as db:
        rows = db.query(Task).filter(Task.id.in_(id_list)).all()
        bnames = {b.id: b.name for b in db.query(Board).all()}
        return {"items": [{"id": t.id, "title": t.title, "board_id": t.board_id,
                           "board_name": bnames.get(t.board_id, "")} for t in rows]}

def _group_for_status(db, board_id, status_val):
    """The group a top-level item moves to for a status — the one whose own status
    matches exactly, or None to leave the item exactly where it is.

    No fallback by stage or by status family. A "back to development" route (todo →
    the in_progress group) was tried and removed: a board can have more than one
    in_progress group — e.g. "בביצוע" and "roadmap" — and picking one by status
    alone is a guess that lands items in the wrong place.

    A group left at the default `todo` status is treated as a plain section, never
    an auto-move destination. Every new group is created as `todo` (see
    create_group_api), so without this a freshly added group silently becomes a
    magnet: setting a todo-family status like "חזרה לפיתוח" would yank the item out
    of "בביצוע" into that new last group. Skipping todo groups keeps the item put.
    """
    sv = status_val.value if hasattr(status_val, "value") else status_val
    groups = db.query(Group).filter(Group.board_id == board_id).order_by(Group.position).all()
    gs_of = lambda g: (g.task_status.value if hasattr(g.task_status, "value") else g.task_status)
    for g in groups:
        if gs_of(g) == "todo":
            continue
        if gs_of(g) == sv:
            return g
    return None


def _st_of(t):
    return t.status.value if hasattr(t.status, "value") else t.status


def _rollup_parent(db, parent_id):
    """When every sub-item of a parent is done, mark the parent done too and move
    it to the matching group."""
    parent = db.query(Task).filter(Task.id == parent_id).first()
    if not parent:
        return
    subs = db.query(Task).filter(Task.parent_id == parent_id).all()
    if not subs:
        return
    if all(_st_of(s) == "done" for s in subs) and _st_of(parent) != "done":
        parent.status = TaskStatus.DONE
        g = _group_for_status(db, parent.board_id, TaskStatus.DONE)
        if g:
            parent.group_id = g.id


@app.patch("/api/tasks/{task_id}")
def update_task(task_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    """Generic item update — title, priority, status, due_date, and custom column
    values (merged into custom_fields). Used by the editable board columns."""
    with Session(engine) as db:
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            raise HTTPException(404, "task not found")
        actor = actor_id
        # ── permission enforcement (item level, and column level for custom_fields) ──
        # Previously the whole block was skipped when the caller omitted user_id,
        # so an anonymous PATCH could edit any item on any board.
        actor_role = _board_role(db, task.board_id, actor)
        editing_builtin = any(k in data for k in ("title", "description", "priority", "status", "due_date", "tags"))
        iperm = _item_perm(task, actor, actor_role)
        if editing_builtin and iperm not in ("edit", "delete"):
            raise HTTPException(403, "אין לך הרשאת עריכה לפריט זה")
        if "custom_fields" in data and isinstance(data["custom_fields"], dict):
            cols = {c["id"]: c for c in (db.query(Board).filter(Board.id == task.board_id).first().settings or {}).get("columns", [])}
            for k in list(data["custom_fields"].keys()):
                col = cols.get(k)
                if col is not None and _col_perm(col, actor, actor_role) != "edit":
                    del data["custom_fields"][k]      # silently drop cols the user can't edit
            if iperm not in ("edit", "delete"):
                data["custom_fields"] = {}
        st_val = lambda v: (v.value if hasattr(v, "value") else v)
        if data.get("title") is not None and data["title"] != task.title:
            _audit(db, task_id, "update", "title", task.title, data["title"], actor)
            task.title = data["title"]
        if data.get("description") is not None:
            task.description = data["description"]
        if data.get("priority"):
            # a board-defined priority arrives as its own key: store the engine
            # value it behaves as, and let its label/colour ride on the item
            pboard = db.query(Board).filter(Board.id == task.board_id).first()
            pval, pcf = _resolve_priority(pboard, data["priority"])
            try:
                nv = Priority(pval)
                if st_val(task.priority) != st_val(nv):
                    _audit(db, task_id, "update", "priority", st_val(task.priority), st_val(nv), actor)
                task.priority = nv
                if pcf:
                    cfd = data.get("custom_fields")
                    if not isinstance(cfd, dict):
                        cfd = {}
                        data["custom_fields"] = cfd
                    for k, v in pcf.items():
                        cfd.setdefault(k, v)
            except ValueError:
                pass
        if data.get("status"):
            try:
                nv = TaskStatus(data["status"])
                if st_val(task.status) != st_val(nv):
                    _audit(db, task_id, "update", "status", st_val(task.status), st_val(nv), actor)
                task.status = nv
                # auto-move a top-level item only to the group that stands for its
                # exact status; with no such group it stays where it is
                if task.parent_id is None:
                    g = _group_for_status(db, task.board_id, nv)
                    if g:
                        task.group_id = g.id
                else:
                    # roll up: parent auto-completes once all its sub-items are done
                    _rollup_parent(db, task.parent_id)
            except ValueError:
                pass
        if "due_date" in data:
            due = data["due_date"]
            old_due = task.due_date.isoformat() if task.due_date else None
            if due:
                try:
                    task.due_date = datetime.fromisoformat(str(due).replace("Z", "+00:00"))
                except ValueError:
                    pass
            else:
                task.due_date = None
            _audit(db, task_id, "update", "due_date", old_due,
                   task.due_date.isoformat() if task.due_date else None, actor)
        if "tags" in data and isinstance(data["tags"], list):
            old_tags = task.tags or []
            if old_tags != data["tags"]:
                _audit(db, task_id, "update", "tags", ", ".join(old_tags), ", ".join(data["tags"]), actor)
            task.tags = data["tags"]
        if "custom_fields" in data and isinstance(data["custom_fields"], dict):
            board = db.query(Board).filter(Board.id == task.board_id).first()
            coldefs = {c["id"]: c for c in ((board.settings or {}).get("columns", []) if board else [])}
            notify_on = _board_notify_on(board)
            def _ids(val):
                out = set()
                for x in (val or []) if isinstance(val, list) else []:
                    try:
                        out.add(int(x))
                    except (TypeError, ValueError):
                        pass
                return out
            cf = dict(task.custom_fields or {})
            for k, v in data["custom_fields"].items():
                old = cf.get(k)
                if v is None:
                    cf.pop(k, None)
                else:
                    cf[k] = v
                if json.dumps(old, ensure_ascii=False) != json.dumps(v, ensure_ascii=False):
                    _audit(db, task_id, "update", "col:" + k,
                           json.dumps(old, ensure_ascii=False), json.dumps(v, ensure_ascii=False), actor)
                    # people column: notify each user newly added to this cell
                    col = coldefs.get(k)
                    if col and col.get("type") == "people" and notify_on:
                        added = _ids(v) - _ids(old)
                        added.discard(actor)
                        for uid2 in added:
                            _notify(db, uid2, "assign", "שויכת לפריט",
                                    f"שויכת לעמודת '{col.get('title', 'אנשים')}' במשימה '{task.title}'.",
                                    board_id=task.board_id, task_id=task_id)
            task.custom_fields = cf
        db.commit()
        return {"id": task.id, "custom_fields": task.custom_fields or {}}

# ── File uploads (for the Files column) ─────────────────────────────
MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10MB per file.
# NOTE: on Vercel the serverless edge rejects request bodies larger than ~4.5MB
# before they reach us, so the full 10MB only applies on hosts without that cap
# (e.g. Azure Container Apps). Larger files would need external blob storage.

# Only these render in the browser. Anything else is served as a download, so an
# uploaded .html/.svg cannot execute script on our origin and steal the session.
# SVG is deliberately NOT here — it can carry <script> and would run same-origin.
INLINE_SAFE_TYPES = {
    "image/png", "image/jpeg", "image/gif", "image/webp", "image/bmp",
    "application/pdf", "text/plain",
}


def _safe_media_type(declared: str) -> tuple[str, bool]:
    """Normalise a client-declared content type into (media_type, may_inline)."""
    ct = (declared or "application/octet-stream").split(";")[0].strip().lower()
    if ct in INLINE_SAFE_TYPES:
        return ct, True
    # Never echo back an active type — text/html, image/svg+xml and friends
    # become an opaque download instead.
    return "application/octet-stream", False


@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...), actor_id: int = Depends(current_user_id)):
    """Store the upload in the DB so it persists and is shared across serverless
    instances (Vercel's local filesystem is ephemeral and per-instance)."""
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, "הקובץ גדול מדי (מקסימום 10MB)")
    media_type, _ = _safe_media_type(file.content_type)
    token = uuid.uuid4().hex
    with Session(engine) as db:
        db.add(UploadedFile(token=token, name=file.filename or "file",
                            content_type=media_type,
                            data=content, size=len(content)))
        db.commit()
    return {"name": file.filename, "url": f"/api/files/{token}", "type": media_type}

@app.get("/api/files/{token}")
def serve_file(token: str, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        f = db.query(UploadedFile).filter(UploadedFile.token == token).first()
        if not f:
            raise HTTPException(404, "file not found")
        from urllib.parse import quote
        media_type, may_inline = _safe_media_type(f.content_type)
        how = "inline" if may_inline else "attachment"
        disp = f"{how}; filename*=UTF-8''{quote(f.name or 'file')}"
        return Response(content=f.data, media_type=media_type,
                        headers={"Content-Disposition": disp,
                                 # stop the browser sniffing a download back into HTML
                                 "X-Content-Type-Options": "nosniff",
                                 "Content-Security-Policy": "default-src 'none'; sandbox",
                                 # attachments are per-user now — keep them out of shared caches
                                 "Cache-Control": "private, max-age=31536000"})

@app.post("/api/tasks/{task_id}/permissions")
def set_item_permissions(task_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    """Set/clear a per-user permission on a single item. Board managers (admin) only.
    perm ∈ view|edit|delete|none  (none = pass to null to remove the override)."""
    with Session(engine) as db:
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            raise HTTPException(404, "task not found")
        actor = actor_id
        if _board_role(db, task.board_id, actor) != "admin":
            raise HTTPException(403, "רק מנהל לוח יכול לקבוע הרשאות פריט")
        perms = dict(task.permissions or {})
        uid = str(data.get("user_id"))
        perm = data.get("perm")
        if perm in ("view", "edit", "delete", "none"):
            perms[uid] = perm
        else:
            perms.pop(uid, None)   # clear override → falls back to board role
        task.permissions = perms
        db.commit()
        return {"id": task.id, "permissions": task.permissions}

@app.delete("/api/tasks/{task_id}")
def delete_task(task_id: int, user_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            raise HTTPException(404, "task not found")
        if user_id is not None:
            role = _board_role(db, task.board_id, user_id)
            if _item_perm(task, user_id, role) != "delete":
                raise HTTPException(403, "אין לך הרשאת מחיקה לפריט זה")
        # remove sub-items along with the parent
        for s in db.query(Task).filter(Task.parent_id == task_id).all():
            db.delete(s)
        db.delete(task)
        db.commit()
        return {"status": "deleted"}

# ── Item conversation (comments), files & activity log ──────────────

def _acting_user(db, uid):
    """The User row for the authenticated caller. `uid` now always comes from the
    auth dependency — it used to be read from the request body."""
    return db.query(User).filter(User.id == uid).first() if uid is not None else None

def _can_comment(db, board, user):
    """Board-scoped: only members who are not view-only may comment.
    A user who isn't a board member has no access at all."""
    if not user or not board:
        return False
    role = _board_role(db, board.id, user.id)
    if role is None:          # not invited to this board
        return False
    if role == "viewer":      # view-only at the board level
        return False
    if (board.settings or {}).get("view_only"):
        return role == "admin"
    return True

def _audit(db, entity_id, action, field=None, old=None, new=None, user_id=None):
    db.add(AuditLog(entity_type="task", entity_id=entity_id, action=action,
                    field_name=field,
                    old_value=(None if old is None else str(old)),
                    new_value=(None if new is None else str(new)),
                    changed_by=user_id))

def _serialize_comment(c, db):
    u = db.query(User).filter(User.id == c.user_id).first()
    # read receipts: everyone who saw the message except its own author
    seen_ids = [i for i in (c.seen_by or []) if i != c.user_id]
    seen_users = []
    if seen_ids:
        for su in db.query(User).filter(User.id.in_(seen_ids)).all():
            seen_users.append({"id": su.id, "name": su.name, "avatar_url": su.avatar_url})
    return {
        "id": c.id, "task_id": c.task_id, "parent_id": c.parent_id,
        "content": c.content, "attachments": c.attachments or [],
        "mentions": c.mentions or [], "likes": c.likes or [],
        "seen_users": seen_users,
        "user_id": c.user_id, "user_name": u.name if u else "—",
        "user_role": u.role if u else None, "created_at": c.created_at,
    }

@app.get("/api/tasks/{task_id}/comments")
def list_comments(task_id: int):
    with Session(engine) as db:
        cs = db.query(Comment).filter(Comment.task_id == task_id).order_by(Comment.created_at).all()
        items = [_serialize_comment(c, db) for c in cs]
        by_id = {i["id"]: {**i, "replies": []} for i in items}
        roots = []
        for i in items:
            node = by_id[i["id"]]
            if i["parent_id"] and i["parent_id"] in by_id:
                by_id[i["parent_id"]]["replies"].append(node)
            else:
                roots.append(node)
        # newest main comment pinned on top
        roots.sort(key=lambda x: str(x["created_at"]), reverse=True)
        return {"comments": roots, "count": len(items)}

@app.post("/api/tasks/{task_id}/comments")
def add_comment(task_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            raise HTTPException(404, "task not found")
        board = db.query(Board).filter(Board.id == task.board_id).first()
        user = _acting_user(db, actor_id)
        if not _can_comment(db, board, user):
            raise HTTPException(403, "אין הרשאת תגובה (צפייה בלבד)")
        content = (data.get("content") or "").strip()
        atts = data.get("attachments") or []
        if not content and not atts:
            raise HTTPException(400, "תגובה ריקה")
        c = Comment(task_id=task_id, user_id=user.id, parent_id=data.get("parent_id"),
                    content=content, attachments=atts,
                    mentions=data.get("mentions") or [], likes=[], seen_by=[])
        db.add(c)
        _audit(db, task_id, "comment", field="reply" if data.get("parent_id") else "comment",
               new=(content[:120] or "(קובץ)"), user_id=user.id)
        # notify every tagged user (explicit @-picks + any @Name typed in the text),
        # excluding the author, honoring the board's notification switch
        mentioned = set()
        for x in (data.get("mentions") or []):
            try:
                mentioned.add(int(x))
            except (TypeError, ValueError):
                continue
        if content:
            for u in sorted(db.query(User).all(), key=lambda x: -len(x.name or "")):
                if u.name and ("@" + u.name) in content:
                    mentioned.add(u.id)
        mentioned.discard(user.id)
        if mentioned and _board_notify_on(board):
            for mid in mentioned:
                _notify(db, mid, "mention", "תויגת בשיחה",
                        f"{user.name} תייג/ה אותך במשימה '{task.title}'.",
                        board_id=task.board_id, task_id=task_id)
        db.commit(); db.refresh(c)
        return _serialize_comment(c, db)

@app.post("/api/comments/{cid}/like")
def like_comment(cid: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        c = db.query(Comment).filter(Comment.id == cid).first()
        if not c:
            raise HTTPException(404, "comment not found")
        user = _acting_user(db, actor_id)
        if not user:
            raise HTTPException(403, "אין משתמש")
        likes = list(c.likes or [])
        if user.id in likes:
            likes.remove(user.id)
        else:
            likes.append(user.id)
        c.likes = likes
        db.commit()
        return {"id": cid, "likes": likes}

@app.post("/api/tasks/{task_id}/comments/seen")
def mark_comments_seen(task_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    """Mark every message in the item as seen by the acting user (read receipts).
    A user's own messages are skipped — you don't 'see' your own."""
    with Session(engine) as db:
        user = _acting_user(db, actor_id)
        if not user:
            raise HTTPException(403, "אין משתמש")
        changed = False
        for c in db.query(Comment).filter(Comment.task_id == task_id).all():
            if c.user_id == user.id:
                continue
            seen = list(c.seen_by or [])
            if user.id not in seen:
                seen.append(user.id)
                c.seen_by = seen
                changed = True
        if changed:
            db.commit()
        return {"ok": True}

@app.patch("/api/comments/{cid}")
def update_comment(cid: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        c = db.query(Comment).filter(Comment.id == cid).first()
        if not c:
            raise HTTPException(404, "comment not found")
        user = _acting_user(db, actor_id)
        if not user or (c.user_id != user.id and user.role not in ("admin", "manager")):
            raise HTTPException(403, "אין הרשאה לערוך תגובה")

        def _mentioned(content, explicit):
            """Users tagged either explicitly or via '@Name' inside the text."""
            s = set()
            for x in (explicit or []):
                try:
                    s.add(int(x))
                except (TypeError, ValueError):
                    continue
            if content:
                for u in db.query(User).all():
                    if u.name and ("@" + u.name) in content:
                        s.add(u.id)
            return s

        before = _mentioned(c.content, c.mentions)   # who was already tagged
        if "content" in data:
            c.content = data["content"]
        if "attachments" in data:
            c.attachments = data.get("attachments") or []
        if "mentions" in data:
            c.mentions = data.get("mentions") or []
        # notify anyone newly tagged in this edit (not the editor, honoring the switch)
        newly = _mentioned(c.content, c.mentions) - before
        newly.discard(user.id)
        if newly:
            task = db.query(Task).filter(Task.id == c.task_id).first()
            board = db.query(Board).filter(Board.id == task.board_id).first() if task else None
            if task and _board_notify_on(board):
                for mid in newly:
                    _notify(db, mid, "mention", "תויגת בשיחה",
                            f"{user.name} תייג/ה אותך במשימה '{task.title}'.",
                            board_id=task.board_id, task_id=task.id)
        db.commit()
        return _serialize_comment(c, db)

@app.delete("/api/comments/{cid}")
def delete_comment(cid: int, user_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        c = db.query(Comment).filter(Comment.id == cid).first()
        if not c:
            raise HTTPException(404, "comment not found")
        user = db.query(User).filter(User.id == user_id).first()
        if not user or (c.user_id != user.id and user.role not in ("admin", "manager")):
            raise HTTPException(403, "אין הרשאה למחוק תגובה")
        for r in db.query(Comment).filter(Comment.parent_id == cid).all():
            db.delete(r)
        db.delete(c)
        db.commit()
        return {"status": "deleted"}

@app.get("/api/tasks/{task_id}/files")
def task_files(task_id: int):
    with Session(engine) as db:
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            raise HTTPException(404, "task not found")
        out = []
        for c in db.query(Comment).filter(Comment.task_id == task_id).order_by(Comment.created_at.desc()).all():
            u = db.query(User).filter(User.id == c.user_id).first()
            for f in (c.attachments or []):
                out.append({**f, "source": "שיחה", "user_name": u.name if u else None,
                            "user_id": c.user_id, "created_at": c.created_at})
        board = db.query(Board).filter(Board.id == task.board_id).first()
        cols = (board.settings or {}).get("columns", []) if board else []
        cf = task.custom_fields or {}
        for col in [c for c in cols if c.get("type") == "files"]:
            for f in (cf.get(col["id"]) or []):
                out.append({**f, "source": "עמודה: " + (col.get("title") or ""), "user_name": None,
                            "user_id": None, "created_at": task.updated_at})
        return {"files": out}

@app.post("/api/tasks/{task_id}/files/delete")
def delete_task_file(task_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    """Remove a file from the item — from the comment it's attached to or a files
    column — and delete the stored blob. The uploader or a board admin/manager may
    delete a chat attachment; a board editor+/admin may delete a column file."""
    with Session(engine) as db:
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            raise HTTPException(404, "task not found")
        user = _acting_user(db, actor_id)
        if not user:
            raise HTTPException(403, "אין משתמש")
        url = (data.get("url") or "").strip()
        if not url:
            raise HTTPException(400, "חסר קובץ למחיקה")
        role = _board_role(db, task.board_id, user.id)
        is_priv = role == "admin" or user.role in ("admin", "manager")
        removed = False
        # 1) chat attachment: find the owning comment (uploader or admin may delete)
        for c in db.query(Comment).filter(Comment.task_id == task_id).all():
            atts = c.attachments or []
            if any((a or {}).get("url") == url for a in atts):
                if not (is_priv or c.user_id == user.id):
                    raise HTTPException(403, "רק מעלה הקובץ או מנהל הלוח יכול למחוק אותו")
                c.attachments = [a for a in atts if (a or {}).get("url") != url]
                removed = True
        # 2) files-column value (board editor+/admin)
        board = db.query(Board).filter(Board.id == task.board_id).first()
        cols = (board.settings or {}).get("columns", []) if board else []
        cf = dict(task.custom_fields or {})
        for col in [c for c in cols if c.get("type") == "files"]:
            arr = cf.get(col["id"]) or []
            if any((a or {}).get("url") == url for a in arr):
                if not (is_priv or role in ("editor",)):
                    raise HTTPException(403, "אין לך הרשאה למחוק קובץ זה")
                cf[col["id"]] = [a for a in arr if (a or {}).get("url") != url] or None
                task.custom_fields = cf
                removed = True
        # 3) delete the stored blob (best-effort, by token in the url)
        token = url.rsplit("/", 1)[-1]
        uf = db.query(UploadedFile).filter(UploadedFile.token == token).first()
        if uf:
            db.delete(uf)
        db.commit()
        return {"status": "deleted", "removed": removed}

@app.get("/api/tasks/{task_id}/activity")
def task_activity(task_id: int, user_id: int = Depends(current_user_id), date: Optional[str] = None):
    with Session(engine) as db:
        q = db.query(AuditLog).filter(AuditLog.entity_type == "task", AuditLog.entity_id == task_id)
        if user_id:
            q = q.filter(AuditLog.changed_by == user_id)
        out = []
        for l in q.order_by(AuditLog.created_at.desc()).all():
            if date and (not l.created_at or str(l.created_at)[:10] != date):
                continue
            u = db.query(User).filter(User.id == l.changed_by).first()
            out.append({"id": l.id, "action": l.action, "field": l.field_name,
                        "old": l.old_value, "new": l.new_value,
                        "user_id": l.changed_by, "user_name": u.name if u else "מערכת",
                        "can_undo": l.action == "update", "created_at": l.created_at})
        return {"activity": out}

@app.get("/api/tasks/{task_id}/activity/export")
def export_activity(task_id: int):
    with Session(engine) as db:
        logs = db.query(AuditLog).filter(AuditLog.entity_type == "task", AuditLog.entity_id == task_id)\
                 .order_by(AuditLog.created_at.desc()).all()
        buf = io.StringIO()
        buf.write("﻿")  # BOM so Excel reads Hebrew UTF-8
        w = csv.writer(buf)
        w.writerow(["תאריך", "משתמש", "פעולה", "שדה", "מ-", "ל-"])
        for l in logs:
            u = db.query(User).filter(User.id == l.changed_by).first()
            w.writerow([str(l.created_at)[:19], u.name if u else "מערכת", l.action,
                        l.field_name or "", l.old_value or "", l.new_value or ""])
        return Response(content=buf.getvalue(), media_type="text/csv",
                        headers={"Content-Disposition": f"attachment; filename=activity_task_{task_id}.csv"})

def _revert_field(task, field, old):
    if field == "title":
        task.title = old or ""; return True
    if field == "status":
        try: task.status = TaskStatus(old); return True
        except ValueError: return False
    if field == "priority":
        try: task.priority = Priority(old); return True
        except ValueError: return False
    if field == "due_date":
        task.due_date = datetime.fromisoformat(old) if old else None; return True
    if field == "tags":
        task.tags = [x for x in (old or "").split(", ") if x]; return True
    if field and field.startswith("col:"):
        cid = field[4:]; cf = dict(task.custom_fields or {})
        val = json.loads(old) if old not in (None, "", "null") else None
        if val is None: cf.pop(cid, None)
        else: cf[cid] = val
        task.custom_fields = cf; return True
    return False

@app.post("/api/activity/{log_id}/undo")
def undo_activity(log_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        user = _acting_user(db, actor_id)
        if not user or user.role not in ("admin", "manager"):
            raise HTTPException(403, "רק מנהל יכול לבצע ביטול")
        log = db.query(AuditLog).filter(AuditLog.id == log_id).first()
        if not log or log.entity_type != "task" or log.action != "update":
            raise HTTPException(400, "לא ניתן לבטל פעולה זו")
        task = db.query(Task).filter(Task.id == log.entity_id).first()
        if not task:
            raise HTTPException(404, "task not found")
        if not _revert_field(task, log.field_name, log.old_value):
            raise HTTPException(400, "שדה לא נתמך לביטול")
        _audit(db, task.id, "undo", field=log.field_name, old=log.new_value, new=log.old_value, user_id=user.id)
        db.commit()
        return {"status": "reverted", "field": log.field_name}

@app.post("/api/tasks/{task_id}/assignees")
def task_assignees(task_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    """Add or remove a user from a task (monday-style people column)."""
    with Session(engine) as db:
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            raise HTTPException(404, "task not found")
        user = db.query(User).filter(User.id == data.get("user_id")).first()
        if not user:
            raise HTTPException(404, "user not found")
        actor = actor_id
        _require_board_edit(db, task.board_id, actor)   # only board editors may (un)assign
        was = user in task.assignees
        if data.get("action") == "remove":
            if was:
                task.assignees.remove(user)
        else:
            if not was:
                task.assignees.append(user)
                # notify the newly-assigned person (skip self-assignment / re-adds)
                if user.id != actor:
                    board = db.query(Board).filter(Board.id == task.board_id).first()
                    if _board_notify_on(board):
                        actor_u = db.query(User).filter(User.id == actor).first()
                        _notify(db, user.id, "assign", "שויכת למשימה",
                                f"{(actor_u.name + ' ') if actor_u else ''}שייך/ה אותך למשימה '{task.title}'.",
                                board_id=task.board_id, task_id=task_id)
        db.commit()
        return {"assignees": [{"id": u.id, "name": u.name, "avatar_url": u.avatar_url}
                              for u in task.assignees]}

@app.post("/api/tasks/{task_id}/move")
def move_task(task_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            raise HTTPException(404)
        actor = actor_id
        _require_board_edit(db, task.board_id, actor)
        target = data.get("board_id")
        if target and int(target) != task.board_id:
            # ── move the item to another board ──
            target = int(target)
            if task.parent_id:
                raise HTTPException(400, "תת-פריט עובר יחד עם הפריט שלו")
            tb = db.query(Board).filter(Board.id == target).first()
            if not tb:
                raise HTTPException(404, "לוח היעד לא נמצא")
            _require_board_edit(db, target, actor)
            groups = db.query(Group).filter(Group.board_id == target).order_by(Group.position).all()
            wanted = data.get("group_id")
            grp = next((g for g in groups if g.id == wanted), None) or (groups[0] if groups else None)
            subs = db.query(Task).filter(Task.parent_id == task.id).all()
            for t in [task] + subs:
                t.board_id = target
                t.group_id = grp.id if grp else None
            if grp:
                task.status = grp.task_status
            maxpos = (db.query(func.max(Task.position))
                      .filter(Task.board_id == target, Task.parent_id == None,
                              Task.group_id == (grp.id if grp else None)).scalar())
            task.position = (maxpos + 1) if maxpos is not None else 0
            # the identifier belongs to the board it was minted for: the item (and
            # its sub-items) get new ones here, and new ones again if they return
            for t in [task] + subs:
                _ensure_item_uid(db, t)
            db.commit()
            return {"status": "moved", "board_id": target, "item_uid": task.item_uid}
        if "group_id" in data:
            task.group_id = data["group_id"]
        if "position" in data:
            task.position = data["position"]
        if "status" in data:
            task.status = data["status"]
        db.commit()
        return {"status": "moved"}

# ── Groups (board columns / קבוצות) ─────────────────────────────────

def _group_dict(g):
    return {"id": g.id, "name": g.name, "position": g.position, "color": g.color,
            "task_status": g.task_status.value if hasattr(g.task_status, 'value') else g.task_status}

@app.post("/api/groups")
def create_group_api(data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        board_id = data.get("board_id")
        board = db.query(Board).filter(Board.id == board_id).first()
        if not board:
            raise HTTPException(404, "board not found")
        _require_board_edit(db, board_id, actor_id)
        pos = data.get("position")
        if pos is None:
            pos = db.query(Group).filter(Group.board_id == board_id).count()
        try:
            task_status = TaskStatus(data.get("task_status", "todo"))
        except ValueError:
            task_status = TaskStatus.TODO
        g = Group(
            board_id=board_id,
            name=data.get("name", "קבוצה חדשה"),
            position=pos,
            color=data.get("color", "#0073ea"),
            task_status=task_status,
        )
        db.add(g)
        db.commit()
        db.refresh(g)
        return _group_dict(g)

@app.patch("/api/groups/{group_id}")
def update_group_api(group_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        g = db.query(Group).filter(Group.id == group_id).first()
        if not g:
            raise HTTPException(404, "group not found")
        _require_board_edit(db, g.board_id, actor_id)
        if data.get("name") is not None:
            g.name = data["name"]
        if data.get("color") is not None:
            g.color = data["color"]
        if data.get("position") is not None:
            g.position = data["position"]
        db.commit()
        db.refresh(g)
        return _group_dict(g)

@app.post("/api/groups/reorder")
def reorder_groups_api(data: dict, actor_id: int = Depends(current_user_id)):
    """Persist a new group ordering. data = {order: [groupId, ...]}"""
    with Session(engine) as db:
        order = data.get("order") or []
        actor = actor_id
        if order and actor is not None:
            first = db.query(Group).filter(Group.id == order[0]).first()
            if first:
                _require_board_edit(db, first.board_id, actor)
        for idx, gid in enumerate(order):
            g = db.query(Group).filter(Group.id == gid).first()
            if g:
                g.position = idx
        db.commit()
        return {"status": "reordered", "order": order}

@app.delete("/api/groups/{group_id}")
def delete_group_api(group_id: int, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        g = db.query(Group).filter(Group.id == group_id).first()
        if not g:
            raise HTTPException(404, "group not found")
        _require_board_edit(db, g.board_id, actor_id)   # board editors+ only
        # Detach tasks instead of deleting them — they fall back to "ללא קבוצה"
        for t in db.query(Task).filter(Task.group_id == group_id).all():
            t.group_id = None
        db.delete(g)
        db.commit()
        return {"status": "deleted"}

# ── Citizens & Permits ──────────────────────────────────────────────

@app.get("/api/citizen-requests")
def list_citizen_requests():
    with Session(engine) as db:
        reqs = db.query(CitizenRequest).order_by(CitizenRequest.created_at.desc()).limit(50).all()
        return [{
            "id": r.id, "request_type": r.request_type.value if hasattr(r.request_type, 'value') else r.request_type,
            "citizen_name": r.citizen_name, "title": r.title,
            "description": r.description,
            "location_lat": r.location_lat, "location_lng": r.location_lng,
            "address": r.address, "status": r.status,
            "priority": r.priority.value if hasattr(r.priority, 'value') else r.priority,
            "created_at": r.created_at,
        } for r in reqs]

@app.get("/api/permits")
def list_permits():
    with Session(engine) as db:
        permits = db.query(Permit).order_by(Permit.created_at.desc()).limit(50).all()
        return [{
            "id": p.id, "permit_type": p.permit_type.value if hasattr(p.permit_type, 'value') else p.permit_type,
            "permit_number": p.permit_number,
            "applicant_name": p.applicant_name,
            "property_address": p.property_address,
            "description": p.description,
            "status": p.status,
            "submitted_at": p.submitted_at,
        } for p in permits]

# ── GIS / Transport ─────────────────────────────────────────────────

@app.get("/api/transport/stops")
def transport_stops():
    with Session(engine) as db:
        stops = db.query(PublicTransportStop).all()
        features = []
        for s in stops:
            features.append({
                "type": "Feature",
                "geometry": {"type": "Point", "coordinates": [s.longitude, s.latitude]},
                "properties": {
                    "id": s.id, "stop_code": s.stop_code,
                    "name": s.name, "routes": s.routes or []
                }
            })
        return {"type": "FeatureCollection", "features": features}

@app.get("/api/infrastructure/assets")
def list_assets():
    with Session(engine) as db:
        assets = db.query(InfrastructureAsset).all()
        features = []
        for a in assets:
            features.append({
                "type": "Feature",
                "geometry": {"type": "Point", "coordinates": [a.location_lng, a.location_lat]},
                "properties": {
                    "id": a.id, "name": a.name,
                    "asset_type": a.asset_type, "condition": a.condition,
                    "status": a.status, "properties": a.properties or {}
                }
            })
        return {"type": "FeatureCollection", "features": features}

# ── Users ────────────────────────────────────────────────────────────

@app.get("/api/users")
def list_users():
    with Session(engine) as db:
        users = db.query(User).all()
        dept_names = {d.id: d.name for d in db.query(Department).all()}
        # most-recent login per user (for the admin directory "last login" column)
        from sqlalchemy import func as _func
        last_logins = dict(
            db.query(LoginEvent.user_id, _func.max(LoginEvent.logged_in_at)).group_by(LoginEvent.user_id).all()
        )
        def _iso(dt):
            return dt.isoformat() if dt else None
        return [{
            "id": u.id, "name": u.name, "email": u.email,
            "role": u.role, "avatar_url": u.avatar_url,
            "department_id": u.department_id,
            "department_name": dept_names.get(u.department_id),
            "is_active": bool(u.is_active) if u.is_active is not None else True,
            "phone": u.phone, "title": u.title,
            "last_login": _iso(last_logins.get(u.id)),
            "email_notifications": bool(u.email_notifications) if u.email_notifications is not None else True,
            "notif_prefs": u.notif_prefs or {},
        } for u in users]

@app.post("/api/auth/login")
async def record_login(request: Request, actor_id: int = Depends(current_user_id)):
    """Record a successful login for the admin login-history view.

    The recorded identity is the authenticated one — the client used to name the
    user it wanted logged, which made the audit trail forgeable."""
    with Session(engine) as db:
        u = db.query(User).filter(User.id == actor_id).first()
        if not u:
            raise HTTPException(404, "user not found")
        ev = LoginEvent(
            user_id=actor_id,
            ip=(request.client.host if request.client else None),
            user_agent=(request.headers.get("user-agent") or "")[:400],
        )
        db.add(ev)
        db.commit()
        return {"status": "ok", "logged_in_at": ev.logged_in_at.isoformat()}

@app.get("/api/auth/me")
def whoami(request: Request):
    """The signed-in user, resolved from the Entra identity on the request.
    The SPA calls this at boot instead of letting the visitor pick a user."""
    u = current_user(request)
    with Session(engine) as db:
        return {
            "id": u.id, "name": u.name, "email": u.email, "role": u.role,
            "avatar_url": u.avatar_url, "title": u.title,
            "ws_role": _ws_role(db, u.id),
            "auth_mode": auth_mode(),
        }


@app.get("/api/users/{uid}/login-history")
def login_history(uid: int, actor_id: int = Depends(current_user_id), limit: int = 50):
    """Login history for a user — visible to system (workspace) admins only."""
    with Session(engine) as db:
        if not _cap(db, actor_id, "manage_system"):
            raise HTTPException(403, "רק מנהל מערכת יכול לראות היסטוריית התחברות")
        if not db.query(User.id).filter(User.id == uid).scalar():
            raise HTTPException(404, "user not found")
        evs = (db.query(LoginEvent).filter(LoginEvent.user_id == uid)
               .order_by(LoginEvent.logged_in_at.desc()).limit(max(1, min(limit, 200))).all())
        return [{
            "logged_in_at": e.logged_in_at.isoformat() if e.logged_in_at else None,
            "ip": e.ip, "user_agent": e.user_agent,
        } for e in evs]

@app.get("/api/login-history")
def all_login_history(actor_id: int = Depends(current_user_id), limit: int = 100):
    """Consolidated login history across all users — system admins only."""
    with Session(engine) as db:
        if not _cap(db, actor_id, "manage_system"):
            raise HTTPException(403, "רק מנהל מערכת יכול לראות היסטוריית התחברות")
        names = {u.id: (u.name, u.avatar_url) for u in db.query(User).all()}
        evs = (db.query(LoginEvent).order_by(LoginEvent.logged_in_at.desc())
               .limit(max(1, min(limit, 500))).all())
        return [{
            "user_id": e.user_id,
            "user_name": names.get(e.user_id, ("—", None))[0],
            "avatar_url": names.get(e.user_id, ("—", None))[1],
            "logged_in_at": e.logged_in_at.isoformat() if e.logged_in_at else None,
            "ip": e.ip, "user_agent": e.user_agent,
        } for e in evs]

@app.post("/api/users")
def create_user(data: dict, actor_id: int = Depends(current_user_id)):
    """Create a new user in the workspace directory. System (workspace) admin only."""
    with Session(engine) as db:
        actor = actor_id
        if not _cap(db, actor, "manage_system"):
            raise HTTPException(403, "רק מנהל מערכת יכול להוסיף משתמש")
        name = (data.get("name") or "").strip()
        email = (data.get("email") or "").strip().lower()
        if not name:
            raise HTTPException(400, "חסר שם משתמש")
        if not email:
            raise HTTPException(400, "חסרה כתובת מייל")
        if db.query(User).filter(User.email == email).first():
            raise HTTPException(409, "כתובת המייל כבר קיימת במערכת")
        role = data.get("role") if data.get("role") in ("admin", "manager", "member", "viewer", "guest") else "member"
        dept_id = data.get("department_id") or None
        # inherit the organization of an existing user so the new user is in the same workspace
        org_id = db.query(User.organization_id).filter(User.organization_id != None).limit(1).scalar()
        u = User(name=name, email=email, role=role, department_id=dept_id,
                 organization_id=org_id, is_active=True)
        db.add(u)
        db.flush()
        # also register workspace membership so the role actually takes effect
        # (system-admin capabilities key off workspace_members, not User.role)
        ws_role = "admin" if role == "admin" else ("viewer" if role == "viewer" else "member")
        db.add(WorkspaceMember(user_id=u.id, role=ws_role))
        db.commit()
        db.refresh(u)
        dept_name = db.query(Department.name).filter(Department.id == u.department_id).scalar() if u.department_id else None
        return {"id": u.id, "name": u.name, "email": u.email, "role": u.role,
                "department_id": u.department_id, "department_name": dept_name,
                "is_active": True, "avatar_url": u.avatar_url}

@app.patch("/api/users/{uid}")
def update_user(uid: int, data: dict, actor_id: int = Depends(current_user_id)):
    """Update a user's profile. A user may edit their own profile; a system
    (workspace) admin may edit anyone's."""
    with Session(engine) as db:
        actor = actor_id
        if actor != uid and not _cap(db, actor, "manage_system"):
            raise HTTPException(403, "אפשר לעדכן רק את הפרופיל שלך")
        u = db.query(User).filter(User.id == uid).first()
        if not u:
            raise HTTPException(404, "user not found")
        for f in ("avatar_url", "phone", "title", "name", "email_notifications"):
            if f in data:
                setattr(u, f, data[f])
        if "email" in data:
            new_email = (data["email"] or "").strip().lower()
            if new_email and new_email != u.email:
                if db.query(User).filter(User.email == new_email, User.id != uid).first():
                    raise HTTPException(409, "כתובת המייל כבר קיימת במערכת")
                u.email = new_email
        if "notif_prefs" in data and isinstance(data["notif_prefs"], dict):
            u.notif_prefs = {**(u.notif_prefs or {}), **data["notif_prefs"]}
        # sensitive fields (role / department / active status) — system admin only,
        # never editable on one's own profile via this path
        admin_fields = {"role", "department_id", "is_active"}
        if admin_fields & set(data):
            if not _cap(db, actor, "manage_system"):
                raise HTTPException(403, "רק מנהל מערכת יכול לשנות תפקיד, מחלקה או סטטוס")
            if "role" in data and data["role"] in ("admin", "manager", "member", "viewer", "guest"):
                u.role = data["role"]
            if "department_id" in data:
                u.department_id = data["department_id"]
            if "is_active" in data:
                if u.id == actor and not data["is_active"]:
                    raise HTTPException(400, "אי אפשר להשבית את המשתמש שלך")
                u.is_active = bool(data["is_active"])
        db.commit()
        dept_name = None
        if u.department_id:
            d = db.query(Department).filter(Department.id == u.department_id).first()
            dept_name = d.name if d else None
        return {"id": u.id, "name": u.name, "email": u.email, "avatar_url": u.avatar_url,
                "phone": u.phone, "title": u.title, "role": u.role,
                "department_id": u.department_id, "department_name": dept_name,
                "is_active": bool(u.is_active),
                "email_notifications": bool(u.email_notifications),
                "notif_prefs": u.notif_prefs or {}}

@app.get("/api/departments")
def list_departments():
    with Session(engine) as db:
        depts = db.query(Department).order_by(Department.name).all()
        from sqlalchemy import func as _func
        ucount = dict(db.query(User.department_id, _func.count(User.id))
                      .filter(User.department_id != None).group_by(User.department_id).all())
        bcount = dict(db.query(Board.department_id, _func.count(Board.id))
                      .filter(Board.department_id != None).group_by(Board.department_id).all())
        return [{
            "id": d.id, "name": d.name, "code": d.code, "color": d.color,
            "organization_id": d.organization_id,
            "user_count": ucount.get(d.id, 0), "board_count": bcount.get(d.id, 0),
        } for d in depts]

DEPT_COLORS = ("#3b82f6", "#8b5cf6", "#ec4899", "#f59e0b", "#10b981",
               "#06b6d4", "#ef4444", "#6366f1", "#14b8a6", "#f97316")


@app.post("/api/departments")
def create_department(data: dict, actor_id: int = Depends(current_user_id)):
    """Create an organizational unit (department). System (workspace) admin only."""
    with Session(engine) as db:
        if not _cap(db, actor_id, "manage_system"):
            raise HTTPException(403, "רק מנהל מערכת יכול להוסיף יחידה ארגונית")
        name = (data.get("name") or "").strip()
        if not name:
            raise HTTPException(400, "חסר שם ליחידה הארגונית")
        if db.query(Department).filter(func.lower(Department.name) == name.lower()).first():
            raise HTTPException(409, "יחידה ארגונית בשם זה כבר קיימת")
        org_id = db.query(User.organization_id).filter(User.organization_id != None).limit(1).scalar()
        color = (data.get("color") or "").strip() or DEPT_COLORS[db.query(Department).count() % len(DEPT_COLORS)]
        d = Department(name=name, code=(data.get("code") or "").strip() or None,
                       color=color, organization_id=org_id)
        db.add(d)
        db.commit()
        db.refresh(d)
        return {"id": d.id, "name": d.name, "code": d.code, "color": d.color,
                "organization_id": d.organization_id, "user_count": 0, "board_count": 0}


@app.patch("/api/departments/{dept_id}")
def update_department(dept_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    """Rename / recolor an organizational unit. System (workspace) admin only."""
    with Session(engine) as db:
        if not _cap(db, actor_id, "manage_system"):
            raise HTTPException(403, "רק מנהל מערכת יכול לערוך יחידה ארגונית")
        d = db.query(Department).filter(Department.id == dept_id).first()
        if not d:
            raise HTTPException(404, "יחידה ארגונית לא נמצאה")
        if "name" in data:
            name = (data["name"] or "").strip()
            if not name:
                raise HTTPException(400, "חסר שם ליחידה הארגונית")
            if db.query(Department).filter(func.lower(Department.name) == name.lower(),
                                           Department.id != dept_id).first():
                raise HTTPException(409, "יחידה ארגונית בשם זה כבר קיימת")
            d.name = name
        if "code" in data:
            d.code = (data["code"] or "").strip() or None
        if "color" in data and data["color"]:
            d.color = data["color"].strip()
        db.commit()
        return {"id": d.id, "name": d.name, "code": d.code, "color": d.color}


@app.delete("/api/departments/{dept_id}")
def delete_department(dept_id: int, actor_id: int = Depends(current_user_id)):
    """Delete an organizational unit. System admin only. Any users/boards attached
    to it are detached (department set to none) rather than deleted."""
    with Session(engine) as db:
        if not _cap(db, actor_id, "manage_system"):
            raise HTTPException(403, "רק מנהל מערכת יכול למחוק יחידה ארגונית")
        d = db.query(Department).filter(Department.id == dept_id).first()
        if not d:
            raise HTTPException(404, "יחידה ארגונית לא נמצאה")
        # detach references so nothing is orphaned or violates a FK
        db.query(User).filter(User.department_id == dept_id).update({User.department_id: None})
        db.query(Board).filter(Board.department_id == dept_id).update({Board.department_id: None})
        db.query(Project).filter(Project.department_id == dept_id).update({Project.department_id: None})
        db.delete(d)
        db.commit()
        return {"status": "deleted", "id": dept_id}

# ── Role permissions (capability matrix) ─────────────────────────────

@app.get("/api/role-permissions")
def get_role_permissions():
    """The capability matrix. Readable by anyone so the client can gate features.
    'admin' is returned as an always-on, locked row."""
    with Session(engine) as db:
        matrix = _role_perms(db)
    roles = [{"key": "admin", "he": ROLE_HE_BE["admin"], "locked": True}] + \
            [{"key": r, "he": ROLE_HE_BE.get(r, r), "locked": False} for r in EDITABLE_ROLES]
    matrix_out = {"admin": {cap: True for cap in CAPABILITIES}, **matrix}
    return {
        "capabilities": [{"key": c, "he": CAP_HE[c]} for c in CAPABILITIES],
        "roles": roles,
        "matrix": matrix_out,
    }

@app.put("/api/role-permissions")
def set_role_permission(data: dict, actor_id: int = Depends(current_user_id)):
    """Toggle one (role, capability) cell. System-management capability required."""
    with Session(engine) as db:
        if not _cap(db, actor_id, "manage_system"):
            raise HTTPException(403, "רק מנהל מערכת יכול לשנות הרשאות תפקידים")
        role = data.get("role")
        cap = data.get("capability")
        if role not in EDITABLE_ROLES:
            raise HTTPException(400, "תפקיד לא ניתן לעריכה")
        if cap not in CAPABILITIES:
            raise HTTPException(400, "יכולת לא מוכרת")
        allowed = bool(data.get("allowed"))
        row = (db.query(RolePermission)
               .filter(RolePermission.role == role, RolePermission.capability == cap).first())
        if row:
            row.allowed = allowed
        else:
            db.add(RolePermission(role=role, capability=cap, allowed=allowed))
        db.commit()
        return {"role": role, "capability": cap, "allowed": allowed, "matrix": _role_perms(db)}

# ── Agent Swarm ───────────────────────────────────────────────────────

from swarm import get_swarm

@app.get("/api/swarm/agents")
def list_agents():
    swarm = get_swarm()
    return {"agents": [
        {"id": k, "name": v.name, "role": v.role}
        for k, v in swarm.agents.items()
    ]}

@app.post("/api/swarm/think")
def swarm_think(data: dict, actor_id: int = Depends(current_user_id)):
    """Run agents on a task. mode: single|all|coordinated"""
    from swarm import swarm_api
    result = swarm_api(
        agent=data.get("agent", ""),
        task=data.get("task", ""),
        context=data.get("context"),
        mode=data.get("mode", "single")
    )
    return result

# ── Form Templates ──────────────────────────────────────────────────

FORM_TEMPLATES = {
    "building_permit": {
        "name": "בקשה להיתר בנייה",
        "fields": [
            {"id": "applicant_name", "label": "שם המבקש", "type": "text", "required": True},
            {"id": "applicant_id", "label": "תז", "type": "text", "required": True},
            {"id": "phone", "label": "טלפון", "type": "tel", "required": True},
            {"id": "email", "label": "אימייל", "type": "email"},
            {"id": "property_address", "label": "כתובת הנכס", "type": "text", "required": True},
            {"id": "gush", "label": "גוש", "type": "text"},
            {"id": "helka", "label": "חלקה", "type": "text"},
            {"id": "permit_type", "label": "סוג היתר", "type": "select", "options": ["בנייה חדשה", "הרחבה", "שינוי ייעוד", "הריסה", "עבודות תשתית"]},
            {"id": "description", "label": "תיאור העבודות", "type": "textarea", "required": True},
            {"id": "area_sqm", "label": "שטח (מטר)", "type": "number"},
            {"id": "attachments", "label": "קבצים מצורפים", "type": "file", "multiple": True},
        ]
    },
    "citizen_request": {
        "name": "פניית תושב",
        "fields": [
            {"id": "citizen_name", "label": "שם מלא", "type": "text", "required": True},
            {"id": "phone", "label": "טלפון", "type": "tel", "required": True},
            {"id": "email", "label": "אימייל", "type": "email"},
            {"id": "request_type", "label": "סוג פנייה", "type": "select", "options": ["תקלה בכביש", "תאורת רחוב", "פסולת/ניקיון", "מים/ביוב", "גינה ציבורית", " רעש", "תחבורה ציבורית", "אחר"]},
            {"id": "title", "label": "כותרת", "type": "text", "required": True},
            {"id": "description", "label": "תיאור", "type": "textarea", "required": True},
            {"id": "address", "label": "כתובת מדויקת", "type": "text"},
            {"id": "location", "label": "מיקום במפה", "type": "location"},
            {"id": "photo", "label": "צילום", "type": "file"},
        ]
    },
    "event_permit": {
        "name": "בקשה לאישור אירוע",
        "fields": [
            {"id": "organizer_name", "label": "שם המארגן", "type": "text", "required": True},
            {"id": "organizer_phone", "label": "טלפון", "type": "tel", "required": True},
            {"id": "event_name", "label": "שם האירוע", "type": "text", "required": True},
            {"id": "event_type", "label": "סוג אירוע", "type": "select", "options": ["הרצאה", "מוזיקה", "ספורט", "יריד", "הפגנה", "אחר"]},
            {"id": "expected_attendees", "label": "משתתפים צפויים", "type": "number"},
            {"id": "event_date", "label": "תאריך האירוע", "type": "date", "required": True},
            {"id": "event_time", "label": "שעה", "type": "time"},
            {"id": "location", "label": "מיקום", "type": "text", "required": True},
            {"id": "description", "label": "תיאור האירוע", "type": "textarea"},
        ]
    },
}

@app.get("/api/forms/templates")
def list_form_templates():
    return {"templates": [
        {"id": k, "name": v["name"], "fields_count": len(v["fields"])}
        for k, v in FORM_TEMPLATES.items()
    ]}

@app.get("/api/forms/templates/{template_id}")
def get_form_template(template_id: str):
    if template_id not in FORM_TEMPLATES:
        raise HTTPException(404, f"Template '{template_id}' not found")
    return FORM_TEMPLATES[template_id]

@app.post("/api/forms/submit")
def submit_form(data: dict):
    """Submit form data and create a task."""
    template_id = data.get("template_id", "")
    form_data = data.get("form_data", {})
    with Session(engine) as db:
        task = Task(
            board_id=data.get("board_id", 3),  # Default to citizen requests
            title=form_data.get("title") or form_data.get("applicant_name") or f"טופס: {template_id}",
            description=json.dumps(form_data, ensure_ascii=False),
            tags=["form", template_id],
            custom_fields=form_data,
        )
        db.add(task)
        db.commit()
        return {"id": task.id, "status": "submitted"}

# ── Visualization Engine ─────────────────────────────────────────────

@app.get("/api/viz/board-insights")
def board_insights(user_id: int = Depends(current_user_id)):
    """Generate visualization data and insights for boards the user may see."""
    with Session(engine) as db:
        visible = _visible_board_ids(db, user_id)
        boards = db.query(Board).filter(
            Board.is_archived == False, Board.id.in_(visible)
        ).all() if visible else []
        insights = []
        for b in boards:
            tasks = db.query(Task).filter(Task.board_id == b.id).all()
            if not tasks:
                continue
            
            # Status distribution
            status_dist = {}
            priority_dist = {}
            overdue = 0
            now = datetime.now(timezone.utc).replace(tzinfo=None)
            dept_name = db.query(Department.name).filter(Department.id == b.department_id).scalar() or ""
            
            for t in tasks:
                s = t.status.value if hasattr(t.status, 'value') else t.status
                p = t.priority.value if hasattr(t.priority, 'value') else t.priority
                status_dist[s] = status_dist.get(s, 0) + 1
                priority_dist[p] = priority_dist.get(p, 0) + 1
                due = t.due_date
                if due and hasattr(due, 'tzinfo') and due.tzinfo:
                    due = due.replace(tzinfo=None)
                if due and due < now and s not in ("done", "cancelled"):
                    overdue += 1
            
            insights.append({
                "board_id": b.id,
                "board_name": b.name,
                "icon": b.icon,
                "department": dept_name,
                "total_tasks": len(tasks),
                "status_distribution": status_dist,
                "priority_distribution": priority_dist,
                "overdue_count": overdue,
                "completion_rate": round(status_dist.get("done", 0) / max(len(tasks), 1) * 100, 1),
            })
        
        return {"boards": insights, "total": len(insights)}

@app.get("/api/viz/timeline")
def viz_timeline(days: int = 30):
    """Task timeline data for Gantt-like visualization."""
    with Session(engine) as db:
        from datetime import timedelta
        cutoff = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=days)
        tasks = db.query(Task).filter(
            Task.created_at >= cutoff
        ).order_by(Task.created_at).all()
        
        # Group by day
        daily = {}
        for t in tasks:
            day = t.created_at.strftime("%Y-%m-%d") if t.created_at else "unknown"
            if day not in daily:
                daily[day] = 0
            daily[day] += 1
        
        return {"timeline": [{"date": d, "count": c} for d, c in sorted(daily.items())]}

# ── Obsidian-like Graph ─────────────────────────────────────────────

@app.get("/api/graph/context")
def graph_context():
    """Generate a knowledge graph of all tasks, boards, and their connections."""
    with Session(engine) as db:
        boards = db.query(Board).filter(Board.is_archived == False).all()
        tasks = db.query(Task).filter(Task.is_archived == False).all()
        users = db.query(User).all()
        
        nodes = []
        edges = []
        
        # Board nodes
        for b in boards:
            nodes.append({
                "id": f"board_{b.id}", "label": b.name, "type": "board",
                "icon": b.icon, "color": b.color
            })
        
        # Task nodes
        for t in tasks:
            nodes.append({
                "id": f"task_{t.id}", "label": t.title[:30], "type": "task",
                "status": t.status.value if hasattr(t.status, 'value') else t.status,
                "priority": t.priority.value if hasattr(t.priority, 'value') else t.priority,
            })
            edges.append({
                "source": f"task_{t.id}", "target": f"board_{t.board_id}",
                "type": "belongs_to"
            })
        
        # User nodes
        for u in users:
            nodes.append({
                "id": f"user_{u.id}", "label": u.name, "type": "user",
                "role": u.role
            })
        
        # Assignee edges
        for t in tasks:
            if t.assignees:
                for a in t.assignees:
                    edges.append({
                        "source": f"user_{a.id}", "target": f"task_{t.id}",
                        "type": "assigned_to"
                    })
        
        return {"nodes": nodes, "edges": edges}

# ── AI Assistant ─────────────────────────────────────────────────────

# ── Tool Definitions ────────────────────────────────────────────────

AI_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "list_boards",
            "description": "רשימת כל הלוחות (boards) במערכת",
            "parameters": {"type": "object", "properties": {}, "required": []}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_board",
            "description": "קבלת פרטי לוח לפי ID",
            "parameters": {
                "type": "object",
                "properties": {"board_id": {"type": "integer", "description": "מזהה הלוח"}},
                "required": ["board_id"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "list_projects",
            "description": "רשימת פרויקטים (ניתן לסנן לפי department_id)",
            "parameters": {
                "type": "object",
                "properties": {
                    "department_id": {"type": "integer", "description": "מזהה אגף (אופציונלי)"}
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_board",
            "description": "יצירת לוח חדש במערכת",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "שם הלוח"},
                    "department_id": {"type": "integer", "description": "מזהה האגף (השתמש ב-1 כברירת מחדל)"},
                    "description": {"type": "string", "description": "תיאור הלוח"},
                    "icon": {"type": "string", "description": "אימוג'י ללוח, ברירת מחדל 📋"},
                    "color": {"type": "string", "description": "צבע לוח. מותר: #0073ea, #00c875, #fdab3d, #e2445c, #579bfc, #c4c4c4"},
                },
                "required": ["name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_group",
            "description": "יצירת עמודה/קבוצה בלוח",
            "parameters": {
                "type": "object",
                "properties": {
                    "board_id": {"type": "integer", "description": "מזהה הלוח"},
                    "name": {"type": "string", "description": "שם הקבוצה/עמודה"},
                    "position": {"type": "integer", "description": "מיקום (0 = ראשון)"},
                    "color": {"type": "string", "description": "צבע"},
                    "task_status": {"type": "string", "description": "סטטוס: backlog/todo/in_progress/review/done/cancelled/on_hold"},
                },
                "required": ["board_id", "name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_task",
            "description": "יצירת משימה (פריט) חדשה בלוח. חובה רק board_id ו-title. אם לא צוינה קבוצה — המשימה תיכנס לקבוצה הראשונה בלוח. מלא רק שדות שהמשתמש ביקש במפורש.",
            "parameters": {
                "type": "object",
                "properties": {
                    "board_id": {"type": "integer", "description": "מזהה הלוח"},
                    "group_id": {"type": "integer", "description": "מזהה הקבוצה/עמודה (אופציונלי; ברירת מחדל: הקבוצה הראשונה בלוח)"},
                    "title": {"type": "string", "description": "כותרת המשימה"},
                    "description": {"type": "string", "description": "תיאור המשימה (אופציונלי — השאר ריק אם לא צוין)"},
                    "priority": {"type": "string", "enum": ["low", "medium", "high", "critical", "emergency"], "description": "עדיפות (אופציונלי; ברירת מחדל: medium)"},
                    "status": {"type": "string", "enum": ["backlog", "todo", "in_progress", "review", "done", "cancelled", "on_hold"], "description": "סטטוס (אופציונלי; ברירת מחדל: backlog)"},
                    "due_date": {"type": "string", "description": "תאריך יעד בפורמט ISO (אופציונלי)"},
                    "tags": {"type": "array", "items": {"type": "string"}, "description": "תגיות (אופציונלי)"},
                },
                "required": ["board_id", "title"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "update_task",
            "description": "עדכון משימה קיימת",
            "parameters": {
                "type": "object",
                "properties": {
                    "task_id": {"type": "integer", "description": "מזהה המשימה"},
                    "title": {"type": "string"},
                    "status": {"type": "string", "enum": ["backlog", "todo", "in_progress", "review", "done", "cancelled", "on_hold"]},
                    "priority": {"type": "string", "enum": ["low", "medium", "high", "critical", "emergency"]},
                    "due_date": {"type": "string"},
                    "description": {"type": "string"},
                    "tags": {"type": "array", "items": {"type": "string"}},
                },
                "required": ["task_id"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "delete_task",
            "description": "מחיקת משימה (ארכון)",
            "parameters": {
                "type": "object",
                "properties": {
                    "task_id": {"type": "integer", "description": "מזהה המשימה למחיקה"}
                },
                "required": ["task_id"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "list_departments",
            "description": "רשימת כל האגפים במערכת",
            "parameters": {"type": "object", "properties": {}, "required": []}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "list_users",
            "description": "רשימת כל המשתמשים במערכת",
            "parameters": {"type": "object", "properties": {}, "required": []}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_project",
            "description": "יצירת פרויקט חדש במסגרת תוכנית העבודה",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "שם הפרויקט"},
                    "department_id": {"type": "integer", "description": "מזהה האגף"},
                    "work_plan_id": {"type": "integer", "description": "מזהה תוכנית עבודה, ברירת מחדל 1"},
                    "planned_budget": {"type": "number", "description": "תקציב מתוכנן"},
                    "status": {"type": "string", "enum": ["draft", "planning", "in_progress", "completed", "cancelled", "on_hold"]},
                    "priority": {"type": "string", "enum": ["low", "medium", "high", "critical"]},
                    "manager_name": {"type": "string", "description": "שם מנהל הפרויקט (יחפש משתמש לפי שם)"}
                },
                "required": ["name", "department_id"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_project_detail",
            "description": "קבלת פרטי פרויקט מלאים (עם שלבים, תקציב, KPI, אישורים)",
            "parameters": {
                "type": "object",
                "properties": {"project_id": {"type": "integer"}},
                "required": ["project_id"]
            }
        }
    },
]


def _ai_board_guard(db, board_id, actor, need="editor"):
    """Board permission check for the AI tools. Returns a Hebrew refusal string
    (not an exception) so the model can relay it, or None when allowed."""
    role = _board_role(db, board_id, actor)
    if role is None:
        return "❌ אין לך גישה ללוח זה."
    if _BROLE_ORDER.get(role, -1) < _BROLE_ORDER.get(need, 1):
        return "❌ אין לך הרשאת עריכה בלוח זה."
    return None


def execute_ai_tool(name: str, args: dict, actor: Optional[int] = None) -> str:
    """Execute an AI tool by name with the given arguments. Returns a descriptive Hebrew result string.

    Every tool runs with the *caller's* permissions. These tools used to ignore
    `actor` entirely, so anyone who could reach the chat endpoint could read or
    delete items on boards they were never invited to — including by way of
    prompt injection planted in a task description the model later read."""
    from datetime import datetime, timezone

    if actor is None:
        return "❌ נדרשת התחברות."

    if name == "list_boards":
        with Session(engine) as db:
            visible = _visible_board_ids(db, actor)
            boards = [b for b in db.query(Board).filter(Board.is_archived == False).all()
                      if b.id in visible]
            if not boards:
                return "❌ לא נמצאו לוחות במערכת."
            lines = []
            for b in boards:
                dept_name = db.query(Department.name).filter(Department.id == b.department_id).scalar() or ""
                task_count = db.query(Task).filter(Task.board_id == b.id).count()
                lines.append(f"  🆔 {b.id} | {b.icon or '📋'} **{b.name}** | אגף: {dept_name} | {task_count} משימות")
            return "📋 **כל הלוחות במערכת:**\n" + "\n".join(lines)

    elif name == "get_board":
        board_id = args.get("board_id")
        with Session(engine) as db:
            b = db.query(Board).filter(Board.id == board_id).first()
            if not b:
                return f"❌ לוח {board_id} לא נמצא."
            if _board_role(db, board_id, actor) is None:
                return "❌ אין לך גישה ללוח זה."
            dept_name = db.query(Department.name).filter(Department.id == b.department_id).scalar() or ""
            groups = db.query(Group).filter(Group.board_id == b.id).order_by(Group.position).all()
            tasks = db.query(Task).filter(Task.board_id == b.id, Task.is_archived == False).all()
            lines = [
                f"📋 **{b.name}**",
                f"   תיאור: {b.description or 'אין תיאור'}",
                f"   אגף: {dept_name}",
                f"   סוג: {b.board_type.value if hasattr(b.board_type, 'value') else b.board_type}",
                f"   צבע: {b.color}",
                f"   סה״כ משימות: {len(tasks)}",
            ]
            if groups:
                lines.append(f"   **עמודות ({len(groups)})**:")
                for g in groups:
                    g_tasks = [t for t in tasks if t.group_id == g.id]
                    lines.append(f"      • {g.name} ({len(g_tasks)} משימות)")
            return "\n".join(lines)

    elif name == "list_projects":
        dept_id = args.get("department_id")
        with Session(engine) as db:
            q = db.query(Project)
            if dept_id:
                q = q.filter(Project.department_id == dept_id)
            projects = q.order_by(Project.id).all()
            if not projects:
                return "❌ לא נמצאו פרויקטים."
            lines = []
            for p in projects:
                dept_name = db.query(Department.name).filter(Department.id == p.department_id).scalar() or ""
                s = p.status.value if hasattr(p.status, 'value') else p.status
                lines.append(f"  🆔 {p.id} | **{p.name}** | אגף: {dept_name} | סטטוס: {s} | תקציב: ₪{p.planned_budget or 0:,}")
            return "📋 **כל הפרויקטים:**\n" + "\n".join(lines)

    elif name == "create_board":
        name = args.get("name", "לוח חדש")
        dept_id = args.get("department_id", 1)
        description = args.get("description", "")
        icon = args.get("icon", "📋")
        color = args.get("color", "#0073ea")
        with Session(engine) as db:
            # same rule as the UI: only a system (workspace) admin may create a
            # board, and the creator is registered as its admin — never orphaned.
            if _ws_role(db, actor) != "admin":
                return "❌ רק מנהל מערכת יכול ליצור לוח חדש."
            b = Board(
                name=name,
                description=description,
                department_id=dept_id,
                board_type=BoardType.KANBAN,
                icon=icon,
                color=color,
                settings={"views": ["table"]},  # new boards start with only the main table
            )
            db.add(b)
            db.flush()

            # Create default groups
            g1 = Group(board_id=b.id, name="בתכנון", position=0, color="#579bfc", task_status=TaskStatus.BACKLOG)
            g2 = Group(board_id=b.id, name="בתהליך", position=1, color="#fdab3d", task_status=TaskStatus.IN_PROGRESS)
            g3 = Group(board_id=b.id, name="הושלם", position=2, color="#00c875", task_status=TaskStatus.DONE)
            db.add_all([g1, g2, g3])
            db.add(BoardMember(board_id=b.id, user_id=actor, role="admin"))
            db.commit()
            db.refresh(b)
            return f"✅ לוח '{name}' נוצר בהצלחה (מזהה: {b.id}) — אתה מוגדר כמנהל הלוח."

    elif name == "create_group":
        board_id = args.get("board_id")
        name = args.get("name", "קבוצה חדשה")
        position = args.get("position", 0)
        color = args.get("color", "#c4c4c4")
        task_status_str = args.get("task_status", "todo")
        try:
            task_status = TaskStatus(task_status_str)
        except ValueError:
            task_status = TaskStatus.TODO
        with Session(engine) as db:
            board = db.query(Board).filter(Board.id == board_id).first()
            if not board:
                return f"❌ לוח {board_id} לא נמצא."
            denied = _ai_board_guard(db, board_id, actor)
            if denied:
                return denied
            g = Group(
                board_id=board_id,
                name=name,
                position=position,
                color=color,
                task_status=task_status,
            )
            db.add(g)
            db.commit()
            db.refresh(g)
            return f"✅ עמודה '{name}' נוצרה בלוח {board_id} (מזהה: {g.id})."

    elif name == "create_task":
        board_id = args.get("board_id")
        group_id = args.get("group_id")
        title = args.get("title", "משימה חדשה")
        description = args.get("description", "")
        # defaults for an item created via the agent: medium priority, "בתכנון"
        # (backlog) status, no assignees, everything else empty.
        priority_str = args.get("priority", "medium")
        status_str = args.get("status", "backlog")
        due_date_str = args.get("due_date")
        tags = args.get("tags", [])
        try:
            priority = Priority(priority_str)
        except ValueError:
            priority = Priority.MEDIUM
        try:
            status = TaskStatus(status_str)
        except ValueError:
            status = TaskStatus.BACKLOG
        due_date = None
        if due_date_str:
            try:
                due_date = datetime.fromisoformat(due_date_str)
            except Exception:
                pass
        with Session(engine) as db:
            board = db.query(Board).filter(Board.id == board_id).first()
            if not board:
                return f"❌ לוח {board_id} לא נמצא."
            denied = _ai_board_guard(db, board_id, actor)
            if denied:
                return denied
            # group_id is optional — fall back to the board's first group so the
            # agent can add an item without needing to resolve column ids first.
            g = None
            if group_id:
                g = db.query(Group).filter(Group.id == group_id, Group.board_id == board_id).first()
            if not g:
                g = db.query(Group).filter(Group.board_id == board_id).order_by(Group.position).first()
            group_id = g.id if g else None
            t = Task(
                board_id=board_id,
                group_id=group_id,
                title=title,
                description=description,
                priority=priority,
                status=status,
                due_date=due_date,
                tags=tags,
            )
            db.add(t)
            db.commit()
            db.refresh(t)
            due_str = f" (יעד: {due_date_str})" if due_date_str else ""
            return f"✅ משימה '{title}' נוצרה בהצלחה{due_str} (מזהה: {t.id})."

    elif name == "update_task":
        task_id = args.get("task_id")
        with Session(engine) as db:
            t = db.query(Task).filter(Task.id == task_id).first()
            if not t:
                return f"❌ משימה {task_id} לא נמצאה."
            role = _board_role(db, t.board_id, actor)
            if _item_perm(t, actor, role) not in ("edit", "delete"):
                return "❌ אין לך הרשאת עריכה לפריט זה."
            changed = []
            for field in ["title", "description", "due_date"]:
                if field in args:
                    old_val = getattr(t, field)
                    if field == "due_date" and args[field]:
                        try:
                            setattr(t, field, datetime.fromisoformat(args[field]))
                        except Exception:
                            pass
                    else:
                        setattr(t, field, args[field])
                    if old_val != getattr(t, field):
                        changed.append(field)
            if "status" in args:
                try:
                    ns = TaskStatus(args["status"])
                    t.status = ns
                    changed.append("status")
                except ValueError:
                    pass
            if "priority" in args:
                try:
                    np = Priority(args["priority"])
                    t.priority = np
                    changed.append("priority")
                except ValueError:
                    pass
            if "tags" in args:
                t.tags = args["tags"]
                changed.append("tags")
            for c in changed:
                db.add(AuditLog(
                    entity_type="task", entity_id=task_id,
                    action="update", field_name=c,
                    new_value=str(getattr(t, c)),
                ))
            db.commit()
            if changed:
                return f"✅ משימה {task_id} עודכנה בהצלחה. שדות שהשתנו: {', '.join(changed)}."
            return f"ℹ️ לא בוצעו שינויים במשימה {task_id}."

    elif name == "delete_task":
        task_id = args.get("task_id")
        with Session(engine) as db:
            t = db.query(Task).filter(Task.id == task_id).first()
            if not t:
                return f"❌ משימה {task_id} לא נמצאה."
            role = _board_role(db, t.board_id, actor)
            if _item_perm(t, actor, role) != "delete":
                return "❌ אין לך הרשאת מחיקה לפריט זה."
            t.is_archived = True
            db.add(AuditLog(
                entity_type="task", entity_id=task_id,
                action="archive", field_name="is_archived",
                new_value="True", changed_by=actor,
            ))
            db.commit()
            return f"✅ משימה '{t.title}' (מזהה: {task_id}) אורכבה בהצלחה."

    elif name == "list_departments":
        with Session(engine) as db:
            depts = db.query(Department).all()
            if not depts:
                return "❌ לא נמצאו אגפים במערכת."
            lines = []
            for d in depts:
                proj_count = db.query(Project).filter(Project.department_id == d.id).count()
                lines.append(f"  🆔 {d.id} | **{d.name}** | קוד: {d.code or '-'} | {proj_count} פרויקטים")
            return "🏢 **כל האגפים:**\n" + "\n".join(lines)

    elif name == "list_users":
        with Session(engine) as db:
            users = db.query(User).all()
            if not users:
                return "❌ לא נמצאו משתמשים במערכת."
            lines = []
            for u in users:
                dept_name = db.query(Department.name).filter(Department.id == u.department_id).scalar() or ""
                lines.append(f"  🆔 {u.id} | **{u.name}** | תפקיד: {u.role or '-'} | אגף: {dept_name}")
            return "👥 **כל המשתמשים:**\n" + "\n".join(lines)

    elif name == "create_project":
        name = args.get("name", "פרויקט חדש")
        dept_id = args.get("department_id")
        work_plan_id = args.get("work_plan_id", 1)
        planned_budget = args.get("planned_budget", 0)
        status_str = args.get("status", "draft")
        priority_str = args.get("priority", "medium")
        manager_name = args.get("manager_name")
        with Session(engine) as db:
            dept = db.query(Department).filter(Department.id == dept_id).first()
            if not dept:
                dept_id = 1
            manager_id = None
            if manager_name:
                user = db.query(User).filter(User.name.ilike(f"%{manager_name}%")).first()
                if user:
                    manager_id = user.id
            try:
                p_status = ProjectStatus(status_str)
            except ValueError:
                p_status = ProjectStatus.DRAFT
            try:
                p_priority = Priority(priority_str)
            except ValueError:
                p_priority = Priority.MEDIUM
            p = Project(
                work_plan_id=work_plan_id,
                department_id=dept_id,
                name=name,
                planned_budget=planned_budget,
                status=p_status,
                priority=p_priority,
                manager_id=manager_id,
            )
            db.add(p)
            db.commit()
            db.refresh(p)
            mgr_text = f" (מנהל: {manager_name})" if manager_id else ""
            return f"✅ פרויקט '{name}' נוצר בהצלחה{mgr_text} (מזהה: {p.id})."

    elif name == "get_project_detail":
        project_id = args.get("project_id")
        with Session(engine) as db:
            p = db.query(Project).filter(Project.id == project_id).first()
            if not p:
                return f"❌ פרויקט {project_id} לא נמצא."
            dept_name = db.query(Department.name).filter(Department.id == p.department_id).scalar() or ""
            mgr_name = ""
            if p.manager_id:
                u = db.query(User).filter(User.id == p.manager_id).first()
                if u:
                    mgr_name = u.name
            s = p.status.value if hasattr(p.status, 'value') else p.status
            pri = p.priority.value if hasattr(p.priority, 'value') else p.priority
            steps = db.query(ProjectStep).filter(ProjectStep.project_id == p.id).order_by(ProjectStep.position).all()
            budget_items = db.query(BudgetLineItem).filter(BudgetLineItem.project_id == p.id).all()
            kpis = db.query(KPI).filter(KPI.project_id == p.id).all()
            lines = [
                f"📊 **{p.name}** (מזהה: {p.id})",
                f"   אגף: {dept_name}",
                f"   מנהל: {mgr_name or 'לא הוגדר'}",
                f"   סטטוס: {s} | עדיפות: {pri}",
                f"   התקדמות: {p.progress_percentage or 0}%",
                f"   תקציב: מתוכנן ₪{p.planned_budget or 0:,} | מאושר ₪{p.approved_budget or 0:,} | בפועל ₪{p.actual_budget or 0:,}",
                f"   תאריכים: {p.start_date.strftime('%d/%m/%Y') if p.start_date else '?'} → {p.end_date.strftime('%d/%m/%Y') if p.end_date else '?'}",
            ]
            if steps:
                lines.append(f"   **שלבים ({len(steps)})**:")
                for st in steps:
                    lines.append(f"      • {st.name} - {st.progress or 0}% ({st.status})")
            if kpis:
                lines.append(f"   **KPI ({len(kpis)})**:")
                for k in kpis:
                    ach = round(k.actual / max(k.target, 1) * 100, 1)
                    lines.append(f"      • {k.name}: {k.actual}/{k.target} {k.unit or ''} ({ach}%)")
            if budget_items:
                lines.append(f"   **תקציב לפי סעיף ({len(budget_items)})**:")
                for bi in budget_items:
                    bt = bi.item_type.value if hasattr(bi.item_type, 'value') else bi.item_type
                    lines.append(f"      • {bi.name or bt}: מאושר ₪{bi.approved_amount or 0:,} | בפועל ₪{bi.actual_amount or 0:,}")
            return "\n".join(lines)

    return f"❌ פונקציה '{name}' לא מוכרת."


@app.post("/api/ai/query")
def ai_query(data: dict, actor_id: int = Depends(current_user_id)):
    """CODE-CAL MISSIONS agent — DeepSeek (cloud) with tool calling, Gemini fallback, or local Ollama."""
    import subprocess
    prompt = data.get("prompt", "")
    context = data.get("context", "")
    deepseek_key = os.environ.get("DEEPSEEK_API_KEY")
    gemini_key = os.environ.get("GEMINI_API_KEY")
    model = data.get("model") or ("kremer" if deepseek_key else "gemma4-coder")

    # Model alias mapping: kremer = DeepSeek, elaine = Gemini
    model_internal = model.lower()

    system = """אתה CODE-CAL MISSIONS, עוזר עירוני חכם לניהול תוכניות עבודה ותקציב.
אתה יכול לשוחח עם המשתמש בעברית וגם לבצע פעולות במערכת באמצעות tools.

הכללים:
1. כשמשתמש מבקש ליצור/לעדכן/למחוק משהו - השתמש ב-tools המתאימים
2. כשמשתמש שואל שאלה - ענה מידע מהמערכת
3. תמיד אשר למשתמש אחרי ביצוע פעולה
4. כשאתה יוצר לוח חדש, צור גם קבוצות (עמודות) מתאימות: "לתכנון" (backlog), "בתהליך" (in_progress), "הושלם" (done)
5. יצירת פריט (משימה): צריך רק **שם**. קח את הלוח מההקשר — בהקשר יש "מזהה הלוח הפעיל" ורשימת קבוצות; העבר את ה-board_id הזה ל-create_task (group_id אופציונלי, ברירת מחדל היא הקבוצה הראשונה). אם אין לוח פעיל בהקשר - בקש מהמשתמש לאיזה לוח.
6. ברירות מחדל לפריט חדש: עדיפות **medium**, סטטוס **backlog** (בתכנון), **בלי אחראים**, ושאר השדות ריקים. אל תמלא תיאור/תאריך/תגיות אלא אם המשתמש ביקש במפורש - אל תמציא.
7. דבר בעברית תמיד
8. השתמש באימוג'ים במידה
9. אם אתה לא יודע איזה department_id או board_id - השתמש ברשימה קודם"""

    if model_internal == "kremer":
        # DeepSeek with tool calling
        if not deepseek_key:
            return {"response": "קרמר לא זמין כרגע (מפתח חסר).", "success": False, "model": model}
        try:
            import httpx
            messages = [
                {"role": "system", "content": system},
            ]
            user_msg = prompt
            if context:
                user_msg = f"הקשר: {context}\n\n{user_msg}"
            messages.append({"role": "user", "content": user_msg})

            r = httpx.post(
                "https://api.deepseek.com/v1/chat/completions",
                headers={"Authorization": f"Bearer {deepseek_key}", "Content-Type": "application/json"},
                json={
                    "model": "deepseek-chat",
                    "temperature": 0.7,
                    "messages": messages,
                    "tools": AI_TOOLS,
                    "tool_choice": "auto",
                },
                timeout=60,
            )
            r.raise_for_status()
            resp_data = r.json()
            choice = resp_data["choices"][0]
            msg = choice["message"]

            tool_calls = msg.get("tool_calls", [])
            executed = []
            tool_results = []

            if tool_calls:
                max_rounds = 5
                rounds = 0
                final_text = ""

                while rounds < max_rounds:
                    rounds += 1
                    # Execute each tool call
                    for tc in tool_calls:
                        fn_name = tc["function"]["name"]
                        try:
                            fn_args = json.loads(tc["function"]["arguments"])
                        except Exception:
                            fn_args = {}
                        result_text = execute_ai_tool(fn_name, fn_args, actor_id)
                        tool_results.append({
                            "tool_call_id": tc.get("id", ""),
                            "function_name": fn_name,
                            "arguments": fn_args,
                            "result": result_text,
                        })
                        executed.append({
                            "name": fn_name,
                            "arguments": fn_args,
                            "result": result_text,
                        })

                    # Send results back to the model
                    messages.append(msg)
                    for tr in tool_results:
                        messages.append({
                            "role": "tool",
                            "tool_call_id": tr["tool_call_id"],
                            "content": tr["result"],
                        })

                    r2 = httpx.post(
                        "https://api.deepseek.com/v1/chat/completions",
                        headers={"Authorization": f"Bearer {deepseek_key}", "Content-Type": "application/json"},
                        json={
                            "model": "deepseek-chat",
                            "temperature": 0.7,
                            "messages": messages,
                            "tools": AI_TOOLS,
                            "tool_choice": "auto",
                        },
                        timeout=60,
                    )
                    r2.raise_for_status()
                    resp_data2 = r2.json()
                    msg = resp_data2["choices"][0]["message"]
                    tool_calls = msg.get("tool_calls", [])
                    tool_results = []

                    if not tool_calls:
                        # No more tools to call
                        final_text = msg.get("content", "").strip()
                        break

                    final_text = msg.get("content", "") or ""

                return {
                    "response": final_text or "הפעולות בוצעו בהצלחה.",
                    "success": True,
                    "tool_calls": executed,
                    "model": model,
                    "provider": "kremer",
                }

            # No tool calls — just a regular chat response
            txt = msg.get("content", "").strip()
            return {
                "response": txt or "לא התקבלה תשובה",
                "success": True,
                "tool_calls": [],
                "model": model,
                "provider": "kremer",
            }

        except Exception as e:
            return {"response": f"קרמר לא זמין: {str(e)[:200]}", "success": False, "model": model, "tool_calls": []}

    if model_internal == "elaine":
        # Gemini (elaine) — no tool calling, just Q&A
        if not gemini_key:
            return {"response": "איליין לא זמינה כרגע (מפתח חסר).", "success": False, "model": model, "tool_calls": []}
        import httpx, time
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={gemini_key}"
        combined = prompt
        if context:
            combined = f"הקשר: {context}\n\n{prompt}"
        body = {"contents": [{"parts": [{"text": f"{system}\n\n{combined}"}]}]}
        max_retries = 3
        last_error = ""
        for attempt in range(max_retries):
            try:
                r = httpx.post(url, json=body, timeout=60)
                if r.status_code == 429 and attempt < max_retries - 1:
                    wait = 2 ** (attempt + 1)
                    last_error = f"rate limited (429), retrying in {wait}s"
                    time.sleep(wait)
                    continue
                r.raise_for_status()
                data_r = r.json()
                candidates = data_r.get("candidates", [])
                if candidates:
                    txt = candidates[0].get("content", {}).get("parts", [{}])[0].get("text", "")
                    return {"response": txt.strip() or "לא התקבלה תשובה", "success": True, "tool_calls": [], "model": model, "provider": "elaine"}
                return {"response": "איליין לא החזירה תוכן.", "success": True, "tool_calls": [], "model": model, "provider": "elaine"}
            except Exception as e:
                if "429" in str(e) and attempt < max_retries - 1:
                    wait = 2 ** (attempt + 1)
                    last_error = str(e)[:100]
                    time.sleep(wait)
                    continue
                last_error = str(e)[:200]
                break
        # Fallback to DeepSeek (kremer) if available
        if deepseek_key:
            try:
                import httpx
                r = httpx.post(
                    "https://api.deepseek.com/v1/chat/completions",
                    headers={"Authorization": f"Bearer {deepseek_key}"},
                    json={"model": "deepseek-chat", "temperature": 0.7, "messages": [
                        {"role": "system", "content": system},
                        {"role": "user", "content": f"הקשר: {context}\n\nשאלה: {prompt}"},
                    ]},
                    timeout=60,
                )
                r.raise_for_status()
                txt = r.json()["choices"][0]["message"]["content"].strip()
                return {"response": txt or "לא התקבלה תשובה", "success": True, "model": "kremer", "fallback": True, "provider": "kremer", "tool_calls": []}
            except Exception:
                pass
        return {"response": f"איליין לא זמינה (מוגבל): {last_error}", "success": False, "model": model, "tool_calls": []}

    # Local Ollama models — no tool calling
    full_prompt = f"{system}\n\nContext: {context}\n\nQuestion: {prompt}\n\nAnswer:"
    try:
        result = subprocess.run(
            ["ollama", "run", model, full_prompt],
            capture_output=True, text=True, timeout=120,
            env={**os.environ, "OLLAMA_NUM_THREADS": "8"},
        )
        return {"response": result.stdout.strip() or "לא התקבלה תשובה", "success": True, "tool_calls": [], "model": model, "provider": "ollama"}
    except FileNotFoundError:
        return {"response": "שירות ה-AI המקומי אינו זמין בשרת זה. אפשר לחבר מודל ענן דרך הגדרות המערכת.", "success": False, "model": model, "tool_calls": []}
    except subprocess.TimeoutExpired:
        return {"response": "מודל ה-AI לא הגיב בזמן. נסה שוב או בחר מודל אחר.", "success": False, "model": model, "tool_calls": []}
    except Exception:
        return {"response": "שירות ה-AI אינו זמין כרגע. נסה שוב מאוחר יותר.", "success": False, "model": model, "tool_calls": []}


@app.get("/api/ai/models")
def ai_models():
    """List available LLMs: kremer (DeepSeek), elaine (Gemini), local models."""
    import subprocess
    models = []
    default = "gemma4-coder"
    if os.environ.get("DEEPSEEK_API_KEY"):
        models += [{"name": "kremer", "size": "ענן · DeepSeek"}]
        default = "kremer"
    if os.environ.get("GEMINI_API_KEY"):
        models += [{"name": "elaine", "size": "ענן · Gemini"}]
    try:
        out = subprocess.run(["ollama", "list"], capture_output=True, text=True, timeout=15)
        for line in out.stdout.splitlines()[1:]:
            parts = line.split()
            if parts:
                models.append({"name": parts[0], "size": (parts[2] + " " + parts[3]) if len(parts) > 3 else "מקומי"})
    except Exception:
        pass
    return {"models": models, "available": bool(models), "default": default}

@app.post("/api/llm/v1/chat/completions")
async def llm_proxy(request: Request, actor_id: int = Depends(current_user_id)):
    """Server-side LLM proxy so in-browser clients (e.g. page-agent) can use the
    model WITHOUT ever seeing the API key. Frontend sets baseURL=/api/llm/v1 and a
    dummy apiKey; the real DEEPSEEK_API_KEY is injected here, server-side only."""
    import httpx
    key = os.environ.get("DEEPSEEK_API_KEY")
    if not key:
        raise HTTPException(503, "שירות ה-AI אינו מוגדר בשרת")
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    payload["model"] = "deepseek-chat"              # force a supported model
    if payload.get("tool_choice") == "required":     # DeepSeek rejects 'required'
        payload["tool_choice"] = "auto"
    try:
        async with httpx.AsyncClient(timeout=120) as client:
            r = await client.post(
                "https://api.deepseek.com/v1/chat/completions",
                json=payload,
                headers={"Authorization": f"Bearer {key}"},
            )
        return Response(content=r.content, status_code=r.status_code,
                        media_type="application/json")
    except Exception as e:
        raise HTTPException(502, "LLM upstream error")

@app.post("/api/ai/train")
def ai_train(data: dict, actor_id: int = Depends(current_user_id)):
    """Prepare a LOCAL fine-tuning dataset (JSONL) from board data.
    Data stays on-device and is used only to train the user's own local models."""
    import json as _json
    if not data.get("confirm"):
        raise HTTPException(400, "training requires explicit confirmation")
    # this bulk-exports item content to disk — system admins only
    with Session(engine) as _db:
        if not _cap(_db, actor_id, "manage_system"):
            raise HTTPException(403, "רק מנהל מערכת יכול לייצא נתוני אימון")
    board_id = data.get("board_id")
    out_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "training")
    os.makedirs(out_dir, exist_ok=True)
    with Session(engine) as db:
        q = db.query(Task).filter(Task.is_archived == False)
        if board_id:
            q = q.filter(Task.board_id == board_id)
        tasks = q.all()
        rows = []
        for t in tasks:
            s = t.status.value if hasattr(t.status, 'value') else t.status
            p = t.priority.value if hasattr(t.priority, 'value') else t.priority
            cf = t.custom_fields or {}
            rows.append({
                "instruction": "סווג משימה עירונית: קבע סטטוס, עדיפות ותגיות.",
                "input": f"{t.title}. {t.description or ''}",
                "output": _json.dumps({"status": cf.get("status_label", s),
                                       "priority": p, "tags": t.tags or []}, ensure_ascii=False),
            })
        fname = f"cityos_train_board_{board_id or 'all'}.jsonl"
        fpath = os.path.join(out_dir, fname)
        with open(fpath, "w", encoding="utf-8") as f:
            for r in rows:
                f.write(_json.dumps(r, ensure_ascii=False) + "\n")
    return {"status": "dataset_ready", "examples": len(rows), "file": fpath,
            "scope": "local-only",
            "disclaimer": "המידע משמש לאימון המודלים המקומיים שלך בלבד ואינו נשלח לצד שלישי."}

# ── BIM Bridge ───────────────────────────────────────────────────────

_IFC_CACHE = {}

@app.get("/api/bim/generate")
def bim_generate(board_id: Optional[int] = None):
    """Generate IFC model from board tasks."""
    from bim_bridge import IFCBuilder
    with Session(engine) as db:
        if board_id:
            tasks = db.query(Task).filter(Task.board_id == board_id).all()
            board_name = db.query(Board.name).filter(Board.id == board_id).scalar() or "Board"
        else:
            tasks = db.query(Task).filter(Task.is_archived == False).all()
            board_name = "All CityOS Tasks"
        
        builder = IFCBuilder(f"CityOS BIM - {board_name}")
        task_list = []
        for t in tasks:
            task_list.append({
                "id": t.id, "title": t.title, "description": t.description,
                "status": t.status.value if hasattr(t.status, 'value') else t.status,
                "priority": t.priority.value if hasattr(t.priority, 'value') else t.priority,
                "location_lat": t.location_lat, "location_lng": t.location_lng,
                "tags": t.tags or [],
            })
        
        count = builder.generate_from_tasks(task_list)
        ifc_path = f"/tmp/cityos_bim_board_{board_id or 'all'}.ifc"
        builder.save(ifc_path)
        
        # Generate BCF topics
        topics = builder.generate_bcf(task_list)
        
        _IFC_CACHE[f"board_{board_id or 'all'}"] = {
            "path": ifc_path,
            "elements": len(task_list),
            "ifc_elements": count,
            "bcf_topics": [t.to_dict() for t in topics],
        }
        
        return {
            "status": "generated",
            "elements": count,
            "bcf_topics": len(topics),
            "ifc_file": ifc_path,
        }

@app.get("/api/bim/viewer-data")
def bim_viewer_data():
    """Get IFC data as Three.js-compatible JSON."""
    from bim_bridge import IFCBuilder
    import ifcopenshell, ifcopenshell.geom
    
    # Use latest generated IFC file
    bims = [v for k, v in _IFC_CACHE.items() if os.path.exists(v.get("path", ""))]
    if not bims:
        # Generate a default one
        with Session(engine) as db:
            tasks = db.query(Task).filter(Task.is_archived == False).all()
            task_list = [{
                "id": t.id, "title": t.title,
                "location_lat": t.location_lat, "location_lng": t.location_lng,
                "tags": t.tags or [],
                "priority": t.priority.value if hasattr(t.priority, 'value') else t.priority,
                "status": t.status.value if hasattr(t.status, 'value') else t.status,
            } for t in tasks if t.location_lat]
            if task_list:
                builder = IFCBuilder("CityOS BIM Viewer")
                builder.generate_from_tasks(task_list)
                ifc_path = "/tmp/cityos_bim_viewer.ifc"
                builder.save(ifc_path)
                _IFC_CACHE["viewer"] = {"path": ifc_path}
                bims = [_IFC_CACHE["viewer"]]
    
    if bims:
        try:
            f = ifcopenshell.open(bims[0]["path"])
            objects = []
            settings = ifcopenshell.geom.settings()
            for elem in f.by_type("IfcBuildingElementProxy"):
                try:
                    shape = ifcopenshell.geom.create_shape(settings, elem)
                    verts = shape.geometry.verts
                    faces = shape.geometry.faces
                    task_id = ""
                    for rel in elem.IsDefinedBy or []:
                        if rel.is_a("IfcRelDefinesByProperties"):
                            ps = rel.RelatingPropertyDefinition
                            for p in ps.HasProperties or []:
                                if p.Name == "TaskID" and p.NominalValue:
                                    task_id = str(p.NominalValue.wrappedValue or "")
                    objects.append({
                        "guid": elem.GlobalId,
                        "name": elem.Name or "",
                        "type": elem.ObjectType or "",
                        "task_id": task_id,
                        "positions": list(verts),
                        "faces": list(faces),
                    })
                except:
                    pass
            return {"objects": objects, "count": len(objects)}
        except Exception as e:
            return {"error": str(e)}
    return {"objects": [], "count": 0}

@app.get("/api/bim/bcf-topics")
def bim_bcf_topics():
    """Get BCF topics for the current BIM model."""
    for k, v in _IFC_CACHE.items():
        if "bcf_topics" in v:
            return {"topics": v["bcf_topics"], "board": k}
    return {"topics": []}

# ── CEO Dashboard — City-Wide Command Center ─────────────────────────

@app.get("/api/ceo/dashboard")
def ceo_dashboard(user_id: int = Depends(current_user_id)):
    """Comprehensive CEO dashboard — city-wide view across boards the user may see."""
    with Session(engine) as db:
        now = datetime.now(timezone.utc).replace(tzinfo=None)

        # ── Board & Task Breakdown (membership-scoped, no leaks) ──
        visible = _visible_board_ids(db, user_id)
        boards = db.query(Board).filter(
            Board.is_archived == False, Board.id.in_(visible)
        ).all() if visible else []
        all_tasks = db.query(Task).filter(
            Task.is_archived == False, Task.board_id.in_(visible)
        ).all() if visible else []

        board_breakdown = []
        total_tasks = 0
        total_done = 0
        total_overdue = 0
        total_high_critical = 0

        for b in boards:
            tasks = [t for t in all_tasks if t.board_id == b.id]
            if not tasks:
                continue
            total_tasks += len(tasks)
            by_status = {}
            by_priority = {}
            overdue = 0
            dept_name = db.query(Department.name).filter(Department.id == b.department_id).scalar() or ""
            dept_color = db.query(Department.color).filter(Department.id == b.department_id).scalar() or "#6366f1"

            for t in tasks:
                s = t.status.value if hasattr(t.status, 'value') else t.status
                p = t.priority.value if hasattr(t.priority, 'value') else t.priority
                by_status[s] = by_status.get(s, 0) + 1
                by_priority[p] = by_priority.get(p, 0) + 1
                due = t.due_date
                if due and hasattr(due, 'tzinfo') and due.tzinfo:
                    due = due.replace(tzinfo=None)
                if due and due < now and s not in ("done", "cancelled"):
                    overdue += 1
                    total_overdue += 1
                if p in ("high", "critical", "emergency"):
                    total_high_critical += 1

            done_count = by_status.get("done", 0)
            total_done += done_count

            board_breakdown.append({
                "board_id": b.id,
                "board_name": b.name,
                "icon": b.icon or "📋",
                "department_name": dept_name,
                "department_color": dept_color,
                "total_tasks": len(tasks),
                "done": done_count,
                "overdue": overdue,
                "completion_rate": round(done_count / max(len(tasks), 1) * 100, 1),
                "status_distribution": by_status,
                "priority_distribution": by_priority,
            })

        # ── Citizen Requests Stats ──
        citizen_reqs = db.query(CitizenRequest).all()
        citizen_open = sum(1 for r in citizen_reqs if r.status in ("new", "assigned", "in_progress"))
        citizen_emergency = sum(1 for r in citizen_reqs if hasattr(r.priority, 'value') and r.priority.value in ("critical", "emergency") and r.status not in ("resolved", "closed")) or 0
        citizen_by_type = {}
        for r in citizen_reqs:
            rt = r.request_type.value if hasattr(r.request_type, 'value') else str(r.request_type)
            citizen_by_type[rt] = citizen_by_type.get(rt, 0) + 1

        # ── Permit Stats ──
        permits = db.query(Permit).all()
        permits_pending = sum(1 for p in permits if p.status in ("draft", "submitted", "in_review"))
        permits_approved = sum(1 for p in permits if p.status == "approved")
        permits_rejected = sum(1 for p in permits if p.status == "rejected")

        # ── Infrastructure Stats ──
        assets = db.query(InfrastructureAsset).all()
        assets_poor = sum(1 for a in assets if a.condition == "poor" or a.status == "maintenance_needed")
        assets_by_type = {}
        for a in assets:
            assets_by_type[a.asset_type] = assets_by_type.get(a.asset_type, 0) + 1

        # ── Transport Stats ──
        stops = db.query(PublicTransportStop).count()

        # ── City Health Score ──
        completion_rate = round(total_done / max(total_tasks, 1) * 100, 1)
        overdue_penalty = min(total_overdue * 3, 30)
        citizen_penalty = min(citizen_open * 2, 20)
        health_score = max(0, min(100, round(
            30 * (completion_rate / 100)
            + 20 * (1 - overdue_penalty / 100)
            + 20 * (1 - citizen_penalty / 100)
            + 15 * (permits_approved / max(len(permits), 1))
            + 15 * (1 - assets_poor / max(len(assets), 1))
        )))

        # Health level
        if health_score >= 80:
            health_level = "excellent"
            health_emoji = "🟢"
        elif health_score >= 60:
            health_level = "good"
            health_emoji = "🟡"
        elif health_score >= 40:
            health_level = "fair"
            health_emoji = "🟠"
        else:
            health_level = "critical"
            health_emoji = "🔴"

        return {
            "city_name": "הוד השרון",
            "timestamp": now.isoformat(),
            "health_score": health_score,
            "health_level": health_level,
            "health_emoji": health_emoji,
            "totals": {
                "departments": db.query(Department).count(),
                "boards": len(boards),
                "tasks": total_tasks,
                "done": total_done,
                "overdue": total_overdue,
                "high_priority": total_high_critical,
                "completion_rate": completion_rate,
                "users": db.query(User).count(),
                "citizen_requests": len(citizen_reqs),
                "citizen_open": citizen_open,
                "permits": len(permits),
                "permits_pending": permits_pending,
                "assets": len(assets),
                "transport_stops": stops,
            },
            "board_breakdown": board_breakdown,
            "citizen": {
                "total": len(citizen_reqs),
                "open": citizen_open,
                "emergency": citizen_emergency,
                "by_type": citizen_by_type,
            },
            "permits": {
                "total": len(permits),
                "pending": permits_pending,
                "approved": permits_approved,
                "rejected": permits_rejected,
            },
            "infrastructure": {
                "total": len(assets),
                "needs_maintenance": assets_poor,
                "by_type": assets_by_type,
            },
        }

# ── Work Plan & Budget Module – API Routes ───────────────────────────
# These routes are inserted into main.py. They implement the 5-level
# hierarchy: AnnualWorkPlan → Department → Project → Step → Task
# plus budget, approvals, KPI, Gantt, dashboards, documents, audit & AI.

# ── 1. Annual Work Plans ────────────────────────────────────────────

@app.get("/api/work-plans")
def list_work_plans():
    with Session(engine) as db:
        wps = db.query(AnnualWorkPlan).order_by(AnnualWorkPlan.year.desc()).all()
        result = []
        for wp in wps:
            depts = db.query(Department).filter(Department.organization_id == wp.organization_id).all()
            dept_count = len(depts)
            projects = db.query(Project).filter(Project.work_plan_id == wp.id).all()
            total_projects = len(projects)
            total_planned = sum(p.planned_budget or 0 for p in projects)
            total_approved = sum(p.approved_budget or 0 for p in projects)
            total_actual = sum(p.actual_budget or 0 for p in projects)
            completed = sum(1 for p in projects if p.status == ProjectStatus.COMPLETED)
            in_progress = sum(1 for p in projects if p.status == ProjectStatus.IN_PROGRESS)
            result.append({
                "id": wp.id, "name": wp.name, "year": wp.year,
                "total_budget": wp.total_budget,
                "strategic_goals": wp.strategic_goals or [],
                "municipal_kpis": wp.municipal_kpis or [],
                "overall_status": wp.overall_status,
                "departments_count": dept_count,
                "total_projects": total_projects,
                "completed_projects": completed,
                "in_progress_projects": in_progress,
                "budget_planned": total_planned,
                "budget_approved": total_approved,
                "budget_actual": total_actual,
                "budget_utilization": round(total_actual / max(total_approved, 1) * 100, 1),
                "created_at": wp.created_at.isoformat() if wp.created_at else None,
            })
        return result

@app.get("/api/work-plans/{wp_id}")
def get_work_plan(wp_id: int):
    with Session(engine) as db:
        wp = db.query(AnnualWorkPlan).filter(AnnualWorkPlan.id == wp_id).first()
        if not wp:
            raise HTTPException(404, "Work plan not found")
        depts = db.query(Department).filter(Department.organization_id == wp.organization_id).all()
        dept_breakdown = []
        for d in depts:
            projects = db.query(Project).filter(
                Project.department_id == d.id, Project.work_plan_id == wp.id
            ).all()
            dept_planned = sum(p.planned_budget or 0 for p in projects)
            dept_approved = sum(p.approved_budget or 0 for p in projects)
            dept_actual = sum(p.actual_budget or 0 for p in projects)
            avg_progress = round(sum(p.progress_percentage or 0 for p in projects) / max(len(projects), 1), 1)
            dept_breakdown.append({
                "id": d.id, "name": d.name, "code": d.code, "color": d.color,
                "manager_name": d.manager_name or "",
                "annual_budget": d.annual_budget or 0,
                "project_count": len(projects),
                "budget_planned": dept_planned,
                "budget_approved": dept_approved,
                "budget_actual": dept_actual,
                "budget_utilization": round(dept_actual / max(dept_approved, 1) * 100, 1),
                "avg_progress": avg_progress,
                "planned_projects": d.planned_projects or 0,
                "completed_projects": d.completed_projects or 0,
            })
        return {
            "id": wp.id, "name": wp.name, "year": wp.year,
            "total_budget": wp.total_budget,
            "strategic_goals": wp.strategic_goals or [],
            "municipal_kpis": wp.municipal_kpis or [],
            "overall_status": wp.overall_status,
            "created_at": wp.created_at.isoformat() if wp.created_at else None,
            "departments": dept_breakdown,
        }

@app.post("/api/work-plans")
def create_work_plan(data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        wp = AnnualWorkPlan(
            organization_id=data.get("organization_id", 1),
            name=data.get("name", "תוכנית עבודה"),
            year=data.get("year", datetime.now(timezone.utc).year),
            total_budget=data.get("total_budget", 0),
            strategic_goals=data.get("strategic_goals", []),
            municipal_kpis=data.get("municipal_kpis", []),
            overall_status=data.get("overall_status", "draft"),
        )
        db.add(wp)
        db.commit()
        db.refresh(wp)
        # Audit log
        db.add(AuditLog(entity_type="work_plan", entity_id=wp.id,
            action="create", field_name="name", new_value=wp.name,
            changed_by=actor_id))
        db.commit()
        return {"id": wp.id, "status": "created", "name": wp.name}

# ── 2. Projects ──────────────────────────────────────────────────────

@app.get("/api/projects")
def list_projects(work_plan_id: Optional[int] = None, department_id: Optional[int] = None, status: Optional[str] = None):
    with Session(engine) as db:
        q = db.query(Project)
        if work_plan_id:
            q = q.filter(Project.work_plan_id == work_plan_id)
        if department_id:
            q = q.filter(Project.department_id == department_id)
        if status:
            q = q.filter(Project.status == status)
        projects = q.order_by(Project.id).all()
        result = []
        for p in projects:
            mgr_name = ""
            if p.manager_id:
                u = db.query(User).filter(User.id == p.manager_id).first()
                if u:
                    mgr_name = u.name
            dept_name = ""
            if p.department_id:
                d = db.query(Department).filter(Department.id == p.department_id).first()
                if d:
                    dept_name = d.name
            step_count = db.query(ProjectStep).filter(ProjectStep.project_id == p.id).count()
            util = round(p.actual_budget / max(p.approved_budget, 1) * 100, 1)
            result.append({
                "id": p.id, "name": p.name, "description": p.description,
                "work_plan_id": p.work_plan_id, "department_id": p.department_id,
                "department_name": dept_name,
                "manager_id": p.manager_id, "manager_name": mgr_name,
                "planned_budget": p.planned_budget, "approved_budget": p.approved_budget,
                "actual_budget": p.actual_budget, "budget_utilization": util,
                "progress_percentage": p.progress_percentage,
                "start_date": p.start_date.isoformat() if p.start_date else None,
                "end_date": p.end_date.isoformat() if p.end_date else None,
                "status": p.status.value if hasattr(p.status, 'value') else p.status,
                "priority": p.priority.value if hasattr(p.priority, 'value') else p.priority,
                "tags": p.tags or [], "color": p.color,
                "step_count": step_count,
                "created_at": p.created_at.isoformat() if p.created_at else None,
            })
        return result

@app.get("/api/projects/{project_id}")
def get_project(project_id: int):
    with Session(engine) as db:
        p = db.query(Project).filter(Project.id == project_id).first()
        if not p:
            raise HTTPException(404, "Project not found")
        mgr_name = ""
        if p.manager_id:
            u = db.query(User).filter(User.id == p.manager_id).first()
            if u:
                mgr_name = u.name
        dept_name = ""
        if p.department_id:
            d = db.query(Department).filter(Department.id == p.department_id).first()
            if d:
                dept_name = d.name
        steps = db.query(ProjectStep).filter(ProjectStep.project_id == p.id).order_by(ProjectStep.position).all()
        budget_items = db.query(BudgetLineItem).filter(BudgetLineItem.project_id == p.id).all()
        kpis = db.query(KPI).filter(KPI.project_id == p.id).all()
        change_requests = db.query(ChangeRequest).filter(ChangeRequest.project_id == p.id).all()
        approvals = db.query(Approval).filter(
            Approval.entity_type == "project", Approval.entity_id == p.id
        ).order_by(Approval.step_order).all()
        documents = db.query(Document).filter(Document.project_id == p.id).all()
        return {
            "id": p.id, "name": p.name, "description": p.description,
            "work_plan_id": p.work_plan_id, "department_id": p.department_id,
            "department_name": dept_name,
            "manager_id": p.manager_id, "manager_name": mgr_name,
            "planned_budget": p.planned_budget, "approved_budget": p.approved_budget,
            "actual_budget": p.actual_budget,
            "budget_utilization": round(p.actual_budget / max(p.approved_budget, 1) * 100, 1),
            "progress_percentage": p.progress_percentage,
            "start_date": p.start_date.isoformat() if p.start_date else None,
            "end_date": p.end_date.isoformat() if p.end_date else None,
            "status": p.status.value if hasattr(p.status, 'value') else p.status,
            "priority": p.priority.value if hasattr(p.priority, 'value') else p.priority,
            "tags": p.tags or [], "color": p.color,
            "steps": [{
                "id": s.id, "name": s.name, "description": s.description,
                "owner_id": s.owner_id,
                "owner_name": (db.query(User.name).filter(User.id == s.owner_id).scalar() or "") if s.owner_id else "",
                "start_date": s.start_date.isoformat() if s.start_date else None,
                "end_date": s.end_date.isoformat() if s.end_date else None,
                "progress": s.progress, "position": s.position,
                "status": s.status,
                "task_count": db.query(Task).filter(Task.step_id == s.id).count(),
            } for s in steps],
            "budget_items": [{
                "id": bi.id, "item_type": bi.item_type.value if hasattr(bi.item_type, 'value') else bi.item_type,
                "name": bi.name, "planned_amount": bi.planned_amount,
                "approved_amount": bi.approved_amount, "actual_amount": bi.actual_amount,
                "notes": bi.notes,
            } for bi in budget_items],
            "kpis": [{
                "id": k.id, "name": k.name, "description": k.description,
                "target": k.target, "actual": k.actual, "unit": k.unit,
                "achievement": round(k.actual / max(k.target, 1) * 100, 1),
            } for k in kpis],
            "change_requests": [{
                "id": cr.id, "title": cr.title, "description": cr.description,
                "amount_change": cr.amount_change, "reason": cr.reason,
                "status": cr.status.value if hasattr(cr.status, 'value') else cr.status,
                "requester_name": (db.query(User.name).filter(User.id == cr.requested_by).scalar() or "") if cr.requested_by else "",
            } for cr in change_requests],
            "approvals": [{
                "id": a.id, "approver_role": a.approver_role,
                "status": a.status.value if hasattr(a.status, 'value') else a.status,
                "step_order": a.step_order,
                "approver_name": (db.query(User.name).filter(User.id == a.approver_user_id).scalar() or "") if a.approver_user_id else "",
                "notes": a.notes,
                "created_at": a.created_at.isoformat() if a.created_at else None,
                "approved_at": a.approved_at.isoformat() if a.approved_at else None,
            } for a in approvals],
            "documents": [{
                "id": d.id, "document_type": d.document_type.value if hasattr(d.document_type, 'value') else d.document_type,
                "name": d.name, "description": d.description,
                "file_url": d.file_url, "file_size": d.file_size,
                "uploader_name": (db.query(User.name).filter(User.id == d.uploaded_by).scalar() or "") if d.uploaded_by else "",
                "created_at": d.created_at.isoformat() if d.created_at else None,
            } for d in documents],
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "updated_at": p.updated_at.isoformat() if p.updated_at else None,
        }

@app.post("/api/projects")
def create_project(data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        status_val = data.get("status", "draft")
        try:
            p_status = ProjectStatus(status_val)
        except ValueError:
            p_status = ProjectStatus.DRAFT
        priority_val = data.get("priority", "medium")
        try:
            p_priority = Priority(priority_val)
        except ValueError:
            p_priority = Priority.MEDIUM
        p = Project(
            work_plan_id=data.get("work_plan_id"),
            department_id=data.get("department_id"),
            name=data.get("name", "פרויקט חדש"),
            description=data.get("description", ""),
            manager_id=data.get("manager_id"),
            planned_budget=data.get("planned_budget", 0),
            approved_budget=data.get("approved_budget", 0),
            actual_budget=data.get("actual_budget", 0),
            progress_percentage=data.get("progress_percentage", 0),
            start_date=datetime.fromisoformat(data["start_date"]) if data.get("start_date") else None,
            end_date=datetime.fromisoformat(data["end_date"]) if data.get("end_date") else None,
            status=p_status,
            priority=p_priority,
            tags=data.get("tags", []),
            color=data.get("color"),
        )
        db.add(p)
        db.commit()
        db.refresh(p)
        db.add(AuditLog(entity_type="project", entity_id=p.id,
            action="create", field_name="name", new_value=p.name,
            changed_by=actor_id))
        db.commit()
        return {"id": p.id, "status": "created", "name": p.name}

@app.patch("/api/projects/{project_id}")
def update_project(project_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        p = db.query(Project).filter(Project.id == project_id).first()
        if not p:
            raise HTTPException(404, "Project not found")
        updatable = ["name", "description", "manager_id", "planned_budget", "approved_budget",
                     "actual_budget", "progress_percentage", "start_date", "end_date", "tags", "color"]
        for field in updatable:
            if field in data:
                old_val = getattr(p, field, None)
                if field in ("start_date", "end_date") and data.get(field):
                    setattr(p, field, datetime.fromisoformat(data[field]))
                else:
                    setattr(p, field, data[field])
                if old_val != data.get(field):
                    db.add(AuditLog(entity_type="project", entity_id=p.id,
                        action="update", field_name=field,
                        old_value=str(old_val) if old_val else None,
                        new_value=str(data[field]) if data.get(field) else None,
                        changed_by=actor_id))
        if "status" in data:
            try:
                p.status = ProjectStatus(data["status"])
            except ValueError:
                pass
            db.add(AuditLog(entity_type="project", entity_id=p.id,
                action="update", field_name="status",
                new_value=data["status"],
                changed_by=actor_id))
        if "priority" in data:
            try:
                p.priority = Priority(data["priority"])
            except ValueError:
                pass
        db.commit()
        return {"id": p.id, "status": "updated"}

@app.post("/api/projects/{project_id}/recalculate")
def recalculate_project_budget(project_id: int):
    with Session(engine) as db:
        p = db.query(Project).filter(Project.id == project_id).first()
        if not p:
            raise HTTPException(404)
        items = db.query(BudgetLineItem).filter(BudgetLineItem.project_id == project_id).all()
        p.planned_budget = sum(i.planned_amount or 0 for i in items)
        p.approved_budget = sum(i.approved_amount or 0 for i in items)
        p.actual_budget = sum(i.actual_amount or 0 for i in items)
        db.commit()
        return {
            "planned_budget": p.planned_budget,
            "approved_budget": p.approved_budget,
            "actual_budget": p.actual_budget,
            "utilization": round(p.actual_budget / max(p.approved_budget, 1) * 100, 1),
        }

# ── 3. Project Steps ─────────────────────────────────────────────────

@app.get("/api/projects/{project_id}/steps")
def list_steps(project_id: int):
    with Session(engine) as db:
        steps = db.query(ProjectStep).filter(ProjectStep.project_id == project_id).order_by(ProjectStep.position).all()
        return [{
            "id": s.id, "project_id": s.project_id,
            "name": s.name, "description": s.description,
            "owner_id": s.owner_id,
            "owner_name": (db.query(User.name).filter(User.id == s.owner_id).scalar() or "") if s.owner_id else "",
            "start_date": s.start_date.isoformat() if s.start_date else None,
            "end_date": s.end_date.isoformat() if s.end_date else None,
            "progress": s.progress, "position": s.position,
            "status": s.status,
            "task_count": db.query(Task).filter(Task.step_id == s.id).count(),
        } for s in steps]

@app.post("/api/projects/{project_id}/steps")
def create_step(project_id: int, data: dict):
    with Session(engine) as db:
        step = ProjectStep(
            project_id=project_id,
            name=data.get("name", "שלב חדש"),
            description=data.get("description"),
            owner_id=data.get("owner_id"),
            start_date=datetime.fromisoformat(data["start_date"]) if data.get("start_date") else None,
            end_date=datetime.fromisoformat(data["end_date"]) if data.get("end_date") else None,
            progress=data.get("progress", 0),
            position=data.get("position", 0),
            status=data.get("status", "pending"),
        )
        db.add(step)
        db.commit()
        db.refresh(step)
        return {"id": step.id, "status": "created"}

@app.patch("/api/steps/{step_id}")
def update_step(step_id: int, data: dict):
    with Session(engine) as db:
        s = db.query(ProjectStep).filter(ProjectStep.id == step_id).first()
        if not s:
            raise HTTPException(404)
        for field in ["name", "description", "owner_id", "progress", "position", "status"]:
            if field in data:
                setattr(s, field, data[field])
        if "start_date" in data:
            s.start_date = datetime.fromisoformat(data["start_date"]) if data["start_date"] else None
        if "end_date" in data:
            s.end_date = datetime.fromisoformat(data["end_date"]) if data["end_date"] else None
        db.commit()
        return {"id": s.id, "status": "updated"}

@app.get("/api/steps/{step_id}/tasks")
def list_step_tasks(step_id: int):
    with Session(engine) as db:
        tasks = db.query(Task).filter(Task.step_id == step_id, Task.is_archived == False).all()
        return [{
            "id": t.id, "title": t.title, "description": t.description,
            "status": t.status.value if hasattr(t.status, 'value') else t.status,
            "priority": t.priority.value if hasattr(t.priority, 'value') else t.priority,
            "due_date": t.due_date.isoformat() if t.due_date else None,
            "assignees": [{"id": u.id, "name": u.name} for u in (t.assignees or [])],
            "tags": t.tags or [],
        } for t in tasks]

# ── 4. Budget Line Items ─────────────────────────────────────────────

@app.get("/api/projects/{project_id}/budget-items")
def list_budget_items(project_id: int):
    with Session(engine) as db:
        items = db.query(BudgetLineItem).filter(BudgetLineItem.project_id == project_id).all()
        totals = {"planned": 0, "approved": 0, "actual": 0}
        for i in items:
            totals["planned"] += i.planned_amount or 0
            totals["approved"] += i.approved_amount or 0
            totals["actual"] += i.actual_amount or 0
        return {
            "items": [{
                "id": bi.id, "item_type": bi.item_type.value if hasattr(bi.item_type, 'value') else bi.item_type,
                "name": bi.name, "planned_amount": bi.planned_amount,
                "approved_amount": bi.approved_amount, "actual_amount": bi.actual_amount,
                "notes": bi.notes,
            } for bi in items],
            "totals": totals,
            "utilization": round(totals["actual"] / max(totals["approved"], 1) * 100, 1),
        }

@app.post("/api/projects/{project_id}/budget-items")
def create_budget_item(project_id: int, data: dict):
    with Session(engine) as db:
        item_type_str = data.get("item_type", "other")
        try:
            item_type = BudgetItemType(item_type_str)
        except ValueError:
            item_type = BudgetItemType.OTHER
        bi = BudgetLineItem(
            project_id=project_id,
            item_type=item_type,
            name=data.get("name"),
            planned_amount=data.get("planned_amount", 0),
            approved_amount=data.get("approved_amount", 0),
            actual_amount=data.get("actual_amount", 0),
            notes=data.get("notes"),
        )
        db.add(bi)
        db.commit()
        db.refresh(bi)
        return {"id": bi.id, "status": "created"}

# ── 4b. Budget — Department & Work Plan Level ────────────────────────

@app.get("/api/budget/department/{dept_id}")
def department_budget(dept_id: int):
    with Session(engine) as db:
        dept = db.query(Department).filter(Department.id == dept_id).first()
        if not dept:
            raise HTTPException(404)
        projects = db.query(Project).filter(Project.department_id == dept_id).all()
        total_planned = sum(p.planned_budget or 0 for p in projects)
        total_approved = sum(p.approved_budget or 0 for p in projects)
        total_actual = sum(p.actual_budget or 0 for p in projects)
        return {
            "department_id": dept.id, "department_name": dept.name,
            "annual_budget": dept.annual_budget or 0,
            "total_planned": total_planned,
            "total_approved": total_approved,
            "total_actual": total_actual,
            "utilization": round(total_actual / max(total_approved, 1) * 100, 1),
            "budget_vs_annual": round(total_approved / max(dept.annual_budget or 1, 1) * 100, 1),
            "project_count": len(projects),
        }

@app.get("/api/budget/work-plan/{wp_id}")
def work_plan_budget(wp_id: int):
    with Session(engine) as db:
        wp = db.query(AnnualWorkPlan).filter(AnnualWorkPlan.id == wp_id).first()
        if not wp:
            raise HTTPException(404)
        projects = db.query(Project).filter(Project.work_plan_id == wp_id).all()
        total_planned = sum(p.planned_budget or 0 for p in projects)
        total_approved = sum(p.approved_budget or 0 for p in projects)
        total_actual = sum(p.actual_budget or 0 for p in projects)
        depts = db.query(Department).filter(Department.organization_id == wp.organization_id).all()
        dept_breakdown = []
        for d in depts:
            dp = [p for p in projects if p.department_id == d.id]
            dept_breakdown.append({
                "department_id": d.id, "department_name": d.name,
                "annual_budget": d.annual_budget or 0,
                "planned": sum(p.planned_budget or 0 for p in dp),
                "approved": sum(p.approved_budget or 0 for p in dp),
                "actual": sum(p.actual_budget or 0 for p in dp),
                "project_count": len(dp),
            })
        return {
            "work_plan_id": wp.id, "work_plan_name": wp.name,
            "total_budget": wp.total_budget,
            "total_planned": total_planned,
            "total_approved": total_approved,
            "total_actual": total_actual,
            "utilization": round(total_actual / max(total_approved, 1) * 100, 1),
            "departments": dept_breakdown,
        }

# ── 5. Approvals ─────────────────────────────────────────────────────

@app.get("/api/approvals")
def list_approvals(entity_type: Optional[str] = None, entity_id: Optional[int] = None, status: Optional[str] = None):
    with Session(engine) as db:
        q = db.query(Approval)
        if entity_type:
            q = q.filter(Approval.entity_type == entity_type)
        if entity_id:
            q = q.filter(Approval.entity_id == entity_id)
        if status:
            q = q.filter(Approval.status == status)
        approvals = q.order_by(Approval.created_at.desc()).limit(100).all()
        return [{
            "id": a.id, "entity_type": a.entity_type, "entity_id": a.entity_id,
            "approver_role": a.approver_role,
            "approver_user_id": a.approver_user_id,
            "approver_name": (db.query(User.name).filter(User.id == a.approver_user_id).scalar() or "") if a.approver_user_id else "",
            "status": a.status.value if hasattr(a.status, 'value') else a.status,
            "notes": a.notes, "step_order": a.step_order,
            "created_at": a.created_at.isoformat() if a.created_at else None,
            "approved_at": a.approved_at.isoformat() if a.approved_at else None,
        } for a in approvals]

@app.post("/api/approvals")
def create_approval_chain(data: dict):
    with Session(engine) as db:
        roles = data.get("approver_roles", [])
        created = []
        for i, role in enumerate(roles):
            existing = db.query(Approval).filter(
                Approval.entity_type == data["entity_type"],
                Approval.entity_id == data["entity_id"],
                Approval.approver_role == role
            ).first()
            if existing:
                continue
            a = Approval(
                entity_type=data["entity_type"],
                entity_id=data["entity_id"],
                approver_role=role,
                status=ApprovalStatus.PENDING,
                step_order=i,
            )
            db.add(a)
            created.append({"role": role, "step_order": i})
        db.commit()
        return {"created": created, "count": len(created)}

@app.post("/api/approvals/{approval_id}/approve")
def approve_step(approval_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        a = db.query(Approval).filter(Approval.id == approval_id).first()
        if not a:
            raise HTTPException(404)
        try:
            a.status = ApprovalStatus(data.get("status", "approved"))
        except ValueError:
            a.status = ApprovalStatus.APPROVED
        a.approver_user_id = actor_id   # the approver is whoever is signed in
        a.notes = data.get("notes", "")
        a.approved_at = datetime.now(timezone.utc)
        db.add(AuditLog(
            entity_type=f"approval.{a.entity_type}", entity_id=a.entity_id,
            action=data.get("status", "approved"),
            field_name=f"approval.{a.approver_role}",
            new_value=data.get("status", "approved"),
            changed_by=actor_id,
        ))
        db.commit()
        return {"id": a.id, "status": a.status.value, "approved_at": a.approved_at.isoformat()}

@app.get("/api/projects/{project_id}/approval-chain")
def project_approval_chain(project_id: int):
    with Session(engine) as db:
        approvals = db.query(Approval).filter(
            Approval.entity_type == "project", Approval.entity_id == project_id
        ).order_by(Approval.step_order).all()
        chain_status = "approved"
        current_step = None
        for a in approvals:
            s = a.status.value if hasattr(a.status, 'value') else a.status
            if s == "pending":
                chain_status = f"awaiting_{a.approver_role}"
                current_step = a.step_order
                break
            elif s == "rejected":
                chain_status = f"rejected_by_{a.approver_role}"
                current_step = a.step_order
                break
        return {
            "project_id": project_id,
            "chain_status": chain_status,
            "current_step": current_step,
            "approvals": [{
                "id": a.id, "approver_role": a.approver_role,
                "status": a.status.value if hasattr(a.status, 'value') else a.status,
                "step_order": a.step_order,
                "approver_name": (db.query(User.name).filter(User.id == a.approver_user_id).scalar() or "") if a.approver_user_id else "",
                "notes": a.notes,
                "created_at": a.created_at.isoformat() if a.created_at else None,
                "approved_at": a.approved_at.isoformat() if a.approved_at else None,
            } for a in approvals],
        }

# ── 6. Change Requests ──────────────────────────────────────────────

@app.get("/api/change-requests")
def list_change_requests(project_id: Optional[int] = None, status: Optional[str] = None):
    with Session(engine) as db:
        q = db.query(ChangeRequest)
        if project_id:
            q = q.filter(ChangeRequest.project_id == project_id)
        if status:
            try:
                q = q.filter(ChangeRequest.status == ChangeRequestStatus(status))
            except ValueError:
                pass
        crs = q.order_by(ChangeRequest.created_at.desc()).all()
        return [{
            "id": cr.id, "project_id": cr.project_id,
            "title": cr.title, "description": cr.description,
            "amount_change": cr.amount_change, "reason": cr.reason,
            "status": cr.status.value if hasattr(cr.status, 'value') else cr.status,
            "requester_name": (db.query(User.name).filter(User.id == cr.requested_by).scalar() or "") if cr.requested_by else "",
            "approver_name": (db.query(User.name).filter(User.id == cr.approved_by).scalar() or "") if cr.approved_by else "",
            "created_at": cr.created_at.isoformat() if cr.created_at else None,
            "approved_at": cr.approved_at.isoformat() if cr.approved_at else None,
        } for cr in crs]

@app.post("/api/change-requests")
def create_change_request(data: dict):
    with Session(engine) as db:
        try:
            cr_status = ChangeRequestStatus(data.get("status", "submitted"))
        except ValueError:
            cr_status = ChangeRequestStatus.SUBMITTED
        cr = ChangeRequest(
            project_id=data["project_id"],
            title=data.get("title", "בקשת שינוי"),
            description=data.get("description"),
            amount_change=data.get("amount_change", 0),
            reason=data.get("reason", ""),
            status=cr_status,
            requested_by=data.get("requested_by"),
        )
        db.add(cr)
        db.commit()
        db.refresh(cr)
        return {"id": cr.id, "status": "created"}

@app.patch("/api/change-requests/{cr_id}")
def update_change_request(cr_id: int, data: dict, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        cr = db.query(ChangeRequest).filter(ChangeRequest.id == cr_id).first()
        if not cr:
            raise HTTPException(404)
        if "status" in data:
            try:
                cr.status = ChangeRequestStatus(data["status"])
            except ValueError:
                pass
            if data["status"] in ("approved", "rejected"):
                cr.approved_by = actor_id
                cr.approved_at = datetime.now(timezone.utc)
        for field in ["title", "description", "amount_change", "reason"]:
            if field in data:
                setattr(cr, field, data[field])
        db.commit()
        return {"id": cr.id, "status": "updated"}

# ── 7. KPIs ─────────────────────────────────────────────────────────

@app.get("/api/kpis")
def list_kpis(project_id: Optional[int] = None, work_plan_id: Optional[int] = None):
    with Session(engine) as db:
        q = db.query(KPI)
        if project_id:
            q = q.filter(KPI.project_id == project_id)
        if work_plan_id:
            q = q.filter(KPI.work_plan_id == work_plan_id)
        kpis = q.all()
        return [{
            "id": k.id, "project_id": k.project_id, "work_plan_id": k.work_plan_id,
            "name": k.name, "description": k.description,
            "target": k.target, "actual": k.actual, "unit": k.unit,
            "achievement": round(k.actual / max(k.target, 1) * 100, 1),
            "measurement_date": k.measurement_date.isoformat() if k.measurement_date else None,
        } for k in kpis]

@app.post("/api/kpis")
def create_kpi(data: dict):
    with Session(engine) as db:
        kpi = KPI(
            project_id=data.get("project_id"),
            work_plan_id=data.get("work_plan_id"),
            name=data.get("name", "KPI"),
            description=data.get("description"),
            target=data.get("target", 100),
            actual=data.get("actual", 0),
            unit=data.get("unit", ""),
            measurement_date=datetime.fromisoformat(data["measurement_date"]) if data.get("measurement_date") else None,
        )
        db.add(kpi)
        db.commit()
        db.refresh(kpi)
        return {"id": kpi.id, "status": "created"}

@app.patch("/api/kpis/{kpi_id}")
def update_kpi(kpi_id: int, data: dict):
    with Session(engine) as db:
        k = db.query(KPI).filter(KPI.id == kpi_id).first()
        if not k:
            raise HTTPException(404)
        if "actual" in data:
            k.actual = data["actual"]
        if "target" in data:
            k.target = data["target"]
        if "measurement_date" in data:
            k.measurement_date = datetime.fromisoformat(data["measurement_date"]) if data["measurement_date"] else None
        db.commit()
        return {"id": k.id, "achievement": round(k.actual / max(k.target, 1) * 100, 1)}

# ── 8. Dependencies (for Gantt) ──────────────────────────────────────

@app.get("/api/dependencies")
def list_dependencies(project_id: Optional[int] = None):
    with Session(engine) as db:
        q = db.query(Dependency)
        if project_id:
            q = q.filter(
                (Dependency.source_type == "project" and Dependency.source_id == project_id) |
                (Dependency.target_type == "project" and Dependency.target_id == project_id)
            )
        deps = q.all()
        result = []
        for dep in deps:
            dep_type = dep.dependency_type.value if hasattr(dep.dependency_type, 'value') else dep.dependency_type
            source_name = ""
            target_name = ""
            if dep.source_type == "project" and dep.source_id:
                source_name = db.query(Project.name).filter(Project.id == dep.source_id).scalar() or ""
            if dep.target_type == "project" and dep.target_id:
                target_name = db.query(Project.name).filter(Project.id == dep.target_id).scalar() or ""
            result.append({
                "id": dep.id, "source_type": dep.source_type, "source_id": dep.source_id,
                "source_name": source_name,
                "target_type": dep.target_type, "target_id": dep.target_id,
                "target_name": target_name,
                "dependency_type": dep_type,
                "lag_days": dep.lag_days,
            })
        return result

@app.post("/api/dependencies")
def create_dependency(data: dict):
    with Session(engine) as db:
        try:
            dep_type = DependencyType(data.get("dependency_type", "finish_to_start"))
        except ValueError:
            dep_type = DependencyType.FINISH_TO_START
        dep = Dependency(
            source_type=data.get("source_type", "project"),
            source_id=data["source_id"],
            target_type=data.get("target_type", "project"),
            target_id=data["target_id"],
            dependency_type=dep_type,
            lag_days=data.get("lag_days", 0),
        )
        db.add(dep)
        db.commit()
        db.refresh(dep)
        return {"id": dep.id, "status": "created"}

@app.delete("/api/dependencies/{dep_id}")
def delete_dependency(dep_id: int):
    with Session(engine) as db:
        dep = db.query(Dependency).filter(Dependency.id == dep_id).first()
        if not dep:
            raise HTTPException(404)
        db.delete(dep)
        db.commit()
        return {"status": "deleted"}

# ── 9. Documents ─────────────────────────────────────────────────────

@app.get("/api/projects/{project_id}/documents")
def list_documents(project_id: int):
    with Session(engine) as db:
        docs = db.query(Document).filter(Document.project_id == project_id).all()
        # Group by document_type
        grouped = {}
        for d in docs:
            dt = d.document_type.value if hasattr(d.document_type, 'value') else d.document_type
            if dt not in grouped:
                grouped[dt] = []
            grouped[dt].append({
                "id": d.id, "name": d.name, "description": d.description,
                "file_url": d.file_url, "file_size": d.file_size,
                "uploader_name": (db.query(User.name).filter(User.id == d.uploaded_by).scalar() or "") if d.uploaded_by else "",
                "created_at": d.created_at.isoformat() if d.created_at else None,
            })
        return {"documents": grouped, "total": len(docs)}

@app.post("/api/projects/{project_id}/documents")
def create_document(project_id: int, data: dict):
    with Session(engine) as db:
        try:
            doc_type = DocumentType(data.get("document_type", "other"))
        except ValueError:
            doc_type = DocumentType.OTHER
        doc = Document(
            project_id=project_id,
            document_type=doc_type,
            name=data.get("name", "מסמך"),
            description=data.get("description"),
            file_url=data.get("file_url"),
            file_size=data.get("file_size"),
            uploaded_by=data.get("uploaded_by"),
        )
        db.add(doc)
        db.commit()
        db.refresh(doc)
        return {"id": doc.id, "status": "created"}

# ── 10. Audit Log ────────────────────────────────────────────────────

@app.get("/api/audit-log")
def list_audit_log(entity_type: Optional[str] = None, entity_id: Optional[int] = None, limit: int = 50):
    with Session(engine) as db:
        q = db.query(AuditLog)
        if entity_type:
            q = q.filter(AuditLog.entity_type == entity_type)
        if entity_id:
            q = q.filter(AuditLog.entity_id == entity_id)
        entries = q.order_by(AuditLog.created_at.desc()).limit(min(limit, 200)).all()
        return [{
            "id": e.id, "entity_type": e.entity_type, "entity_id": e.entity_id,
            "field_name": e.field_name, "old_value": e.old_value,
            "new_value": e.new_value, "action": e.action,
            "changed_by": e.changed_by,
            "changer_name": (db.query(User.name).filter(User.id == e.changed_by).scalar() or "") if e.changed_by else "",
            "created_at": e.created_at.isoformat() if e.created_at else None,
        } for e in entries]

# ── 11. Director Dashboard ──────────────────────────────────────────

@app.get("/api/director/dashboard/{dept_id}")
def director_dashboard(dept_id: int):
    with Session(engine) as db:
        dept = db.query(Department).filter(Department.id == dept_id).first()
        if not dept:
            raise HTTPException(404)
        projects = db.query(Project).filter(Project.department_id == dept_id).all()
        now = datetime.now(timezone.utc)
        total = len(projects)
        in_progress = sum(1 for p in projects if p.status == ProjectStatus.IN_PROGRESS)
        completed = sum(1 for p in projects if p.status == ProjectStatus.COMPLETED)
        drafting = sum(1 for p in projects if p.status == ProjectStatus.DRAFT)
        overdue = 0
        overdue_projects = []
        for p in projects:
            if p.end_date and p.status not in (ProjectStatus.COMPLETED, ProjectStatus.CANCELLED):
                ed = p.end_date
                if hasattr(ed, 'tzinfo') and ed.tzinfo:
                    ed = ed.replace(tzinfo=None)
                now_naive = now.replace(tzinfo=None)
                if ed < now_naive:
                    overdue += 1
                    overdue_projects.append({
                        "id": p.id, "name": p.name,
                        "end_date": p.end_date.isoformat() if p.end_date else None,
                        "progress": p.progress_percentage,
                        "days_overdue": (now_naive - ed).days,
                    })
        total_approved = sum(p.approved_budget or 0 for p in projects)
        total_actual = sum(p.actual_budget or 0 for p in projects)
        # Budget breakdown by item type
        all_items = db.query(BudgetLineItem).filter(
            BudgetLineItem.project_id.in_([p.id for p in projects])
        ).all() if projects else []
        budget_by_type = {}
        for bi in all_items:
            bt = bi.item_type.value if hasattr(bi.item_type, 'value') else bi.item_type
            if bt not in budget_by_type:
                budget_by_type[bt] = {"planned": 0, "approved": 0, "actual": 0}
            budget_by_type[bt]["planned"] += bi.planned_amount or 0
            budget_by_type[bt]["approved"] += bi.approved_amount or 0
            budget_by_type[bt]["actual"] += bi.actual_amount or 0
        return {
            "department_id": dept.id, "department_name": dept.name,
            "manager_name": dept.manager_name or "",
            "annual_budget": dept.annual_budget or 0,
            "projects": {
                "total": total,
                "in_progress": in_progress,
                "completed": completed,
                "draft": drafting,
                "overdue": overdue,
            },
            "budget": {
                "approved": total_approved,
                "actual": total_actual,
                "utilization": round(total_actual / max(total_approved, 1) * 100, 1),
                "budget_vs_annual": round(total_approved / max(dept.annual_budget or 1, 1) * 100, 1),
                "by_type": budget_by_type,
            },
            "overdue_projects": overdue_projects,
            "project_list": [{
                "id": p.id, "name": p.name,
                "status": p.status.value if hasattr(p.status, 'value') else p.status,
                "progress": p.progress_percentage,
                "planned_budget": p.planned_budget,
                "approved_budget": p.approved_budget,
                "actual_budget": p.actual_budget,
                "end_date": p.end_date.isoformat() if p.end_date else None,
                "priority": p.priority.value if hasattr(p.priority, 'value') else p.priority,
            } for p in projects],
        }

# ── 12. CEO Enhanced Dashboard ──────────────────────────────────────

@app.get("/api/ceo/dashboard-enhanced")
def ceo_dashboard_enhanced():
    with Session(engine) as db:
        # Get latest work plan
        wp = db.query(AnnualWorkPlan).order_by(AnnualWorkPlan.year.desc()).first()
        wp_data = None
        if wp:
            projects = db.query(Project).filter(Project.work_plan_id == wp.id).all()
            total_planned = sum(p.planned_budget or 0 for p in projects)
            total_approved = sum(p.approved_budget or 0 for p in projects)
            total_actual = sum(p.actual_budget or 0 for p in projects)
            wp_data = {
                "id": wp.id, "name": wp.name, "year": wp.year,
                "total_budget": wp.total_budget,
                "total_projects": len(projects),
                "budget_planned": total_planned,
                "budget_approved": total_approved,
                "budget_actual": total_actual,
                "utilization": round(total_actual / max(total_approved, 1) * 100, 1),
            }
        # Per-department breakdown
        depts = db.query(Department).all()
        dept_breakdown = []
        now = datetime.now(timezone.utc).replace(tzinfo=None)
        total_overdue_cross = 0
        for d in depts:
            projects = db.query(Project).filter(Project.department_id == d.id).all()
            if not projects:
                continue
            avg_progress = round(sum(p.progress_percentage or 0 for p in projects) / len(projects), 1)
            total_dept_approved = sum(p.approved_budget or 0 for p in projects)
            total_dept_actual = sum(p.actual_budget or 0 for p in projects)
            completed = sum(1 for p in projects if p.status == ProjectStatus.COMPLETED)
            in_progress = sum(1 for p in projects if p.status == ProjectStatus.IN_PROGRESS)
            overdue = 0
            for p in projects:
                if p.end_date and p.status not in (ProjectStatus.COMPLETED, ProjectStatus.CANCELLED):
                    ed = p.end_date
                    if hasattr(ed, 'tzinfo') and ed.tzinfo:
                        ed = ed.replace(tzinfo=None)
                    if ed and ed < now:
                        overdue += 1
            total_overdue_cross += overdue
            dept_breakdown.append({
                "department_id": d.id, "department_name": d.name,
                "color": d.color or "#6366f1",
                "project_count": len(projects),
                "completed": completed,
                "in_progress": in_progress,
                "overdue": overdue,
                "avg_progress": avg_progress,
                "budget_approved": total_dept_approved,
                "budget_actual": total_dept_actual,
                "budget_utilization": round(total_dept_actual / max(total_dept_approved, 1) * 100, 1),
            })
        total_all_projects = sum(d["project_count"] for d in dept_breakdown)
        total_all_completed = sum(d["completed"] for d in dept_breakdown)
        total_all_approved = sum(d["budget_approved"] for d in dept_breakdown)
        total_all_actual = sum(d["budget_actual"] for d in dept_breakdown)
        # All overdue projects list
        all_projects = db.query(Project).all()
        overdue_list = []
        for p in all_projects:
            if p.end_date and p.status not in (ProjectStatus.COMPLETED, ProjectStatus.CANCELLED):
                ed = p.end_date
                if hasattr(ed, 'tzinfo') and ed.tzinfo:
                    ed = ed.replace(tzinfo=None)
                if ed and ed < now:
                    dept_name = db.query(Department.name).filter(Department.id == p.department_id).scalar() or ""
                    mgr_name = db.query(User.name).filter(User.id == p.manager_id).scalar() or "" if p.manager_id else ""
                    overdue_list.append({
                        "project_id": p.id, "project_name": p.name,
                        "department_name": dept_name,
                        "manager_name": mgr_name,
                        "progress": p.progress_percentage,
                        "end_date": p.end_date.isoformat() if p.end_date else None,
                        "days_overdue": (now - ed).days,
                        "priority": p.priority.value if hasattr(p.priority, 'value') else p.priority,
                    })
        overdue_list.sort(key=lambda x: x["days_overdue"], reverse=True)
        return {
            "work_plan": wp_data,
            "summary": {
                "total_departments": len(dept_breakdown),
                "total_projects": total_all_projects,
                "total_completed": total_all_completed,
                "total_overdue": total_overdue_cross,
                "completion_rate": round(total_all_completed / max(total_all_projects, 1) * 100, 1),
                "budget_approved": total_all_approved,
                "budget_actual": total_all_actual,
                "budget_utilization": round(total_all_actual / max(total_all_approved, 1) * 100, 1),
            },
            "departments": dept_breakdown,
            "overdue_projects": overdue_list[:20],
        }

# ── 13. Gantt Data ──────────────────────────────────────────────────

@app.get("/api/gantt/data")
def gantt_data(work_plan_id: Optional[int] = None):
    if work_plan_id is None:
        raise HTTPException(400, "work_plan_id query parameter is required")
    with Session(engine) as db:
        wp = db.query(AnnualWorkPlan).filter(AnnualWorkPlan.id == work_plan_id).first()
        if not wp:
            raise HTTPException(404, "work plan not found")
        items = []
        # Work Plan level
        items.append({
            "id": f"wp_{wp.id}", "parent_id": None, "type": "work_plan",
            "name": wp.name, "start_date": None, "end_date": None,
            "progress": 0, "level": 0,
        })
        depts = db.query(Department).filter(Department.organization_id == wp.organization_id).all()
        for d in depts:
            projects = db.query(Project).filter(
                Project.department_id == d.id, Project.work_plan_id == wp.id
            ).all()
            if not projects:
                continue
            # Department level - use min/max dates from projects
            dept_dates = [p for p in projects if p.start_date or p.end_date]
            dept_start = min((p.start_date for p in dept_dates if p.start_date), default=None)
            dept_end = max((p.end_date for p in dept_dates if p.end_date), default=None)
            dept_progress = round(sum(p.progress_percentage or 0 for p in projects) / len(projects), 1) if projects else 0
            items.append({
                "id": f"dept_{d.id}", "parent_id": f"wp_{wp.id}", "type": "department",
                "name": d.name,
                "start_date": dept_start.isoformat() if dept_start else None,
                "end_date": dept_end.isoformat() if dept_end else None,
                "progress": dept_progress, "level": 1,
            })
            for p in projects:
                items.append({
                    "id": f"proj_{p.id}", "parent_id": f"dept_{d.id}", "type": "project",
                    "name": p.name,
                    "start_date": p.start_date.isoformat() if p.start_date else None,
                    "end_date": p.end_date.isoformat() if p.end_date else None,
                    "progress": p.progress_percentage or 0,
                    "status": p.status.value if hasattr(p.status, 'value') else p.status,
                    "level": 2, "project_id": p.id,
                })
                steps = db.query(ProjectStep).filter(ProjectStep.project_id == p.id).order_by(ProjectStep.position).all()
                for s in steps:
                    items.append({
                        "id": f"step_{s.id}",
                        "parent_id": f"proj_{p.id}",
                        "type": "step",
                        "name": s.name,
                        "start_date": s.start_date.isoformat() if s.start_date else None,
                        "end_date": s.end_date.isoformat() if s.end_date else None,
                        "progress": s.progress or 0,
                        "status": s.status,
                        "level": 3,
                    })
        # Dependencies
        deps = db.query(Dependency).all()
        dependencies = []
        for dep in deps:
            dep_type = dep.dependency_type.value if hasattr(dep.dependency_type, 'value') else dep.dependency_type
            source_icon = f"proj_{dep.source_id}" if dep.source_type == "project" else f"step_{dep.source_id}"
            target_icon = f"proj_{dep.target_id}" if dep.target_type == "project" else f"step_{dep.target_id}"
            dependencies.append({
                "id": dep.id,
                "from": source_icon,
                "to": target_icon,
                "type": dep_type,
                "lag_days": dep.lag_days,
            })
        return {"items": items, "dependencies": dependencies}

# ── 14. AI Project Insights ──────────────────────────────────────────

@app.post("/api/ai/project-insights/{project_id}")
def ai_project_insights(project_id: int, actor_id: int = Depends(current_user_id)):
    with Session(engine) as db:
        p = db.query(Project).filter(Project.id == project_id).first()
        if not p:
            raise HTTPException(404, "Project not found")
        insights = {"summary": "", "risks": [], "recommendations": [], "budget_health": "good", "schedule_health": "good"}
        mgr_name = db.query(User.name).filter(User.id == p.manager_id).scalar() or "לא הוגדר" if p.manager_id else "לא הוגדר"
        dept_name = db.query(Department.name).filter(Department.id == p.department_id).scalar() or ""
        status = p.status.value if hasattr(p.status, 'value') else p.status
        progress = p.progress_percentage or 0
        now = datetime.now(timezone.utc)
        budget_util = round(p.actual_budget / max(p.approved_budget, 1) * 100, 1)
        # Summary
        summary_parts = [
            f"פרויקט: {p.name}",
            f"אגף: {dept_name}",
            f"מנהל: {mgr_name}",
            f"סטטוס: {status}",
            f"התקדמות: {progress}%",
        ]
        # Schedule analysis
        if p.end_date and p.status not in (ProjectStatus.COMPLETED, ProjectStatus.CANCELLED):
            end = p.end_date
            if hasattr(end, 'tzinfo') and end.tzinfo:
                end = end.replace(tzinfo=None)
            now_naive = now.replace(tzinfo=None)
            if end < now_naive:
                days_overdue = (now_naive - end).days
                insights["schedule_health"] = "critical"
                insights["risks"].append(f"⚠️ הפרויקט באיחור של {days_overdue} ימים")
                insights["recommendations"].append("נדרשת עדכון לוח זמנים ואישור מנהל אגף")
                summary_parts.append(f"איחור: {days_overdue} ימים")
            elif progress < 90 and (end - now_naive).days < 30:
                insights["schedule_health"] = "warning"
                insights["risks"].append(f"⚠️ צפוי איחור - נותרו {(end - now_naive).days} ימים בלבד")
                insights["recommendations"].append("האץ את הקצב או עדכן את לוח הזמנים")
                summary_parts.append(f"נותרו {(end - now_naive).days} ימים")
            else:
                days_left = (end - now_naive).days
                summary_parts.append(f"נראה בלוח זמנים, נותרו {max(days_left, 0)} ימים")
        else:
            summary_parts.append("לוח זמנים: הושלם או לא מוגדר")
        # Budget analysis
        if budget_util > 100:
            insights["budget_health"] = "critical"
            insights["risks"].append(f"💰 חריגת תקציב: מנוצל {budget_util}% (התקציב המאושר {p.approved_budget:,} ₪)")
            insights["recommendations"].append("דרוש אישור גזבר לחריגת תקציב")
            summary_parts.append(f"חריגת תקציב: {budget_util}%")
        elif budget_util > 85:
            insights["budget_health"] = "warning"
            insights["risks"].append(f"💰 ניצול תקציב גבוה: {budget_util}%")
            insights["recommendations"].append("מומלץ לבצע עדכון תקציבי בהקדם")
            summary_parts.append(f"ניצול תקציב: {budget_util}%")
        else:
            summary_parts.append(f"ניצול תקציב: {budget_util}%")
        # Line item analysis
        items = db.query(BudgetLineItem).filter(BudgetLineItem.project_id == project_id).all()
        for bi in items:
            item_util = round(bi.actual_amount / max(bi.approved_amount, 1) * 100, 1)
            if item_util > 100:
                item_name = bi.name or (bi.item_type.value if hasattr(bi.item_type, 'value') else bi.item_type)
                insights["risks"].append(f"💰 סעיף '{item_name}' חורג: {item_util}% ניצול")
                insights["recommendations"].append(f"בדוק את סעיף '{item_name}' - נדרשת אישור לחריגה")
        # Step analysis
        steps = db.query(ProjectStep).filter(ProjectStep.project_id == project_id).order_by(ProjectStep.position).all()
        for s in steps:
            if s.status == "pending" and s.position > 0:
                prev = db.query(ProjectStep).filter(
                    ProjectStep.project_id == project_id,
                    ProjectStep.position == s.position - 1
                ).first()
                if prev and prev.status != "completed" and prev.end_date and prev.end_date < now.replace(tzinfo=None):
                    insights["risks"].append(f"שלב '{s.name}' ממתין לשלב קודם שלא הושלם")
        if not insights["risks"]:
            insights["summary"] = "✅ " + " · ".join(summary_parts)
        else:
            insights["summary"] = " · ".join(summary_parts)
        if not insights["recommendations"] and insights["schedule_health"] == "good" and insights["budget_health"] == "good":
            insights["recommendations"].append("הפרויקט במצב תקין, אין צורך בפעולה מיידית")
        return insights
# ── Main ─────────────────────────────────────────────────────────────

# Serve frontend
frontend_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend")
if os.path.exists(frontend_path):
    app.mount("/", StaticFiles(directory=frontend_path, html=True), name="frontend")

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    print(f"🚀 CityOS running on http://localhost:{port}")
    uvicorn.run(app, host="0.0.0.0", port=port)
